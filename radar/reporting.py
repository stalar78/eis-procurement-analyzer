from __future__ import annotations

import json
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any
import shutil

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from radar import radar_version
from radar.historical import HistoricalAssessmentBundle
from radar.models import ArtifactRecord, DeepAssessment, NoCompetitionOpportunity, RadarAssessment, RadarCard, RadarDecision


REPORT_COLUMNS = [
    "rank",
    "procurement_number",
    "title",
    "customer",
    "law",
    "procedure_type",
    "nmck",
    "published_at",
    "application_deadline",
    "days_to_deadline",
    "search_profiles",
    "score",
    "radar_decision",
    "positive_reasons",
    "negative_reasons",
    "hard_reject_reasons",
    "manual_review_questions",
    "is_new",
    "is_changed",
    "source_url",
]


def _joined(values: list[str]) -> str:
    return "; ".join(values)


def _row(rank: int, card: RadarCard, assessment: RadarAssessment) -> list[Any]:
    return [
        rank,
        card.procurement_number,
        card.title,
        card.customer,
        card.law,
        card.procedure_type,
        card.nmck,
        card.published_at,
        card.application_deadline,
        assessment.days_to_deadline,
        _joined(card.search_profiles),
        assessment.total_score,
        assessment.radar_decision.value,
        _joined(assessment.positive_reasons),
        _joined(assessment.negative_reasons),
        _joined(assessment.hard_reject_reasons),
        _joined(assessment.manual_review_questions),
        assessment.is_new,
        assessment.is_changed,
        card.source_url,
    ]


def _autosize(sheet) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        width = min(70, max(12, max(len(str(cell.value or "")) for cell in column) + 2))
        sheet.column_dimensions[letter].width = width


def _write_table(sheet, rows: list[list[Any]], headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        sheet.append(row)
    _autosize(sheet)


def _excel_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def build_summary(
    run_id: str,
    started_at: datetime,
    finished_at: datetime,
    as_of: datetime,
    profiles: list[str],
    diagnostics: dict[str, Any],
    assessments: list[RadarAssessment],
    deep_assessments: list[DeepAssessment] | None = None,
    historical_bundles: list[HistoricalAssessmentBundle] | None = None,
    opportunities: list[NoCompetitionOpportunity] | None = None,
) -> dict[str, Any]:
    eligibility = Counter(item.eligibility_status.value for item in assessments)
    decisions = Counter(item.radar_decision.value for item in assessments)
    deep_assessments = deep_assessments or []
    historical_bundles = historical_bundles or []
    opportunities = opportunities or []
    deep_decisions = Counter(item.final_radar_decision.value for item in deep_assessments)
    transitions = Counter(f"{item.preliminary_decision.value} -> {item.final_radar_decision.value}" for item in deep_assessments)
    return {
        "radar_version": radar_version,
        "run_id": run_id,
        "run_started": started_at.isoformat(timespec="seconds"),
        "run_finished": finished_at.isoformat(timespec="seconds"),
        "as_of": as_of.isoformat(timespec="seconds"),
        "profiles": profiles,
        "queries_attempted": diagnostics.get("queries_attempted", 0),
        "queries_successful": diagnostics.get("queries_successful", 0),
        "raw_cards": diagnostics.get("raw_cards", 0),
        "unique_cards": diagnostics.get("unique_cards", 0),
        "open": eligibility.get("OPEN", 0),
        "deadline_too_close": eligibility.get("DEADLINE_TOO_CLOSE", 0),
        "closed": eligibility.get("CLOSED", 0),
        "deadline_unknown": eligibility.get("DEADLINE_UNKNOWN", 0),
        "priority": decisions.get("PRIORITY", 0),
        "review": decisions.get("REVIEW", 0),
        "watch": decisions.get("WATCH", 0),
        "reject": decisions.get("REJECT", 0),
        "insufficient": decisions.get("INSUFFICIENT_DATA", 0),
        "new": sum(1 for item in assessments if item.is_new),
        "changed": sum(1 for item in assessments if item.is_changed),
        "errors": diagnostics.get("errors", []),
        "candidates_selected_for_enrichment": diagnostics.get("candidates_selected_for_enrichment", 0),
        "enrichments_complete": diagnostics.get("enrichments_complete", 0),
        "enrichments_partial": diagnostics.get("enrichments_partial", 0),
        "enrichments_failed_retryable": diagnostics.get("enrichments_failed_retryable", 0),
        "enrichments_failed_final": diagnostics.get("enrichments_failed_final", 0),
        "cached_enrichments_reused": diagnostics.get("cached_enrichments_reused", 0),
        "documents_downloaded": diagnostics.get("documents_downloaded", 0),
        "total_downloaded_bytes": diagnostics.get("total_downloaded_bytes", 0),
        "deep_priority": deep_decisions.get("PRIORITY", 0),
        "deep_review": deep_decisions.get("REVIEW", 0),
        "deep_watch": deep_decisions.get("WATCH", 0),
        "deep_reject": deep_decisions.get("REJECT", 0),
        "deep_insufficient": deep_decisions.get("INSUFFICIENT_DATA", 0),
        "historical_candidates": len(historical_bundles),
        "historical_selected_analogs": sum(len(bundle.historical_analogs) for bundle in historical_bundles),
        "historical_low_risk": sum(1 for bundle in historical_bundles if bundle.dumping_risk_assessment.risk_level == "LOW"),
        "historical_moderate_risk": sum(1 for bundle in historical_bundles if bundle.dumping_risk_assessment.risk_level == "MODERATE"),
        "historical_high_risk": sum(1 for bundle in historical_bundles if bundle.dumping_risk_assessment.risk_level == "HIGH"),
        "historical_extreme_risk": sum(1 for bundle in historical_bundles if bundle.dumping_risk_assessment.risk_level == "EXTREME"),
        "preliminary_to_final_decision_changes": sum(1 for item in deep_assessments if item.preliminary_decision != item.final_radar_decision),
        "opportunities_found": len(opportunities),
        "decision_transitions": dict(transitions),
        "discovery_mode": diagnostics.get("discovery_mode", ""),
        "search_window": diagnostics.get("search_window", {}),
        "status_filter_requested": diagnostics.get("status_filter_requested", []),
        "cards_returned_by_search": diagnostics.get("raw_cards", 0),
        "cards_with_active_raw_status": diagnostics.get("cards_with_active_raw_status", 0),
        "cards_with_future_deadline": diagnostics.get("cards_with_future_deadline", 0),
        "provisionally_open": diagnostics.get("provisionally_open", 0),
        "detail_verifications_attempted": diagnostics.get("detail_verifications_attempted", 0),
        "verified_open": diagnostics.get("verified_open", 0),
        "verified_closed": diagnostics.get("verified_closed", 0),
        "verified_cancelled": diagnostics.get("verified_cancelled", 0),
        "status_conflicts": diagnostics.get("status_conflicts", 0),
        "deadline_conflicts": diagnostics.get("deadline_conflicts", 0),
        "detail_unavailable": diagnostics.get("detail_unavailable", 0),
        "no_open_candidate_reason": diagnostics.get("no_open_candidate_reason", ""),
    }


def write_reports(
    output_dir: str | Path,
    run_id: str,
    started_at: datetime,
    finished_at: datetime,
    as_of: datetime,
    profiles: list[str],
    diagnostics: dict[str, Any],
    cards: list[RadarCard],
    assessments: list[RadarAssessment],
    historical_bundles: list[HistoricalAssessmentBundle] | None = None,
    deep_assessments: list[DeepAssessment] | None = None,
    opportunities: list[NoCompetitionOpportunity] | None = None,
    artifacts: list[ArtifactRecord] | None = None,
    enrichment_plan: dict[str, Any] | None = None,
    dry_run: bool = False,
    publish_blocked: bool = False,
) -> dict[str, Path]:
    base = Path(output_dir)
    run_quality = diagnostics.get("run_quality_status") or diagnostics.get("historical_live_validation", {}).get("diagnostics", {}).get("run_quality_status", "")
    publishable = dry_run or run_quality in {"", "SUCCESS", "PARTIAL_SUCCESS"} or (publish_blocked and run_quality == "BLOCKED_EXTERNAL")
    attempt = base / "preview" / run_id if dry_run else base / "runs" / f"{run_id}.tmp"
    archive = base / "runs" / run_id
    target = attempt
    target.mkdir(parents=True, exist_ok=True)

    historical_bundles = historical_bundles or []
    opportunities = opportunities or []
    assessment_by_number = {item.procurement_number: item for item in assessments}
    deep_by_number = {item.procurement_number: item for item in (deep_assessments or [])}
    historical_by_number = {item.procurement_number: item for item in historical_bundles}
    pairs = [(card, assessment_by_number[card.procurement_number]) for card in cards if card.procurement_number in assessment_by_number]
    pairs.sort(key=lambda item: item[1].total_score, reverse=True)
    opportunity_by_number: dict[str, list[NoCompetitionOpportunity]] = {}
    for opportunity in opportunities:
        opportunity_by_number.setdefault(opportunity.current_procurement_number, []).append(opportunity)
    summary = build_summary(run_id, started_at, finished_at, as_of, profiles, diagnostics, assessments, deep_assessments, historical_bundles, opportunities)

    json_payload = {
        "summary": summary,
        "items": [
            {
                "card": card.to_dict(),
                "assessment": assessment.to_dict(),
                "preliminary_assessment": assessment.to_dict(),
                "historical_search": [item.to_dict() for item in historical_by_number[card.procurement_number].historical_search] if card.procurement_number in historical_by_number else [],
                "historical_analogs": [item.to_dict() for item in historical_by_number[card.procurement_number].historical_analogs] if card.procurement_number in historical_by_number else [],
                "competition_metrics": historical_by_number[card.procurement_number].competition_metrics.to_dict() if card.procurement_number in historical_by_number else None,
                "customer_history": historical_by_number[card.procurement_number].customer_history.to_dict() if card.procurement_number in historical_by_number and historical_by_number[card.procurement_number].customer_history else None,
                "supplier_history": [item.to_dict() for item in historical_by_number[card.procurement_number].supplier_history] if card.procurement_number in historical_by_number else [],
                "dumping_risk_assessment": historical_by_number[card.procurement_number].dumping_risk_assessment.to_dict() if card.procurement_number in historical_by_number else None,
                "history_adjusted_assessment": historical_by_number[card.procurement_number].history_adjusted_assessment.to_dict() if card.procurement_number in historical_by_number and historical_by_number[card.procurement_number].history_adjusted_assessment else None,
                "repeated_procurements": [item.to_dict() for item in historical_by_number[card.procurement_number].repeated_procurements] if card.procurement_number in historical_by_number else [],
                "failure_republication_assessment": [item.to_dict() for item in opportunity_by_number.get(card.procurement_number, [])],
                "enrichment_status": deep_by_number[card.procurement_number].enrichment_status.value if card.procurement_number in deep_by_number else "NOT_SELECTED",
                "deep_assessment": deep_by_number[card.procurement_number].to_dict() if card.procurement_number in deep_by_number else None,
                "final_assessment": deep_by_number[card.procurement_number].to_dict() if card.procurement_number in deep_by_number else assessment.to_dict(),
            }
            for card, assessment in pairs
        ],
    }
    json_path = target / "latest.json"
    json_path.write_text(json.dumps(json_payload, ensure_ascii=False, indent=2), encoding="utf-8")

    md_path = target / "latest.md"
    md_path.write_text(render_markdown(summary, pairs, historical_bundles, deep_assessments or [], opportunities), encoding="utf-8")

    xlsx_path = target / "latest.xlsx"
    write_xlsx(xlsx_path, summary, pairs, diagnostics, historical_bundles, deep_assessments or [], opportunities, artifacts or [])

    if enrichment_plan is not None:
        plan_path = target / "enrichment_plan.json"
        plan_path.write_text(json.dumps(enrichment_plan, ensure_ascii=False, indent=2), encoding="utf-8")

    if diagnostics.get("search_diagnostics") is not None:
        search_json = target / "search_diagnostics.json"
        search_json.write_text(json.dumps(diagnostics.get("search_diagnostics", []), ensure_ascii=False, indent=2), encoding="utf-8")
        write_simple_csv(target / "search_diagnostics.csv", diagnostics.get("search_diagnostics", []))
    if diagnostics.get("status_audit") is not None:
        write_simple_csv(target / "status_audit.csv", diagnostics.get("status_audit", []))
    if diagnostics.get("open_verifications") is not None:
        open_json = target / "open_verifications.json"
        open_json.write_text(json.dumps(diagnostics.get("open_verifications", []), ensure_ascii=False, indent=2), encoding="utf-8")
        write_simple_csv(target / "open_verifications.csv", diagnostics.get("open_verifications", []))
    historical_diag_rows = [{key: value} for key, value in diagnostics.items() if key.startswith("historical_")]
    if historical_diag_rows:
        historical_diag_json = target / "historical_diagnostics.json"
        historical_diag_json.write_text(json.dumps(historical_diag_rows, ensure_ascii=False, indent=2), encoding="utf-8")
        write_simple_csv(target / "historical_diagnostics.csv", historical_diag_rows)
    live_validation = diagnostics.get("historical_live_validation")
    if live_validation:
        for name, payload in [
            ("historical_query_plan.json", live_validation.get("historical_query_plan", [])),
            ("source_validation.json", live_validation.get("source_validation", {})),
            ("historical_candidates_raw.json", live_validation.get("raw_candidates", [])),
            ("historical_candidates_unique.json", live_validation.get("unique_candidates", [])),
            ("historical_candidates_scored.json", live_validation.get("scored_candidates", [])),
            ("selected_analogs.json", live_validation.get("bundle", {}).get("historical_analogs", [])),
            ("competition_metric_evidence.json", live_validation.get("competition_metric_evidence", {})),
            ("historical_live_diagnostics.json", live_validation.get("diagnostics", {})),
        ]:
            (target / name).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        write_simple_csv(target / "analog_review.csv", live_validation.get("analog_review", []))
        write_simple_csv(target / "historical_live_diagnostics.csv", [live_validation.get("diagnostics", {})])

    manifest = {
        "run_id": run_id,
        "run_quality_status": run_quality,
        "publishable": publishable,
        "latest_published": bool(publishable and not dry_run),
        "latest_publish_reason": "publishable run quality" if publishable else "blocked/failed run does not replace latest",
    }
    (target / "run_manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    if not dry_run:
        if archive.exists():
            suffix = 1
            while (base / "runs" / f"{run_id}_{suffix}").exists():
                suffix += 1
            archive = base / "runs" / f"{run_id}_{suffix}"
        target.replace(archive)
        attempt_json = base / "latest_attempt.json"
        attempt_payload = json.loads((archive / "latest.json").read_text(encoding="utf-8"))
        attempt_payload["run_manifest"] = manifest
        attempt_json.write_text(json.dumps(attempt_payload, ensure_ascii=False, indent=2), encoding="utf-8")
        if publishable:
            for source_name, latest_name in [("latest.json", "latest.json"), ("latest.md", "latest.md"), ("latest.xlsx", "latest.xlsx")]:
                shutil.copy2(archive / source_name, base / latest_name)
            for sidecar in archive.iterdir():
                if sidecar.name.startswith("latest.") or sidecar.name == "run_manifest.json":
                    continue
                if sidecar.is_file():
                    shutil.copy2(sidecar, base / sidecar.name)
            diagnostics["latest_published"] = True
            diagnostics["latest_publish_reason"] = manifest["latest_publish_reason"]
            return {"json": base / "latest.json", "markdown": base / "latest.md", "xlsx": base / "latest.xlsx", "run_dir": archive, "latest_attempt": attempt_json}
        failed_root = base / "runs_failed"
        failed_root.mkdir(parents=True, exist_ok=True)
        failed_target = failed_root / archive.name
        if failed_target.exists():
            shutil.rmtree(failed_target)
        shutil.move(str(archive), str(failed_target))
        diagnostics["latest_published"] = False
        diagnostics["latest_publish_reason"] = manifest["latest_publish_reason"]
        return {"json": failed_target / "latest.json", "markdown": failed_target / "latest.md", "xlsx": failed_target / "latest.xlsx", "run_dir": failed_target, "latest_attempt": attempt_json}

    return {"json": json_path, "markdown": md_path, "xlsx": xlsx_path}


def write_xlsx(
    path: Path,
    summary: dict[str, Any],
    pairs: list[tuple[RadarCard, RadarAssessment]],
    diagnostics: dict[str, Any],
    historical_bundles: list[HistoricalAssessmentBundle],
    deep_assessments: list[DeepAssessment],
    opportunities: list[NoCompetitionOpportunity],
    artifacts: list[ArtifactRecord],
) -> None:
    wb = Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    _write_table(summary_sheet, [[key, _excel_value(value)] for key, value in summary.items()], ["metric", "value"])
    deep_by_number = {item.procurement_number: item for item in deep_assessments}
    for title, decision in [
        ("Final Priority", RadarDecision.PRIORITY),
        ("Final Review", RadarDecision.REVIEW),
        ("Final Watch", RadarDecision.WATCH),
        ("Final Rejected", RadarDecision.REJECT),
    ]:
        sheet = wb.create_sheet(title)
        rows = [
            final_row(index + 1, card, assessment, deep_by_number.get(card.procurement_number))
            for index, (card, assessment) in enumerate(pairs)
            if (deep_by_number.get(card.procurement_number).final_radar_decision if deep_by_number.get(card.procurement_number) else assessment.radar_decision) == decision
        ]
        _write_table(sheet, rows, FINAL_COLUMNS)

    preliminary_sheet = wb.create_sheet("Preliminary assessment")
    _write_table(preliminary_sheet, [_row(index + 1, card, assessment) for index, (card, assessment) in enumerate(pairs)], REPORT_COLUMNS)

    deep_sheet = wb.create_sheet("Deep assessment")
    _write_table(deep_sheet, [deep_row(item) for item in deep_assessments], DEEP_COLUMNS)

    status_sheet = wb.create_sheet("Enrichment status")
    _write_table(status_sheet, [[item.procurement_number, item.enrichment_status.value, item.error_code, item.error_message] for item in deep_assessments], ["procurement_number", "status", "error_code", "error_message"])

    artifact_sheet = wb.create_sheet("Downloaded artifacts")
    _write_table(artifact_sheet, [[a.procurement_number, a.artifact_type, a.original_filename, a.size_bytes, a.sha256, a.document_type, a.local_path] for a in artifacts], ["procurement_number", "artifact_type", "original_filename", "size_bytes", "sha256", "document_type", "local_path"])

    completeness_sheet = wb.create_sheet("Document completeness")
    _write_table(completeness_sheet, [[d.procurement_number, d.technical_specification_status, d.contract_status, d.application_requirements_status, d.nmck_status, d.protocol_status, d.document_completeness_score, d.document_reliability] for d in deep_assessments], ["procurement_number", "technical_specification", "contract", "application_requirements", "nmck", "protocol", "completeness_score", "reliability"])

    evidence_sheet = wb.create_sheet("Evidence summary")
    _write_table(evidence_sheet, [[d.procurement_number, d.evidence_count, d.high_confidence_evidence_count, _joined(d.key_positive_factors), _joined(d.key_risks), _joined(d.blocking_factors)] for d in deep_assessments], ["procurement_number", "evidence_count", "high_confidence", "positive", "risks", "blockers"])

    historical_summary = wb.create_sheet("Historical Summary")
    _write_table(
        historical_summary,
        [
            [
                bundle.procurement_number,
                bundle.competition_metrics.analog_count_total,
                bundle.competition_metrics.strong_analog_count,
                bundle.competition_metrics.median_participants,
                bundle.competition_metrics.median_reduction_percent,
                bundle.dumping_risk_assessment.risk_level,
                bundle.dumping_risk_assessment.risk_score,
                bundle.history_adjusted_assessment.history_adjusted_score if bundle.history_adjusted_assessment else "",
            ]
            for bundle in historical_bundles
        ],
        ["procurement_number", "analog_count", "strong_analog_count", "median_participants", "median_reduction_percent", "risk_level", "risk_score", "history_adjusted_score"],
    )

    historical_analogs = wb.create_sheet("Historical Analogs")
    _write_table(
        historical_analogs,
        [
            [
                bundle.procurement_number,
                analog.analog_procurement_number,
                analog.title,
                analog.customer,
                analog.similarity_score,
                analog.participant_count,
                analog.reduction_percent,
                analog.result_data_status,
            ]
            for bundle in historical_bundles
            for analog in bundle.historical_analogs
        ],
        ["source_procurement_number", "analog_procurement_number", "title", "customer", "similarity_score", "participant_count", "reduction_percent", "result_data_status"],
    )

    competition_sheet = wb.create_sheet("Competition Metrics")
    _write_table(
        competition_sheet,
        [
            [
                bundle.procurement_number,
                bundle.competition_metrics.analog_count_total,
                bundle.competition_metrics.analog_count_with_complete_results,
                bundle.competition_metrics.participant_sample_size,
                bundle.competition_metrics.reduction_sample_size,
                bundle.competition_metrics.winner_sample_size,
                bundle.competition_metrics.median_participants,
                bundle.competition_metrics.median_reduction_percent,
                bundle.competition_metrics.extreme_reduction_rate,
                bundle.competition_metrics.no_application_rate,
                bundle.competition_metrics.repeated_winner_share,
            ]
            for bundle in historical_bundles
        ],
        ["procurement_number", "analog_count_total", "complete_results", "participant_sample_size", "reduction_sample_size", "winner_sample_size", "median_participants", "median_reduction_percent", "extreme_reduction_rate", "no_application_rate", "repeated_winner_share"],
    )

    dumping_sheet = wb.create_sheet("Dumping Risk")
    _write_table(
        dumping_sheet,
        [
            [
                bundle.procurement_number,
                bundle.dumping_risk_assessment.risk_level,
                bundle.dumping_risk_assessment.risk_score,
                bundle.dumping_risk_assessment.confidence,
                bundle.dumping_risk_assessment.participant_metric_confidence,
                bundle.dumping_risk_assessment.reduction_metric_confidence,
                bundle.dumping_risk_assessment.winner_metric_confidence,
                _joined(bundle.dumping_risk_assessment.positive_signals),
                _joined(bundle.dumping_risk_assessment.negative_signals),
            ]
            for bundle in historical_bundles
        ],
        ["procurement_number", "risk_level", "risk_score", "confidence", "participant_metric_confidence", "reduction_metric_confidence", "winner_metric_confidence", "positive_signals", "negative_signals"],
    )

    if diagnostics.get("analog_result_resolution"):
        resolution_sheet = wb.create_sheet("Analog Result Resolution")
        _write_table(
            resolution_sheet,
            [
                [
                    item.get("procurement_number", ""),
                    item.get("law", ""),
                    item.get("resolution_strategy", ""),
                    item.get("resolution_status", ""),
                    item.get("result_url", ""),
                    item.get("protocol_url", ""),
                    item.get("contract_url", ""),
                    item.get("final_resolved_url", ""),
                    item.get("result_source_type", ""),
                ]
                for item in diagnostics.get("analog_result_resolution", [])
            ],
            ["procurement_number", "law", "resolution_strategy", "resolution_status", "result_url", "protocol_url", "contract_url", "final_resolved_url", "result_source_type"],
        )
    if diagnostics.get("protocol_extraction_diagnostics"):
        protocol_sheet = wb.create_sheet("Protocol Extraction Diagnostics")
        _write_table(
            protocol_sheet,
            [
                [
                    item.get("procurement_number", ""),
                    item.get("document_url", ""),
                    item.get("document_type", ""),
                    item.get("classification_score", ""),
                    item.get("parser_used", ""),
                    item.get("tables_found", ""),
                    item.get("rows_inspected", ""),
                    _excel_value(item.get("extracted_fields", {})),
                    _excel_value(item.get("parser_warnings", [])),
                ]
                for item in diagnostics.get("protocol_extraction_diagnostics", [])
            ],
            ["procurement_number", "document_url", "document_type", "classification_score", "parser_used", "tables_found", "rows_inspected", "extracted_fields", "parser_warnings"],
        )
    if diagnostics.get("assembled_historical_results"):
        assembled_sheet = wb.create_sheet("Assembled Historical Results")
        _write_table(
            assembled_sheet,
            [
                [
                    item.get("procurement_number", ""),
                    item.get("completeness", ""),
                    item.get("confidence", ""),
                    item.get("nmck", ""),
                    item.get("final_price", ""),
                    item.get("participant_count", ""),
                    item.get("admitted_participant_count", ""),
                    item.get("winner_name", ""),
                    item.get("reduction_percent", ""),
                    _excel_value(item.get("conflicts", [])),
                    _excel_value(item.get("warnings", [])),
                ]
                for item in diagnostics.get("assembled_historical_results", [])
            ],
            ["procurement_number", "completeness", "confidence", "nmck", "final_price", "participant_count", "admitted_participant_count", "winner_name", "reduction_percent", "conflicts", "warnings"],
        )
    if diagnostics.get("competition_metric_samples"):
        sample_sheet = wb.create_sheet("Competition Metric Samples")
        sample_rows = [[key, _excel_value(value)] for key, value in diagnostics.get("competition_metric_samples", {}).items()]
        _write_table(sample_sheet, sample_rows, ["metric", "value"])

    customer_sheet = wb.create_sheet("Customer History")
    _write_table(
        customer_sheet,
        [
            [bundle.procurement_number, bundle.customer_history.normalized_customer_name if bundle.customer_history else "", bundle.customer_history.total_completed_procurements if bundle.customer_history else 0, bundle.customer_history.median_reduction_percent if bundle.customer_history else "", bundle.customer_history.history_confidence if bundle.customer_history else ""]
            for bundle in historical_bundles
        ],
        ["procurement_number", "customer", "completed", "median_reduction_percent", "history_confidence"],
    )

    supplier_sheet = wb.create_sheet("Supplier History")
    _write_table(
        supplier_sheet,
        [
            [bundle.procurement_number, supplier.supplier_name, supplier.known_wins, supplier.median_reduction_percent, supplier.confidence]
            for bundle in historical_bundles
            for supplier in bundle.supplier_history
        ],
        ["procurement_number", "supplier_name", "known_wins", "median_reduction_percent", "confidence"],
    )

    repeated_sheet = wb.create_sheet("Repeated Procurements")
    _write_table(
        repeated_sheet,
        [
            [bundle.procurement_number, link.previous_procurement_number, link.similarity_score, link.relation_type, link.confidence]
            for bundle in historical_bundles
            for link in bundle.repeated_procurements
        ],
        ["current_procurement_number", "previous_procurement_number", "similarity_score", "relation_type", "confidence"],
    )

    no_participant_sheet = wb.create_sheet("No Participant Opportunities")
    _write_table(
        no_participant_sheet,
        [
            [bundle.procurement_number, bundle.dumping_risk_assessment.no_application_opportunity, bundle.competition_metrics.no_application_rate, bundle.dumping_risk_assessment.risk_level]
            for bundle in historical_bundles
        ],
        ["procurement_number", "no_application_opportunity", "no_application_rate", "risk_level"],
    )

    if opportunities:
        opportunity_sheet = wb.create_sheet("No Competition Opportunities")
        _write_table(
            opportunity_sheet,
            [
                [
                    item.current_procurement_number,
                    item.previous_procurement_number,
                    item.current_customer,
                    item.current_title,
                    item.current_nmck,
                    item.previous_nmck,
                    item.current_deadline,
                    item.previous_failure_type,
                    item.previous_application_count,
                    item.republication_confidence,
                    item.opportunity_score,
                    item.opportunity_level,
                    _joined(item.positive_signals),
                    _joined(item.risks),
                    _joined(item.warnings),
                ]
                for item in opportunities
            ],
            ["current_procurement_number", "previous_procurement_number", "customer", "title", "current_nmck", "previous_nmck", "current_deadline", "previous_failure_type", "previous_application_count", "republication_confidence", "opportunity_score", "opportunity_level", "positive_signals", "risks", "warnings"],
        )

    historical_diag_sheet = wb.create_sheet("Historical Diagnostics")
    _write_table(
        historical_diag_sheet,
        [[key, json.dumps(value, ensure_ascii=False)] for key, value in diagnostics.items() if key.startswith("historical")],
        ["key", "value"],
    )

    live_validation = diagnostics.get("historical_live_validation") or {}
    live_summary = live_validation.get("diagnostics", {}) if isinstance(live_validation, dict) else {}
    live_sheet = wb.create_sheet("Live Historical Validation")
    _write_table(live_sheet, [[key, _excel_value(value)] for key, value in live_summary.items()], ["metric", "value"])

    source_validation_sheet = wb.create_sheet("Source Validation")
    source_validation = live_validation.get("source_validation", {}) if isinstance(live_validation, dict) else {}
    _write_table(source_validation_sheet, [[key, _excel_value(value)] for key, value in source_validation.items()], ["field", "value"])

    query_plan_sheet = wb.create_sheet("Historical Query Plan")
    query_plan = live_validation.get("historical_query_plan", []) if isinstance(live_validation, dict) else []
    query_headers = search_headers(query_plan)
    _write_table(query_plan_sheet, rows_for_headers(query_plan, query_headers), query_headers)

    analog_review_sheet = wb.create_sheet("Analog Review")
    analog_review = live_validation.get("analog_review", []) if isinstance(live_validation, dict) else []
    analog_headers = search_headers(analog_review)
    _write_table(analog_review_sheet, rows_for_headers(analog_review, analog_headers), analog_headers)

    metric_evidence_sheet = wb.create_sheet("Competition Metric Evidence")
    metric_evidence = live_validation.get("competition_metric_evidence", {}) if isinstance(live_validation, dict) else {}
    _write_table(metric_evidence_sheet, [[key, _excel_value(value)] for key, value in metric_evidence.items()], ["metric", "evidence"])

    live_diag_sheet = wb.create_sheet("Live Historical Diagnostics")
    _write_table(live_diag_sheet, [[key, _excel_value(value)] for key, value in live_summary.items()], ["key", "value"])

    search_sheet = wb.create_sheet("Search diagnostics")
    search_diag = diagnostics.get("search_diagnostics", [])
    search_diag_headers = search_headers(search_diag)
    _write_table(search_sheet, rows_for_headers(search_diag, search_diag_headers), search_diag_headers)

    audit_sheet = wb.create_sheet("Status audit")
    audit_rows = diagnostics.get("status_audit", [])
    audit_headers = search_headers(audit_rows)
    _write_table(audit_sheet, rows_for_headers(audit_rows, audit_headers), audit_headers)

    verify_sheet = wb.create_sheet("Open verification")
    verify_rows = diagnostics.get("open_verifications", [])
    verify_headers = search_headers(verify_rows)
    _write_table(verify_sheet, rows_for_headers(verify_rows, verify_headers), verify_headers)

    all_sheet = wb.create_sheet("All current")
    _write_table(all_sheet, [_row(index + 1, card, assessment) for index, (card, assessment) in enumerate(pairs)], REPORT_COLUMNS)

    changed_sheet = wb.create_sheet("New and changed")
    changed_pairs = [(card, item) for card, item in pairs if item.is_new or item.is_changed]
    _write_table(changed_sheet, [_row(index + 1, card, item) for index, (card, item) in enumerate(changed_pairs)], REPORT_COLUMNS)

    coverage = wb.create_sheet("Search coverage")
    coverage_rows = []
    for card, _assessment in pairs:
        for query in card.search_queries:
            coverage_rows.append([card.procurement_number, _joined(card.search_profiles), query])
    _write_table(coverage, coverage_rows, ["procurement_number", "profiles", "query"])

    diagnostics_sheet = wb.create_sheet("Run diagnostics")
    _write_table(diagnostics_sheet, [[key, json.dumps(value, ensure_ascii=False)] for key, value in diagnostics.items()], ["key", "value"])
    wb.save(path)


def render_markdown(
    summary: dict[str, Any],
    pairs: list[tuple[RadarCard, RadarAssessment]],
    historical_bundles: list[HistoricalAssessmentBundle] | None = None,
    deep_assessments: list[DeepAssessment] | None = None,
    opportunities: list[NoCompetitionOpportunity] | None = None,
) -> str:
    historical_bundles = historical_bundles or []
    deep_assessments = deep_assessments or []
    opportunities = opportunities or []
    deep_by_number = {item.procurement_number: item for item in deep_assessments}
    historical_by_number = {item.procurement_number: item for item in historical_bundles}
    title = "# R3A Live Historical Validation" if summary.get("decision_context") == "HISTORICAL_VALIDATION" else ("# EIS Procurement Radar — Enriched Digest" if deep_assessments else "# EIS Procurement Radar")
    lines = [title, ""]
    lines.append(f"Run: `{summary['run_id']}` | as of `{summary['as_of']}` | version `{summary['radar_version']}`")
    lines.append("")
    for heading, decision in [("Final Priority", RadarDecision.PRIORITY), ("Final Review", RadarDecision.REVIEW)]:
        lines.append(f"## {heading}")
        selected = [(card, item) for card, item in pairs if item.radar_decision == decision][:20]
        if deep_assessments:
            selected = [
                (card, item)
                for card, item in pairs
                if deep_by_number.get(card.procurement_number)
                and deep_by_number[card.procurement_number].final_radar_decision == decision
            ][:20]
        if not selected:
            lines.append("No items.")
        for card, item in selected:
            deep = deep_by_number.get(card.procurement_number)
            score = deep.deep_score if deep else item.total_score
            reasons = _joined((deep.final_decision_reasons if deep else item.positive_reasons)[:4]) or "no reasons recorded"
            lines.append(f"- **{card.procurement_number}** ({score}) {card.title} — {reasons}")
            if card.source_url:
                lines.append(f"  {card.source_url}")
        lines.append("")

    if deep_assessments:
        lines.append("## Decision changes after document analysis")
        changed = [item for item in deep_assessments if item.preliminary_decision != item.final_radar_decision]
        if not changed:
            lines.append("No preliminary-to-final decision changes.")
        for item in changed:
            lines.append(f"- **{item.procurement_number}** {item.preliminary_decision.value} -> {item.final_radar_decision.value}: {_joined(item.final_decision_reasons)}")
        lines.append("")

        lines.append("## Enrichment failures")
        failures = [item for item in deep_assessments if item.error_code]
        if not failures:
            lines.append("No enrichment failures.")
        for item in failures:
            lines.append(f"- **{item.procurement_number}** {item.error_code}: {item.error_message}")
        lines.append("")

        lines.append("## Resource usage")
        lines.append(f"- procurements enriched: {len(deep_assessments)}")
        lines.append(f"- documents downloaded: {summary.get('documents_downloaded', 0)}")
        lines.append(f"- bytes downloaded: {summary.get('total_downloaded_bytes', 0)}")
        lines.append(f"- cached results reused: {summary.get('cached_enrichments_reused', 0)}")
        lines.append("")

    if historical_bundles:
        lines.append("## Historical competition intelligence")
        for card, _assessment in pairs[:20]:
            bundle = historical_by_number.get(card.procurement_number)
            if not bundle:
                continue
            lines.append(
                f"- **{card.procurement_number}** analogs={bundle.competition_metrics.analog_count_total}, "
                f"strong={bundle.competition_metrics.strong_analog_count}, "
                f"median participants={bundle.competition_metrics.median_participants}, "
                f"median reduction={bundle.competition_metrics.median_reduction_percent}, "
                f"risk={bundle.dumping_risk_assessment.risk_level}, "
                f"confidence={bundle.dumping_risk_assessment.confidence}"
            )
        lines.append("")

    if opportunities:
        lines.append("## No competition opportunities")
        for item in opportunities[:20]:
            lines.append(
                f"- **{item.current_procurement_number}** <- {item.previous_procurement_number} "
                f"score={item.opportunity_score}, level={item.opportunity_level}, confidence={item.republication_confidence}"
            )
        lines.append("")

    lines.append("## New and changed")
    selected = [(card, item) for card, item in pairs if item.is_new or item.is_changed][:30]
    if not selected:
        lines.append("No new or changed items.")
    for card, item in selected:
        marker = "new" if item.is_new else "changed"
        lines.append(f"- `{marker}` **{card.procurement_number}** {card.title} ({item.radar_decision.value}, {item.total_score})")
    lines.append("")

    lines.append("## Rejected summary")
    reject_reasons = Counter()
    for _card, item in pairs:
        if item.radar_decision == RadarDecision.REJECT:
            reject_reasons.update(item.hard_reject_reasons or item.negative_reasons[:1] or ["low score"])
    if not reject_reasons:
        lines.append("No rejected items.")
    for reason, count in reject_reasons.most_common(12):
        lines.append(f"- {reason}: {count}")
    lines.append("")

    lines.append("## Run warnings")
    errors = summary.get("errors") or []
    if not errors:
        lines.append("No query errors recorded.")
    for error in errors:
        lines.append(f"- {error}")
    return "\n".join(lines) + "\n"


def flatten_dict_row(row: dict[str, Any]) -> dict[str, Any]:
    return {key: _excel_value(value) for key, value in row.items()}


def search_headers(rows: list[dict[str, Any]]) -> list[str]:
    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)
    return headers or ["empty"]


def rows_for_headers(rows: list[dict[str, Any]], headers: list[str]) -> list[list[Any]]:
    return [[_excel_value(row.get(header, "")) for header in headers] for row in rows]


def write_simple_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    import csv

    headers = search_headers(rows)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(flatten_dict_row(row))


FINAL_COLUMNS = [
    "rank",
    "procurement_number",
    "title",
    "customer",
    "nmck",
    "application_deadline",
    "days_to_deadline",
    "preliminary_score",
    "preliminary_decision",
    "deep_score",
    "final_radar_decision",
    "technical_participation_verdict",
    "overall_recommendation",
    "document_reliability",
    "technical_complexity_score",
    "solo_developer_fit_score",
    "ai_fit_score",
    "recommended_min_price",
    "recommended_comfort_price",
    "nmck_viability",
    "price_margin_percent",
    "specific_platform",
    "required_integrations",
    "key_positive_factors",
    "key_risks",
    "participation_conditions",
    "unanswered_questions",
    "source_url",
]


DEEP_COLUMNS = [
    "procurement_number",
    "preliminary_score",
    "preliminary_decision",
    "deep_score",
    "final_radar_decision",
    "document_reliability",
    "technical_participation_verdict",
    "overall_recommendation",
    "nmck_viability",
    "technical_specification_status",
    "contract_status",
    "application_requirements_status",
    "protocol_status",
    "manual_review_required",
]


def final_row(rank: int, card: RadarCard, preliminary: RadarAssessment, deep: DeepAssessment | None) -> list[Any]:
    if deep is None:
        return [
            rank,
            card.procurement_number,
            card.title,
            card.customer,
            card.nmck,
            card.application_deadline,
            preliminary.days_to_deadline,
            preliminary.total_score,
            preliminary.radar_decision.value,
            "",
            preliminary.radar_decision.value,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            card.source_url,
        ]
    return [
        rank,
        card.procurement_number,
        card.title,
        card.customer,
        card.nmck,
        card.application_deadline,
        preliminary.days_to_deadline,
        preliminary.total_score,
        preliminary.radar_decision.value,
        deep.deep_score,
        deep.final_radar_decision.value,
        deep.technical_participation_verdict,
        deep.overall_recommendation,
        deep.document_reliability,
        deep.technical_complexity_score,
        deep.solo_developer_fit_score,
        deep.ai_fit_score,
        deep.recommended_min_price,
        deep.recommended_comfort_price,
        deep.nmck_viability,
        deep.price_margin_percent,
        deep.specific_platform,
        _joined(deep.required_integrations),
        _joined(deep.key_positive_factors),
        _joined(deep.key_risks),
        _joined(deep.participation_conditions),
        _joined(deep.unanswered_questions),
        card.source_url,
    ]


def deep_row(deep: DeepAssessment) -> list[Any]:
    return [
        deep.procurement_number,
        deep.preliminary_score,
        deep.preliminary_decision.value,
        deep.deep_score,
        deep.final_radar_decision.value,
        deep.document_reliability,
        deep.technical_participation_verdict,
        deep.overall_recommendation,
        deep.nmck_viability,
        deep.technical_specification_status,
        deep.contract_status,
        deep.application_requirements_status,
        deep.protocol_status,
        deep.manual_review_required,
    ]

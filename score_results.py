from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
INPUT_FILE = BASE_DIR / "data" / "web_application.xlsx"
OUTPUT_FILE = BASE_DIR / "data" / "web_application_scored.xlsx"


POSITIVE_RULES = {
    "создание веб-продукта": (
        8,
        [
            r"\bсоздани[ея]\b.*\bсайт",
            r"\bсоздани[ея]\b.*\bвеб[-\s]?прилож",
            r"\bсоздани[ея]\b.*\bпортал",
            r"\bсоздани[ея]\b.*\bличн\w*\s+кабинет",
            r"\bразработк[аи]\b.*\bсайт",
            r"\bразработк[аи]\b.*\bвеб[-\s]?прилож",
            r"\bразработк[аи]\b.*\bпортал",
            r"\bразработк[аи]\b.*\bвеб[-\s]?верс",
        ],
    ),
    "доработка или развитие": (
        6,
        [
            r"\bдоработк",
            r"\bмодернизац",
            r"\bразвити[ея]\b.*\bсайт",
            r"\bразвити[ея]\b.*\bпортал",
            r"\bразвити[ея]\b.*\bинформационн\w*\s+систем",
            r"\bрасширени[ея]\s+функциональн",
            r"\bредизайн",
        ],
    ),
    "веб-модуль": (
        5,
        [
            r"\bвеб[-\s]?клиент",
            r"\bвеб[-\s]?модул",
            r"\bмобильн\w*\s+приложени",
            r"\bличн\w*\s+кабинет",
            r"\bинтернет[-\s]?портал",
            r"\bгеопортал",
        ],
    ),
    "внедрение": (
        4,
        [
            r"\bвнедрени",
            r"\bпроектировани",
        ],
    ),
    "сопровождение с развитием": (
        3,
        [
            r"\bсопровождени.*\bдоработ",
            r"\bподдержк.*\bмодернизац",
            r"\bтехническ\w*\s+поддержк.*\bразработ",
        ],
    ),
}


NEGATIVE_RULES = {
    "лицензии и права": (
        -8,
        [
            r"\bнеисключительн\w*\s+прав",
            r"\bправ[ао]\s+на\s+использован",
            r"\bлицензи",
            r"\bпродлени",
            r"\bпередач[аи]\s+прав",
        ],
    ),
    "защита веб-приложений": (
        -8,
        [
            r"\bмежсетев\w*\s+экран",
            r"\bwaf\b",
            r"\bзащит[аы]\s+веб[-\s]?прилож",
            r"\bapplication firewall",
            r"\bнесанкционированн\w*\s+доступ",
        ],
    ),
    "товарная закупка": (
        -10,
        [
            r"^\s*поставка\b",
            r"\bкнижн\w*\s+продукц",
            r"\bиздательск\w*\s+продукц",
            r"\bоборудован",
            r"\bкомпьютерн\w*\s+техник",
            r"\bкартридж",
        ],
    ),
    "обучение": (
        -7,
        [
            r"\bобучени",
            r"\bповышени[ея]\s+квалификац",
            r"\bпрофессиональн\w*\s+образован",
            r"\bтренинг",
        ],
    ),
    "доступ к готовой системе": (
        -6,
        [
            r"\bпредоставлени[ея]\s+доступ",
            r"\bдоступ[а]?\s+к\s+веб[-\s]?прилож",
            r"\bготов\w*\s+программ",
        ],
    ),
    "только сопровождение": (
        -2,
        [
            r"^\s*(оказание\s+услуг\s+по\s+)?сопровождени",
            r"^\s*(оказание\s+услуг\s+по\s+)?техническ\w*\s+поддержк",
            r"^\s*обслуживани",
        ],
    ),
}


def normalize(value: object) -> str:
    if value is None:
        return ""

    return re.sub(r"\s+", " ", str(value)).strip().lower()


def match_rules(text: str, rules: dict[str, tuple[int, list[str]]]):
    score = 0
    reasons: list[str] = []

    for reason, (points, patterns) in rules.items():
        if any(re.search(pattern, text, flags=re.IGNORECASE) for pattern in patterns):
            score += points
            reasons.append(f"{reason}: {points:+d}")

    return score, reasons


def project_type(text: str) -> str:
    checks = [
        ("Сайт и мобильное приложение", [
            r"\bмобильн\w*\s+приложени",
            r"\bсайт",
        ]),
        ("Веб-приложение", [
            r"\bвеб[-\s]?прилож",
            r"\bвеб[-\s]?верс",
        ]),
        ("Сайт", [
            r"\bсайт",
        ]),
        ("Портал", [
            r"\bпортал",
            r"\bгеопортал",
        ]),
        ("Личный кабинет", [
            r"\bличн\w*\s+кабинет",
        ]),
        ("Информационная система", [
            r"\bинформационн\w*\s+систем",
        ]),
        ("Сопровождение", [
            r"\bсопровождени",
            r"\bтехническ\w*\s+поддержк",
        ]),
    ]

    # Сначала отдельно проверяем комбинацию.
    if re.search(r"\bмобильн\w*\s+приложени", text) and re.search(
        r"\bсайт", text
    ):
        return "Сайт и мобильное приложение"

    for label, patterns in checks[1:]:
        if any(re.search(pattern, text) for pattern in patterns):
            return label

    return "Другое"


def classify(score: int, text: str) -> tuple[str, str]:
    hard_exclusions = [
        r"^\s*поставка\b",
        r"\bмежсетев\w*\s+экран",
        r"\bwaf\b",
        r"\bзащит[аы]\s+веб[-\s]?прилож",
        r"\bобучени",
        r"\bнеисключительн\w*\s+прав",
    ]

    if any(re.search(pattern, text) for pattern in hard_exclusions):
        return "D", "Исключить"

    if score >= 8:
        return "A", "Открыть в первую очередь"

    if score >= 4:
        return "B", "Проверить карточку"

    if score >= 1:
        return "C", "Смежная или неоднозначная закупка"

    return "D", "Исключить"


def parse_year(date_value: object) -> int | None:
    if not date_value:
        return None

    match = re.search(r"\b(20\d{2})\b", str(date_value))
    return int(match.group(1)) if match else None


def main() -> None:
    if not INPUT_FILE.exists():
        raise FileNotFoundError(
            f"Не найден входной файл: {INPUT_FILE}"
        )

    workbook = load_workbook(INPUT_FILE)
    sheet = workbook.active

    headers = {
        str(cell.value): cell.column
        for cell in sheet[1]
        if cell.value is not None
    }

    required = ["object_name", "raw_text", "published_date"]

    for name in required:
        if name not in headers:
            raise ValueError(f"Нет обязательного столбца: {name}")

    new_headers = [
        "relevance_score",
        "relevance_group",
        "project_type",
        "positive_reasons",
        "negative_reasons",
        "market_period",
        "recommendation",
    ]

    start_column = sheet.max_column + 1

    for offset, name in enumerate(new_headers):
        cell = sheet.cell(row=1, column=start_column + offset, value=name)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E4DFEC")

    for row in range(2, sheet.max_row + 1):
        object_name = normalize(
            sheet.cell(row=row, column=headers["object_name"]).value
        )
        raw_text = normalize(
            sheet.cell(row=row, column=headers["raw_text"]).value
        )
        text = f"{object_name} {raw_text}"

        positive_score, positive_reasons = match_rules(
            text,
            POSITIVE_RULES,
        )
        negative_score, negative_reasons = match_rules(
            text,
            NEGATIVE_RULES,
        )

        score = positive_score + negative_score
        group, recommendation = classify(score, text)

        date_value = sheet.cell(
            row=row,
            column=headers["published_date"],
        ).value
        year = parse_year(date_value)

        if year is None:
            market_period = "Дата не определена"
        elif year >= 2024:
            market_period = "Актуальная: 2024–2026"
            score += 2
        elif year >= 2021:
            market_period = "Условно актуальная: 2021–2023"
        else:
            market_period = "Архивная: до 2021"
            score -= 2

        # Пересчитываем группу после поправки на год.
        group, recommendation = classify(score, text)

        values = [
            score,
            group,
            project_type(text),
            "; ".join(positive_reasons),
            "; ".join(negative_reasons),
            market_period,
            recommendation,
        ]

        for offset, value in enumerate(values):
            sheet.cell(
                row=row,
                column=start_column + offset,
                value=value,
            )

    sheet.auto_filter.ref = sheet.dimensions
    sheet.freeze_panes = "A2"

    group_column = start_column + 1

    fills = {
        "A": PatternFill("solid", fgColor="C6EFCE"),
        "B": PatternFill("solid", fgColor="FFEB9C"),
        "C": PatternFill("solid", fgColor="FCE4D6"),
        "D": PatternFill("solid", fgColor="FFC7CE"),
    }

    for row in range(2, sheet.max_row + 1):
        group_cell = sheet.cell(row=row, column=group_column)
        if group_cell.value in fills:
            group_cell.fill = fills[group_cell.value]
            group_cell.font = Font(bold=True)

    widths = {
        start_column: 18,
        start_column + 1: 17,
        start_column + 2: 30,
        start_column + 3: 55,
        start_column + 4: 55,
        start_column + 5: 30,
        start_column + 6: 32,
    }

    for column, width in widths.items():
        sheet.column_dimensions[
            sheet.cell(row=1, column=column).column_letter
        ].width = width

    workbook.save(OUTPUT_FILE)

    counts = {"A": 0, "B": 0, "C": 0, "D": 0}

    for row in range(2, sheet.max_row + 1):
        group = sheet.cell(row=row, column=group_column).value
        if group in counts:
            counts[group] += 1

    print(f"Готово: {OUTPUT_FILE}")
    print("Распределение:")
    for group, count in counts.items():
        print(f"  {group}: {count}")


if __name__ == "__main__":
    main()

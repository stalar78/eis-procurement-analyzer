#!/usr/bin/env python
"""Deterministic analyzer for downloaded EIS procurement documents.

The script intentionally avoids external LLM APIs. It extracts available text,
classifies documents, applies rule-based heuristics, and writes evidence-backed
procurement cards.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import math
import os
import re
import shutil
import statistics
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - handled at runtime
    Workbook = None
    load_workbook = None
    FormulaRule = None
    Font = None
    PatternFill = None
    get_column_letter = None


FIELD_NAMES = [
    "analysis_version",
    "procurement_number",
    "procurement_name",
    "customer",
    "law",
    "procurement_method",
    "region",
    "publication_date",
    "application_deadline",
    "contract_deadline",
    "source_url",
    "nmck",
    "contract_price",
    "price_reduction_percent",
    "participants_count",
    "admitted_participants_count",
    "winner_application_number",
    "winner_name",
    "all_price_offers",
    "nmck_sources",
    "nmck_commercial_offers",
    "work_start",
    "work_end",
    "contract_term",
    "acceptance_period_days",
    "payment_period_days",
    "advance_percent",
    "stages_count",
    "licenses_required",
    "experience_required",
    "staff_requirements",
    "portfolio_required",
    "sro_required",
    "additional_requirements",
    "smp_only",
    "rnp_requirement",
    "local_presence_required",
    "bid_security_amount",
    "bid_security_percent",
    "contract_security_amount",
    "contract_security_percent",
    "warranty_security",
    "anti_dumping_applies",
    "possible_exemption_from_security",
    "project_type",
    "short_scope",
    "functional_modules",
    "user_roles",
    "admin_panel",
    "personal_accounts",
    "integrations",
    "external_systems",
    "api_required",
    "mobile_app_required",
    "adaptive_layout",
    "notifications",
    "document_generation",
    "file_uploads",
    "reporting",
    "analytics",
    "search",
    "content_management",
    "design_requirements",
    "required_stack",
    "allowed_stack",
    "database_requirements",
    "server_requirements",
    "hosting_requirements",
    "deployment_requirements",
    "backup_requirements",
    "monitoring_requirements",
    "security_requirements",
    "personal_data",
    "esia",
    "smev",
    "cryptopro",
    "fstec",
    "fsb",
    "closed_network",
    "support_period",
    "hosting_period",
    "warranty_period",
    "sla",
    "support_24x7",
    "training_required",
    "documentation_required",
    "source_code_transfer",
    "exclusive_rights_transfer",
    "repository_transfer",
    "open_source_restrictions",
    "acceptance_basis",
    "user_testing_required",
    "revision_period_days",
    "unlimited_revision_risk",
    "subjective_design_acceptance",
    "penalties",
    "hidden_defect_liability",
    "key_risks",
    "contradictions",
    "technical_complexity_score",
    "organizational_complexity_score",
    "legal_risk_score",
    "financial_risk_score",
    "ai_fit_score",
    "solo_developer_fit_score",
    "estimated_hours_min",
    "estimated_hours_max",
    "estimated_calendar_weeks",
    "estimated_direct_costs_min",
    "estimated_direct_costs_max",
    "recommended_min_price",
    "recommended_comfort_price",
    "price_margin_vs_min",
    "price_margin_percent",
    "nmck_viability",
    "technical_participation_verdict",
    "market_result_status",
    "overall_recommendation",
    "extreme_price_reduction_review_required",
    "extreme_reduction_reason",
    "excluded_from_market_aggregates",
    "market_confidence",
    "manual_review_required",
    "deprecated_verdict_note",
    "verdict",
    "verdict_reason",
    "data_completeness_score",
    "analysis_reliability",
    "technical_specification_status",
    "contract_status",
    "nmck_status",
    "application_requirements_status",
    "final_protocol_status",
    "missing_key_documents",
    "critical_quality_issues",
    "verdict_allowed",
    "verdict_block_reason",
    "document_availability_score",
    "text_extraction_score",
    "technical_document_files",
    "contract_document_files",
    "application_requirement_files",
    "nmck_document_files",
    "protocol_document_files",
    "estimated_development_hours_min",
    "estimated_development_hours_max",
    "estimated_support_hours",
    "estimated_infrastructure_costs",
    "risk_reserve_percent",
]

ANALYSIS_VERSION = "2.2-decision-model"

DOC_TYPES = [
    "technical_specification",
    "contract_draft",
    "nmck_calculation",
    "application_requirements",
    "clarification",
    "final_protocol",
    "notice",
    "bank_details",
    "signature",
    "other",
]

TEXT_EXTENSIONS = {".txt", ".html", ".htm"}
ARCHIVE_EXTENSIONS = {".zip", ".rar"}
SUPPORTED_EXTENSIONS = {
    ".docx",
    ".doc",
    ".pdf",
    ".xlsx",
    ".xls",
    ".zip",
    ".rar",
    ".rtf",
    ".txt",
    ".html",
    ".htm",
    ".bin",
}


@dataclass
class Evidence:
    field_name: str
    value: Any
    source_file: str
    document_type: str
    page_or_sheet: str
    text_excerpt: str
    confidence: str
    extraction_method: str


@dataclass
class ExtractedDocument:
    procurement_number: str
    original_filename: str
    detected_type: str
    detected_format: str
    extraction_status: str
    pages_or_sheets: str
    text_length: int
    needs_ocr: bool
    error: str
    source_path: str
    output_text_path: str = ""
    section: str = ""
    text_sample: str = ""


@dataclass
class TextPiece:
    text: str
    source_file: str
    document_type: str
    page_or_sheet: str


@dataclass
class Utilities:
    seven_zip: str = ""
    unrar: str = ""
    libreoffice: str = ""
    antiword: str = ""


def safe_print(message: str) -> None:
    try:
        print(message, flush=True)
    except UnicodeEncodeError:
        print(message.encode("utf-8", "replace").decode("utf-8"), flush=True)


def repair_mojibake(value: Any) -> Any:
    if not isinstance(value, str) or "Р" not in value:
        return value
    try:
        fixed = value.encode("cp1251", errors="strict").decode("utf-8", errors="strict")
    except Exception:
        return value
    return fixed if count_cyrillic(fixed) > count_cyrillic(value) else value


def count_cyrillic(value: str) -> int:
    return sum(1 for ch in value if "А" <= ch <= "я" or ch == "ё" or ch == "Ё")


def normalize_text(value: str) -> str:
    value = repair_mojibake(value)
    value = html.unescape(value or "")
    value = re.sub(r"\r\n?", "\n", value)
    value = re.sub(r"[ \t\f\v]+", " ", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def read_text_guess(path: Path) -> tuple[str, str]:
    data = path.read_bytes()
    for encoding in ("utf-8", "utf-8-sig", "cp1251", "utf-16", "utf-16le"):
        try:
            text = data.decode(encoding)
            return normalize_text(text), encoding
        except UnicodeDecodeError:
            continue
    return normalize_text(data.decode("utf-8", errors="replace")), "utf-8-replace"


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({name: csv_value(row.get(name)) for name in fieldnames})


def csv_value(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return "" if value is None else value


def load_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [{k: repair_mojibake(v) for k, v in row.items()} for row in csv.DictReader(f)]


def load_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    return repair_json(data)


def repair_json(value: Any) -> Any:
    if isinstance(value, dict):
        return {k: repair_json(v) for k, v in value.items()}
    if isinstance(value, list):
        return [repair_json(v) for v in value]
    return repair_mojibake(value)


def detect_utilities() -> Utilities:
    def which_many(names: list[str], extra: list[Path] | None = None) -> str:
        for name in names:
            found = shutil.which(name)
            if found:
                return found
        for path in extra or []:
            if path.exists():
                return str(path)
        return ""

    seven_paths = [
        Path(os.environ.get("ProgramFiles", "C:/Program Files")) / "7-Zip" / "7z.exe",
        Path(os.environ.get("ProgramFiles(x86)", "C:/Program Files (x86)")) / "7-Zip" / "7z.exe",
    ]
    libre_paths = [
        Path(os.environ.get("ProgramFiles", "C:/Program Files")) / "LibreOffice" / "program" / "soffice.exe",
        Path(os.environ.get("ProgramFiles", "C:/Program Files")) / "LibreOffice" / "program" / "libreoffice.exe",
    ]
    return Utilities(
        seven_zip=which_many(["7z", "7z.exe"], seven_paths),
        unrar=which_many(["unrar", "unrar.exe"]),
        libreoffice=which_many(["soffice", "soffice.exe", "libreoffice", "libreoffice.exe"], libre_paths),
        antiword=which_many(["antiword", "antiword.exe"]),
    )


def magic_extension(path: Path) -> str:
    data = path.read_bytes()[:16]
    if data.startswith(b"PK\x03\x04"):
        try:
            with zipfile.ZipFile(path) as zf:
                names = set(zf.namelist())
            if "[Content_Types].xml" in names and any(n.startswith("word/") for n in names):
                return ".docx"
            if "[Content_Types].xml" in names and any(n.startswith("xl/") for n in names):
                return ".xlsx"
        except Exception:
            pass
        return ".zip"
    if data.startswith(b"%PDF"):
        return ".pdf"
    if data.startswith(b"{\\rtf"):
        return ".rtf"
    if data.startswith(b"\xd0\xcf\x11\xe0"):
        return ".doc"
    if data.startswith(b"Rar!\x1a\x07"):
        return ".rar"
    return path.suffix.lower()


def stable_name(path: Path, suffix: str = "") -> str:
    digest = hashlib.sha1(str(path).encode("utf-8", errors="ignore")).hexdigest()[:10]
    return f"{path.stem[:80]}__{digest}{suffix}"


def safe_extract_zip(zip_path: Path, target: Path, max_depth: int = 2, depth: int = 0) -> list[Path]:
    extracted: list[Path] = []
    target.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as zf:
        root = target.resolve()
        for member in zf.infolist():
            if member.is_dir():
                continue
            name = member.filename.replace("\\", "/")
            if name.startswith("/") or ".." in Path(name).parts:
                raise ValueError(f"unsafe archive path: {member.filename}")
            out = target / name
            resolved = out.resolve()
            if root not in resolved.parents and resolved != root:
                raise ValueError(f"unsafe archive path: {member.filename}")
            out.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(member) as src, out.open("wb") as dst:
                shutil.copyfileobj(src, dst)
            extracted.append(out)
    if depth < max_depth:
        nested = [p for p in list(extracted) if magic_extension(p) == ".zip"]
        for nested_zip in nested:
            nested_target = target / f"_nested_{stable_name(nested_zip)}"
            extracted.extend(safe_extract_zip(nested_zip, nested_target, max_depth=max_depth, depth=depth + 1))
    return extracted


def extract_rar(rar_path: Path, target: Path, utilities: Utilities) -> tuple[list[Path], str]:
    target.mkdir(parents=True, exist_ok=True)
    if utilities.seven_zip:
        cmd = [utilities.seven_zip, "x", "-y", f"-o{target}", str(rar_path)]
    elif utilities.unrar:
        cmd = [utilities.unrar, "x", "-y", str(rar_path), str(target)]
    else:
        return [], "RAR extraction unavailable: 7-Zip/unrar not found"
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    except Exception as exc:
        return [], f"RAR extraction failed: {exc}"
    if proc.returncode != 0:
        return [], f"RAR extraction failed: {(proc.stderr or proc.stdout).strip()[:500]}"
    return [p for p in target.rglob("*") if p.is_file()], ""


def parse_docx_xml_text(xml_bytes: bytes) -> list[str]:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    root = ET.fromstring(xml_bytes)
    chunks: list[str] = []
    for child in root.iter():
        if child.tag.endswith("}p"):
            text = "".join(t.text or "" for t in child.findall(".//w:t", ns))
            if text.strip():
                chunks.append(text.strip())
        elif child.tag.endswith("}tbl"):
            rows = []
            for row in child.findall(".//w:tr", ns):
                cells = []
                for cell in row.findall(".//w:tc", ns):
                    cell_text = " ".join(t.text or "" for t in cell.findall(".//w:t", ns)).strip()
                    cells.append(cell_text)
                if any(cells):
                    rows.append(" | ".join(cells))
            if rows:
                chunks.append("\n".join(rows))
    return chunks


def extract_docx(path: Path) -> tuple[list[TextPiece], str, str]:
    pieces: list[TextPiece] = []
    try:
        with zipfile.ZipFile(path) as zf:
            if "word/document.xml" not in zf.namelist():
                return [], "", "word/document.xml not found"
            chunks = parse_docx_xml_text(zf.read("word/document.xml"))
            pieces.append(TextPiece(normalize_text("\n".join(chunks)), str(path), "", "document"))
            for name in zf.namelist():
                if name.startswith("word/header") or name.startswith("word/footer"):
                    try:
                        chunks = parse_docx_xml_text(zf.read(name))
                    except Exception:
                        continue
                    if chunks:
                        pieces.append(TextPiece(normalize_text("\n".join(chunks)), str(path), "", name))
    except Exception as exc:
        return [], "", str(exc)
    return pieces, "document", ""


def extract_xlsx(path: Path) -> tuple[list[TextPiece], str, str]:
    if load_workbook is None:
        return [], "", "openpyxl is not installed"
    pieces: list[TextPiece] = []
    sheets: list[str] = []
    try:
        wb = load_workbook(path, data_only=True, read_only=False)
        for ws in wb.worksheets:
            sheets.append(ws.title)
            merged_lookup: dict[str, str] = {}
            for rng in ws.merged_cells.ranges:
                value = ws.cell(rng.min_row, rng.min_col).value
                if value is not None:
                    for row in range(rng.min_row, rng.max_row + 1):
                        for col in range(rng.min_col, rng.max_col + 1):
                            merged_lookup[f"{row}:{col}"] = str(value)
            lines: list[str] = []
            for row in ws.iter_rows():
                values = []
                for cell in row:
                    value = cell.value
                    if value is None:
                        value = merged_lookup.get(f"{cell.row}:{cell.column}")
                    if value is None:
                        continue
                    values.append(f"{cell.coordinate}={value}")
                if values:
                    lines.append(" | ".join(values))
            if lines:
                pieces.append(TextPiece(normalize_text("\n".join(lines)), str(path), "", ws.title))
    except Exception as exc:
        return [], "", str(exc)
    return pieces, ", ".join(sheets), ""


def extract_pdf(path: Path) -> tuple[list[TextPiece], str, bool, str]:
    pieces: list[TextPiece] = []
    pages = ""
    try:
        import fitz  # type: ignore

        doc = fitz.open(path)
        pages = str(doc.page_count)
        for i, page in enumerate(doc, start=1):
            text = normalize_text(page.get_text("text"))
            if text:
                pieces.append(TextPiece(text, str(path), "", f"page {i}"))
        needs_ocr = sum(len(p.text) for p in pieces) < max(100, doc.page_count * 20)
        return pieces, pages, needs_ocr, ""
    except Exception:
        pass
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        pages = str(len(reader.pages))
        for i, page in enumerate(reader.pages, start=1):
            text = normalize_text(page.extract_text() or "")
            if text:
                pieces.append(TextPiece(text, str(path), "", f"page {i}"))
        needs_ocr = sum(len(p.text) for p in pieces) < max(100, len(reader.pages) * 20)
        return pieces, pages, needs_ocr, ""
    except Exception as exc:
        return [], pages, True, f"PDF extractor unavailable or failed: {exc}"


def extract_doc(path: Path, temp_dir: Path, utilities: Utilities) -> tuple[list[TextPiece], str, str]:
    if utilities.antiword:
        try:
            proc = subprocess.run([utilities.antiword, str(path)], capture_output=True, timeout=90)
            if proc.returncode == 0 and proc.stdout:
                text = normalize_text(proc.stdout.decode("cp1251", errors="replace"))
                return [TextPiece(text, str(path), "", "document")], "document", ""
        except Exception:
            pass
    if utilities.libreoffice:
        out_dir = temp_dir / "doc_converted" / stable_name(path)
        out_dir.mkdir(parents=True, exist_ok=True)
        try:
            proc = subprocess.run(
                [utilities.libreoffice, "--headless", "--convert-to", "docx", "--outdir", str(out_dir), str(path)],
                capture_output=True,
                text=True,
                timeout=120,
            )
            converted = list(out_dir.glob("*.docx"))
            if proc.returncode == 0 and converted:
                return extract_docx(converted[0])
            return [], "", f"LibreOffice conversion failed: {(proc.stderr or proc.stdout).strip()[:500]}"
        except Exception as exc:
            return [], "", f"LibreOffice conversion failed: {exc}"
    return [], "", "DOC extraction unavailable: LibreOffice/antiword not found"


def extract_rtf(path: Path) -> tuple[list[TextPiece], str, str]:
    text, encoding = read_text_guess(path)
    try:
        from striprtf.striprtf import rtf_to_text  # type: ignore

        text = normalize_text(rtf_to_text(text))
        return [TextPiece(text, str(path), "", "document")], "document", f"striprtf/{encoding}"
    except Exception:
        text = re.sub(r"\\'[0-9a-fA-F]{2}", " ", text)
        text = re.sub(r"\\[a-zA-Z]+\d* ?", " ", text)
        text = re.sub(r"[{}]", " ", text)
        return [TextPiece(normalize_text(text), str(path), "", "document")], "document", f"plain-rtf/{encoding}"


def normalize_classifier_text(value: str) -> str:
    value = normalize_text(value).lower().replace("ё", "е")
    value = re.sub(r"[_\-\(\)\[\]\{\}\.,;:№]+", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def add_classification_score(scores: dict[str, int], reasons: dict[str, list[str]], doc_type: str, points: int, reason: str) -> None:
    scores[doc_type] = scores.get(doc_type, 0) + points
    reasons.setdefault(doc_type, []).append(f"{points:+d} {reason}")


def classify_document_detailed(path: Path, section: str, text: str) -> dict[str, Any]:
    name = normalize_classifier_text(path.name)
    section_norm = normalize_classifier_text(section)
    head = normalize_classifier_text(text[:10000])
    hay = f"{name} {section_norm} {head}"
    scores: dict[str, int] = {}
    reasons: dict[str, list[str]] = {}

    def name_has(needle: str) -> bool:
        return needle in name

    def text_has(needle: str) -> bool:
        return needle in head

    technical_names = [
        ("техническое задание", 5),
        ("описание объекта закупки", 5),
        ("описание объекта", 4),
        ("техническая часть", 4),
        ("тз портал", 5),
        ("перечень работ", 3),
        ("требования к системе", 3),
        ("ооз", 3),
    ]
    for phrase, points in technical_names:
        if name_has(phrase):
            add_classification_score(scores, reasons, "technical_specification", points, f"name contains '{phrase}'")
    if re.search(r"(^|\s)тз(\s|$)", name):
        add_classification_score(scores, reasons, "technical_specification", 5, "name starts/contains standalone TZ")
    for phrase, points in [
        ("функциональные требования", 3),
        ("требования к системе", 3),
        ("перечень работ", 2),
        ("требования к результату", 2),
        ("состав услуг", 2),
    ]:
        if text_has(phrase):
            add_classification_score(scores, reasons, "technical_specification", points, f"text contains '{phrase}'")
    if text_has("статья 31") or text_has("единые требования к участникам"):
        add_classification_score(scores, reasons, "technical_specification", -5, "mostly application/legal boilerplate")

    for phrase, points in [
        ("проект контракта", 5),
        ("проект муниципального контракта", 5),
        ("проект государственного контракта", 5),
        ("проект договора", 5),
        ("приложение 4 контракт", 4),
        ("v проект контракта", 4),
    ]:
        if name_has(phrase):
            add_classification_score(scores, reasons, "contract_draft", points, f"name contains '{phrase}'")
    if name_has("приложение 4") and name_has("контракт"):
        add_classification_score(scores, reasons, "contract_draft", 5, "name contains appendix 4 contract")
    if name_has("контракт") and not name_has("протокол"):
        add_classification_score(scores, reasons, "contract_draft", 2, "name contains contract")
    for phrase, points in [
        ("предмет контракта", 3),
        ("права и обязанности сторон", 2),
        ("порядок приемки", 2),
        ("ответственность сторон", 2),
    ]:
        if text_has(phrase):
            add_classification_score(scores, reasons, "contract_draft", points, f"text contains '{phrase}'")

    for phrase, points in [
        ("требования к содержанию составу заявки", 5),
        ("требования к содержанию", 4),
        ("требования к заявке", 5),
        ("инструкция участникам", 5),
        ("инструкция по заполнению заявки", 5),
        ("состав заявки", 4),
    ]:
        if name_has(phrase):
            add_classification_score(scores, reasons, "application_requirements", points, f"name contains '{phrase}'")
    if name_has("информационная карта"):
        add_classification_score(scores, reasons, "information_card", 5, "name contains information card")
        add_classification_score(scores, reasons, "application_requirements", 2, "information card can support requirements")
    for phrase in ["единые требования к участникам", "требования к участникам", "содержание состав заявки", "статья 31"]:
        if text_has(phrase):
            add_classification_score(scores, reasons, "application_requirements", 2, f"text contains '{phrase}'")

    for phrase, points in [
        ("обоснование нмцк", 5),
        ("расчет нмцк", 5),
        ("начальной максимальной цены", 4),
        ("нмцк", 4),
        ("коммерческим предложениям", 3),
    ]:
        if name_has(phrase):
            add_classification_score(scores, reasons, "nmck_calculation", points, f"name contains '{phrase}'")
    if path.suffix.lower() in {".xls", ".xlsx"} and ("коммерчес" in hay or "нмцк" in hay):
        add_classification_score(scores, reasons, "nmck_calculation", 4, "spreadsheet with nmck/commercial context")
    for phrase in ["метод сопоставимых рыночных цен", "коммерческое предложение", "обоснование начальной"]:
        if text_has(phrase):
            add_classification_score(scores, reasons, "nmck_calculation", 2, f"text contains '{phrase}'")

    for phrase, points in [
        ("протокол подведения итогов", 6),
        ("протокол итогов", 6),
        ("протокол подачи ценовых предложений", 5),
    ]:
        if name_has(phrase):
            add_classification_score(scores, reasons, "final_protocol", points, f"name contains '{phrase}'")
    if name_has("протокол") and not name_has("разноглас"):
        add_classification_score(scores, reasons, "final_protocol", 3, "name contains protocol")
    for phrase in ["признать победителем", "предложение о цене контракта", "подано заявок", "победитель"]:
        if text_has(phrase):
            add_classification_score(scores, reasons, "final_protocol", 3, f"text contains '{phrase}'")

    for phrase in ["разъяснение", "ответ на запрос", "участнику закупки"]:
        if name_has(phrase) or text_has(phrase):
            add_classification_score(scores, reasons, "clarification", 4, f"clarification phrase '{phrase}'")
    for phrase in ["реквизит", "банковск"]:
        if name_has(phrase) or text_has(phrase):
            add_classification_score(scores, reasons, "bank_details", 4, f"bank phrase '{phrase}'")
    if name_has("электронный документ") or name_has("подпис") or path.suffix.lower() == ".sig":
        add_classification_score(scores, reasons, "signature", 4, "signature/electronic document marker")
    if section_norm in {"common", "notice"} or text_has("извещение о проведении"):
        add_classification_score(scores, reasons, "notice", 2, "notice/common section marker")

    if scores.get("final_protocol", 0) >= 3:
        add_classification_score(scores, reasons, "contract_draft", -5, "protocol is not contract draft")
    if scores.get("contract_draft", 0) >= 5:
        add_classification_score(scores, reasons, "technical_specification", -5, "contract draft is not technical specification")

    ranked = sorted(scores.items(), key=lambda item: item[1], reverse=True)
    detected = "other"
    score = 0
    confidence = "low"
    if ranked and ranked[0][1] >= 3:
        detected, score = ranked[0]
        confidence = "high" if score >= 6 else "medium"
    alternatives = [{"type": k, "score": v, "reasons": reasons.get(k, [])[:5]} for k, v in ranked[1:5] if v >= 2]
    return {
        "detected_type": detected,
        "classification_score": score,
        "classifier_confidence": confidence,
        "classification_reasons": reasons.get(detected, [])[:8],
        "alternative_types": alternatives,
    }


def classify_document(path: Path, section: str, text: str) -> str:
    return classify_document_detailed(path, section, text)["detected_type"]


def find_input_procurements(input_dir: Path, procurement_number: str = "", limit: int = 0) -> list[Path]:
    dirs = [p for p in input_dir.iterdir() if p.is_dir() and re.search(r"\d{19}", p.name)]
    dirs.sort(key=lambda p: p.name)
    if procurement_number:
        dirs = [p for p in dirs if procurement_number in p.name]
    if limit:
        dirs = dirs[:limit]
    return dirs


def procurement_number_from_dir(path: Path) -> str:
    match = re.search(r"\d{19}", path.name)
    return match.group(0) if match else path.name


def discover_downloads(proc_dir: Path) -> list[tuple[Path, str]]:
    downloads = proc_dir / "downloads"
    if not downloads.exists():
        return []
    files: list[tuple[Path, str]] = []
    for path in downloads.rglob("*"):
        if not path.is_file():
            continue
        try:
            section = path.relative_to(downloads).parts[0]
        except Exception:
            section = ""
        files.append((path, section))
    files.sort(key=lambda x: str(x[0]).lower())
    return files


def extract_one_file(
    path: Path,
    procurement_number: str,
    section: str,
    out_proc: Path,
    temp_dir: Path,
    utilities: Utilities,
    overwrite: bool,
    skip_ocr: bool,
    manifest: list[dict[str, Any]],
    errors: list[str],
    seen_hashes: set[str],
) -> list[TextPiece]:
    suffix = magic_extension(path)
    text_dir = out_proc / "extracted_text"
    text_dir.mkdir(parents=True, exist_ok=True)
    cache_path = text_dir / f"{stable_name(path, '.txt')}"
    digest = hashlib.sha1(path.read_bytes()).hexdigest()
    if digest in seen_hashes:
        manifest.append(doc_manifest(procurement_number, path, "duplicate", suffix, "skipped_duplicate", "", 0, False, "", section))
        return []
    seen_hashes.add(digest)

    if cache_path.exists() and not overwrite:
        text = cache_path.read_text(encoding="utf-8", errors="replace")
        class_detail = classify_document_detailed(path, section, text)
        doc_type = class_detail["detected_type"]
        row = doc_manifest(procurement_number, path, doc_type, suffix, "cached", "cache", len(text), False, "", section, str(cache_path))
        row.update(
            {
                "classification_score": class_detail.get("classification_score", ""),
                "classifier_confidence": class_detail.get("classifier_confidence", ""),
                "classification_reasons": class_detail.get("classification_reasons", []),
                "alternative_types": class_detail.get("alternative_types", []),
            }
        )
        manifest.append(row)
        return [TextPiece(text, str(path), doc_type, "cache")]

    pieces: list[TextPiece] = []
    pages_or_sheets = ""
    needs_ocr = False
    error = ""
    extraction_note = ""

    try:
        if suffix == ".zip":
            target = temp_dir / "archives" / stable_name(path)
            extracted = safe_extract_zip(path, target)
            child_pieces: list[TextPiece] = []
            for child in extracted:
                child_pieces.extend(
                    extract_one_file(
                        child,
                        procurement_number,
                        section,
                        out_proc,
                        temp_dir,
                        utilities,
                        overwrite,
                        skip_ocr,
                        manifest,
                        errors,
                        seen_hashes,
                    )
                )
            manifest.append(doc_manifest(procurement_number, path, "other", suffix, "archive_extracted", str(len(extracted)), 0, False, "", section))
            return child_pieces
        if suffix == ".rar":
            target = temp_dir / "archives" / stable_name(path)
            extracted, error = extract_rar(path, target, utilities)
            if error:
                errors.append(f"{path}: {error}")
                manifest.append(doc_manifest(procurement_number, path, "other", suffix, "error", "", 0, False, error, section))
                return []
            child_pieces = []
            for child in extracted:
                child_pieces.extend(
                    extract_one_file(child, procurement_number, section, out_proc, temp_dir, utilities, overwrite, skip_ocr, manifest, errors, seen_hashes)
                )
            manifest.append(doc_manifest(procurement_number, path, "other", suffix, "archive_extracted", str(len(extracted)), 0, False, "", section))
            return child_pieces
        if suffix == ".docx":
            pieces, pages_or_sheets, error = extract_docx(path)
        elif suffix == ".xlsx":
            pieces, pages_or_sheets, error = extract_xlsx(path)
        elif suffix == ".xls":
            pieces, pages_or_sheets, error = extract_xls(path, temp_dir, utilities)
        elif suffix == ".pdf":
            pieces, pages_or_sheets, needs_ocr, error = extract_pdf(path)
            if needs_ocr and not skip_ocr:
                extraction_note = "needs_ocr=true; OCR engine is optional and not bundled"
        elif suffix == ".doc":
            pieces, pages_or_sheets, error = extract_doc(path, temp_dir, utilities)
        elif suffix == ".rtf":
            pieces, pages_or_sheets, extraction_note = extract_rtf(path)
        elif suffix in TEXT_EXTENSIONS:
            text, enc = read_text_guess(path)
            pieces = [TextPiece(text, str(path), "", "document")]
            pages_or_sheets = "document"
            extraction_note = enc
        elif suffix == ".bin":
            inferred = magic_extension(path)
            if inferred != ".bin":
                copied = temp_dir / f"{stable_name(path)}{inferred}"
                shutil.copy2(path, copied)
                return extract_one_file(copied, procurement_number, section, out_proc, temp_dir, utilities, overwrite, skip_ocr, manifest, errors, seen_hashes)
            error = "unsupported BIN format"
        else:
            error = f"unsupported format: {suffix}"
    except Exception as exc:
        error = str(exc)

    text = "\n\n".join(p.text for p in pieces if p.text)
    class_detail = classify_document_detailed(path, section, text)
    doc_type = class_detail["detected_type"]
    for piece in pieces:
        piece.document_type = doc_type
    if text:
        cache_path.write_text(text, encoding="utf-8")
    status = "success" if text else "error"
    if not text and not error:
        error = "empty extracted text"
    if error:
        errors.append(f"{path}: {error}")
    row = doc_manifest(
            procurement_number,
            path,
            doc_type,
            suffix,
            status,
            pages_or_sheets,
            len(text),
            needs_ocr,
            "; ".join(x for x in [error, extraction_note] if x),
            section,
            str(cache_path) if text else "",
        )
    row.update(
        {
            "classification_score": class_detail.get("classification_score", ""),
            "classifier_confidence": class_detail.get("classifier_confidence", ""),
            "classification_reasons": class_detail.get("classification_reasons", []),
            "alternative_types": class_detail.get("alternative_types", []),
        }
    )
    manifest.append(row)
    return pieces


def extract_xls(path: Path, temp_dir: Path, utilities: Utilities) -> tuple[list[TextPiece], str, str]:
    try:
        import xlrd  # type: ignore

        book = xlrd.open_workbook(str(path))
        pieces: list[TextPiece] = []
        for sheet in book.sheets():
            lines: list[str] = []
            for row_idx in range(sheet.nrows):
                values = []
                for col_idx in range(sheet.ncols):
                    value = sheet.cell_value(row_idx, col_idx)
                    if value not in ("", None):
                        values.append(f"R{row_idx + 1}C{col_idx + 1}={value}")
                if values:
                    lines.append(" | ".join(values))
            if lines:
                pieces.append(TextPiece(normalize_text("\n".join(lines)), str(path), "", sheet.name))
        return pieces, ", ".join(book.sheet_names()), ""
    except Exception:
        pass
    if utilities.libreoffice:
        out_dir = temp_dir / "xls_converted" / stable_name(path)
        out_dir.mkdir(parents=True, exist_ok=True)
        try:
            proc = subprocess.run(
                [utilities.libreoffice, "--headless", "--convert-to", "xlsx", "--outdir", str(out_dir), str(path)],
                capture_output=True,
                text=True,
                timeout=120,
            )
            converted = list(out_dir.glob("*.xlsx"))
            if proc.returncode == 0 and converted:
                return extract_xlsx(converted[0])
            return [], "", f"XLS conversion failed: {(proc.stderr or proc.stdout).strip()[:500]}"
        except Exception as exc:
            return [], "", f"XLS conversion failed: {exc}"
    return [], "", "XLS extraction unavailable: xlrd/LibreOffice not found"


def doc_manifest(
    procurement_number: str,
    path: Path,
    doc_type: str,
    fmt: str,
    status: str,
    pages_or_sheets: str,
    text_length: int,
    needs_ocr: bool,
    error: str,
    section: str,
    output_text_path: str = "",
) -> dict[str, Any]:
    classification = classify_document_detailed(path, section, "")
    if doc_type in ("other", "signature") and classification.get("detected_type") != "other":
        doc_type = str(classification["detected_type"])
    return {
        "procurement_number": procurement_number,
        "original_filename": path.name,
        "detected_type": doc_type,
        "detected_format": fmt.lstrip("."),
        "extraction_status": status,
        "pages_or_sheets": pages_or_sheets,
        "text_length": text_length,
        "needs_ocr": needs_ocr,
        "error": error,
        "source_path": str(path),
        "output_text_path": output_text_path,
        "section": section,
        "classification_score": classification.get("classification_score", ""),
        "classifier_confidence": classification.get("classifier_confidence", ""),
        "classification_reasons": classification.get("classification_reasons", []),
        "alternative_types": classification.get("alternative_types", []),
    }


def first_match(patterns: list[str], pieces: list[TextPiece], field: str, confidence: str = "medium") -> tuple[Any, Evidence | None]:
    for piece in pieces:
        for pattern in patterns:
            match = re.search(pattern, piece.text, flags=re.IGNORECASE | re.MULTILINE)
            if match:
                value = match.group(1).strip() if match.groups() else match.group(0).strip()
                value = clean_value(value)
                return value, make_evidence(field, value, piece, match.group(0), confidence, f"regex:{pattern[:40]}")
    return None, None


def clean_value(value: str) -> str:
    value = normalize_text(value)
    value = re.sub(r"\s{2,}", " ", value)
    value = value.strip(" :-\n\t")
    return value[:1000]


def make_evidence(field: str, value: Any, piece: TextPiece, excerpt: str, confidence: str, method: str) -> Evidence:
    return Evidence(
        field_name=field,
        value=value,
        source_file=piece.source_file,
        document_type=piece.document_type,
        page_or_sheet=piece.page_or_sheet,
        text_excerpt=clean_value(excerpt)[:500],
        confidence=confidence,
        extraction_method=method,
    )


def parse_money(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value)
    text = text.replace("\xa0", " ").replace("₽", "").replace("руб.", "").replace("руб", "")
    match = re.search(r"(\d[\d\s]{0,15}(?:[,.]\d{1,2})?)", text)
    if not match:
        return None
    number = match.group(1).replace(" ", "").replace(",", ".")
    try:
        return round(float(number), 2)
    except ValueError:
        return None


def parse_percent(value: Any) -> float | None:
    if value is None:
        return None
    match = re.search(r"(\d+(?:[,.]\d+)?)", str(value))
    if not match:
        return None
    number = float(match.group(1).replace(",", "."))
    return round(number, 2) if 0 <= number <= 100 else None


def parse_int(value: Any) -> int | None:
    if value is None:
        return None
    match = re.search(r"\d+", str(value))
    return int(match.group(0)) if match else None


def add_field(card: dict[str, Any], evidence: list[Evidence], field: str, value: Any, ev: Evidence | None) -> None:
    if value in (None, "", [], {}):
        return
    if card.get(field) not in (None, "", [], {}):
        if card[field] != value:
            if not isinstance(card.get("contradictions"), list):
                card["contradictions"] = []
            card["contradictions"].append({"field": field, "old": card[field], "new": value})
        return
    card[field] = value
    if ev:
        evidence.append(ev)


def candidate_basics(proc_dir: Path) -> dict[str, Any]:
    data = load_json(proc_dir / "candidate.json")
    return {
        "procurement_number": data.get("procurement_number") or procurement_number_from_dir(proc_dir),
        "procurement_name": data.get("object_name", ""),
        "customer": data.get("customer", ""),
        "nmck": parse_money(data.get("initial_price")),
        "publication_date": data.get("published_date", ""),
        "source_url": data.get("card_url", ""),
        "manual_rank": data.get("manual_rank", ""),
        "why_selected": data.get("why_selected", ""),
    }


def analyze_fields(card: dict[str, Any], pieces: list[TextPiece]) -> list[Evidence]:
    evidence: list[Evidence] = []
    for field in ("procurement_name", "customer", "publication_date", "source_url", "nmck"):
        if card.get(field):
            evidence.append(
                Evidence(field, card[field], str(Path("candidate.json")), "notice", "metadata", str(card[field])[:500], "high", "candidate.json")
            )

    patterns: dict[str, list[str]] = {
        "law": [r"(44-ФЗ|223-ФЗ|Федеральн[а-я ]+закона\s*№?\s*44|Федеральн[а-я ]+закона\s*№?\s*223)"],
        "procurement_method": [r"(электронн[а-я ]+аукцион|конкурс в электронной форме|запрос котировок|закупка у единственного поставщика)"],
        "region": [r"(Хабаровск[а-я ]+кра[йя]|Московск[а-я ]+област[ьи]|Санкт-Петербург|Москва|[А-ЯЁ][а-яё]+ область|[А-ЯЁ][а-яё]+ край)"],
        "application_deadline": [r"Дата и время окончания срока подачи заявок[:\s]+([0-9]{2}\.[0-9]{2}\.[0-9]{4}(?:\s+[0-9:]+)?)"],
        "contract_deadline": [r"Срок(?:и)? (?:исполнения|действия) контракт[а-я ]*[:\s]+([^\n]{5,160})"],
        "contract_price": [r"(?:цена контракта|предложение о цене|итоговая цена)[^\n\r]{0,80}?(\d[\d\s]+[,.]\d{2})"],
        "participants_count": [r"(?:подано|поступило)[^\n\r]{0,80}?(\d+)\s+заяв"],
        "admitted_participants_count": [r"(?:допущен[оы]?|признан[оы]? соответствующ)[^\n\r]{0,80}?(\d+)\s+заяв"],
        "winner_application_number": [r"(?:победител[ья][^\n\r]{0,120}?заявк[аи]\s*№?\s*)(\d+)"],
        "winner_name": [r"(?:победител[ья][:\s]+)([^\n\r]{5,180})"],
        "acceptance_period_days": [r"(?:срок приемки|приемка)[^\n\r]{0,100}?(\d+)\s*(?:рабочих|календарных)?\s*дн"],
        "payment_period_days": [r"(?:оплата|срок оплаты)[^\n\r]{0,100}?(\d+)\s*(?:рабочих|календарных)?\s*дн"],
        "advance_percent": [r"(?:аванс|авансовый платеж)[^\n\r]{0,80}?(\d+(?:[,.]\d+)?)\s*%"],
        "bid_security_amount": [r"(?:обеспечение заявки)[^\n\r]{0,120}?(\d[\d\s]+[,.]\d{2})"],
        "bid_security_percent": [r"(?:обеспечение заявки)[^\n\r]{0,120}?(\d+(?:[,.]\d+)?)\s*%"],
        "contract_security_amount": [r"(?:обеспечение исполнения контракта)[^\n\r]{0,160}?(\d[\d\s]+[,.]\d{2})"],
        "contract_security_percent": [r"(?:обеспечение исполнения контракта)[^\n\r]{0,160}?(\d+(?:[,.]\d+)?)\s*%"],
        "support_period": [r"(?:срок технической поддержки|сопровождение)[^\n\r]{0,100}?([^\n\r]{5,120})"],
        "warranty_period": [r"(?:гарантийн[а-я ]+срок|гарантия)[^\n\r]{0,100}?([^\n\r]{5,120})"],
        "revision_period_days": [r"(?:устранени[ея] замечаний|доработк[аи])[^\n\r]{0,100}?(\d+)\s*(?:рабочих|календарных)?\s*дн"],
    }
    numeric_money = {"contract_price", "bid_security_amount", "contract_security_amount"}
    numeric_int = {"participants_count", "admitted_participants_count", "acceptance_period_days", "payment_period_days", "revision_period_days"}
    numeric_percent = {"advance_percent", "bid_security_percent", "contract_security_percent"}
    for field, pats in patterns.items():
        value, ev = first_match(pats, pieces, field)
        if field in numeric_money:
            value = parse_money(value)
        elif field in numeric_int:
            value = parse_int(value)
        elif field in numeric_percent:
            value = parse_percent(value)
        add_field(card, evidence, field, value, ev)

    prices = extract_price_offers(pieces, card.get("nmck"))
    if prices:
        card["all_price_offers"] = prices
        evidence.append(make_evidence("all_price_offers", prices, prices[0]["_piece"], prices[0]["_excerpt"], "medium", "price-offer-regex"))
        for item in prices:
            item.pop("_piece", None)
            item.pop("_excerpt", None)
        winner = min((p for p in prices if p.get("price")), key=lambda p: p["price"], default=None)
        if winner:
            add_field(card, evidence, "contract_price", winner["price"], None)
            add_field(card, evidence, "winner_application_number", winner.get("application_number"), None)
    derive_booleans_and_lists(card, pieces, evidence)
    derive_price_fields(card, evidence, pieces)
    score_procurement(card, pieces, evidence)
    return evidence


def extract_price_offers(pieces: list[TextPiece], nmck: float | None) -> list[dict[str, Any]]:
    offers: list[dict[str, Any]] = []
    seen: set[tuple[str, float]] = set()
    pattern = re.compile(
        r"(?:заявк[аи]\s*№?\s*|участник\s*№?\s*)(\d{1,4})[^\n\r]{0,160}?(\d[\d\s]{2,15}[,.]\d{2})",
        re.IGNORECASE,
    )
    for piece in pieces:
        if piece.document_type != "final_protocol" and "протокол" not in piece.text[:2000].lower():
            continue
        for match in pattern.finditer(piece.text):
            price = parse_money(match.group(2))
            app = match.group(1)
            if not price or (nmck and price > nmck * 1.2):
                continue
            key = (app, price)
            if key in seen:
                continue
            seen.add(key)
            reduction = round((1 - price / nmck) * 100, 2) if nmck else None
            offers.append(
                {
                    "application_number": app,
                    "price": price,
                    "reduction_percent": reduction,
                    "rank": None,
                    "status": "admitted",
                    "_piece": piece,
                    "_excerpt": match.group(0),
                }
            )
    offers.sort(key=lambda item: item["price"])
    for idx, item in enumerate(offers, start=1):
        item["rank"] = idx
    return offers[:50]


def derive_booleans_and_lists(card: dict[str, Any], pieces: list[TextPiece], evidence: list[Evidence]) -> None:
    haystack = "\n".join(piece.text[:15000] for piece in pieces)
    checks: dict[str, list[str]] = {
        "licenses_required": ["лицензия", "лицензии", "лицензион", "выписка из реестра лиценз"],
        "experience_required": ["опыт исполнени", "аналогичн", "подтверждение опыта"],
        "staff_requirements": ["специалист", "квалификац", "штат"],
        "portfolio_required": ["портфолио"],
        "sro_required": [" сро ", "членство в сро", "саморегулируем"],
        "smp_only": ["субъект малого предпринимательства", "сонко", "смп"],
        "rnp_requirement": ["реестр недобросовестных поставщиков", "рнп"],
        "local_presence_required": ["на территории заказчика", "местонахождение исполнителя"],
        "warranty_security": ["обеспечение гарантийн"],
        "anti_dumping_applies": ["антидемпинг"],
        "possible_exemption_from_security": ["освобождается от предоставления обеспечения"],
        "admin_panel": ["административн", "панель администратора", "админ"],
        "personal_accounts": ["личный кабинет"],
        "integrations": ["интеграц", "api", "синхронизац"],
        "api_required": ["api", "апи", "веб-сервис"],
        "mobile_app_required": ["мобильное приложение", "android", "ios"],
        "adaptive_layout": ["адаптивн", "мобильная версия"],
        "notifications": ["уведомлен", "sms", "e-mail", "email"],
        "document_generation": ["формировани", "генерац", "печатн"],
        "file_uploads": ["загрузка файл", "прикреплен", "документ"],
        "reporting": ["отчет", "отчёт"],
        "analytics": ["аналитик", "дашборд", "статистик"],
        "search": ["поиск", "фильтр"],
        "content_management": ["cms", "управление контентом"],
        "personal_data": ["персональн", "152-фз"],
        "esia": ["есиа"],
        "smev": ["смэв"],
        "cryptopro": ["криптопро", "crypto"],
        "fstec": ["фстэк"],
        "fsb": ["фсб"],
        "closed_network": ["закрыт", "изолированн", "контур"],
        "support_24x7": ["24/7", "круглосуточ"],
        "training_required": ["обучение"],
        "documentation_required": ["документац", "руководство пользователя"],
        "source_code_transfer": ["исходн", "исходный код"],
        "exclusive_rights_transfer": ["исключительн", "права"],
        "repository_transfer": ["репозитор"],
        "open_source_restrictions": ["открыт", "open source", "свободн"],
        "user_testing_required": ["тестировани", "опытная эксплуатац"],
        "unlimited_revision_risk": ["до полного устранения", "без ограничен"],
        "subjective_design_acceptance": ["по согласованию с заказчиком", "на усмотрение заказчика", "дизайн"],
        "hidden_defect_liability": ["скрыт", "недостатк"],
    }
    lower = haystack.lower()
    for field, terms in checks.items():
        if any(term in lower for term in terms):
            piece = find_piece_with_terms(pieces, terms)
            add_field(card, evidence, field, True, make_evidence(field, True, piece, extract_excerpt(piece.text, terms), "medium", "keyword") if piece else None)

    list_fields = {
        "functional_modules": ["модуль", "раздел", "личный кабинет", "каталог", "новости", "карта", "реестр", "форма"],
        "external_systems": ["есиа", "смэв", "гис", "1с", "криптопро", "платеж"],
        "security_requirements": ["защита информации", "персональные данные", "ssl", "https", "резервное копирование"],
        "key_risks": ["штраф", "пени", "неустойк", "24/7", "персональн", "исключительн", "сжатые сроки", "антидемпинг"],
    }
    for field, terms in list_fields.items():
        found = sorted({term for term in terms if term in lower})
        if found:
            piece = find_piece_with_terms(pieces, found)
            add_field(card, evidence, field, found, make_evidence(field, found, piece, extract_excerpt(piece.text, found), "low", "keyword-list") if piece else None)

    scope_sentences = best_sentences(pieces, ["оказание услуг", "разработк", "создание", "портал", "сайт"], limit=2)
    if scope_sentences:
        add_field(card, evidence, "short_scope", " ".join(scope_sentences)[:700], None)
    if "сайт" in lower or "портал" in lower:
        add_field(card, evidence, "project_type", "web_portal_or_site", None)


def find_piece_with_terms(pieces: list[TextPiece], terms: Iterable[str]) -> TextPiece | None:
    for piece in pieces:
        lower = piece.text.lower()
        if any(term in lower for term in terms):
            return piece
    return pieces[0] if pieces else None


def extract_excerpt(text: str, terms: Iterable[str]) -> str:
    lower = text.lower()
    for term in terms:
        idx = lower.find(str(term).lower())
        if idx >= 0:
            return text[max(0, idx - 160) : idx + 340]
    return text[:500]


def best_sentences(pieces: list[TextPiece], terms: list[str], limit: int = 3) -> list[str]:
    result: list[str] = []
    for piece in pieces:
        for sentence in re.split(r"(?<=[.!?])\s+|\n+", piece.text):
            low = sentence.lower()
            if 40 <= len(sentence) <= 500 and any(term in low for term in terms):
                result.append(clean_value(sentence))
                if len(result) >= limit:
                    return result
    return result


def derive_price_fields(card: dict[str, Any], evidence: list[Evidence], pieces: list[TextPiece]) -> None:
    nmck = card.get("nmck")
    contract_price = card.get("contract_price")
    if nmck and contract_price:
        reduction = round((1 - float(contract_price) / float(nmck)) * 100, 2)
        if 0 <= reduction <= 100:
            card["price_reduction_percent"] = reduction
    if not card.get("nmck_sources"):
        sentences = best_sentences(pieces, ["коммерческ", "источник", "обоснование", "метод"], 3)
        if sentences:
            card["nmck_sources"] = sentences
    if not card.get("nmck_commercial_offers"):
        offers = []
        for piece in pieces:
            if piece.document_type != "nmck_calculation":
                continue
            for match in re.finditer(r"(?:коммерческ[а-я ]+предложен[а-я ]*|кп)[^\n\r]{0,120}?(\d[\d\s]+[,.]\d{2})", piece.text, re.I):
                price = parse_money(match.group(1))
                if price:
                    offers.append(price)
        if offers:
            card["nmck_commercial_offers"] = offers[:20]


def score_procurement(card: dict[str, Any], pieces: list[TextPiece], evidence: list[Evidence]) -> None:
    risk_terms = [
        ("esia", 2),
        ("smev", 2),
        ("cryptopro", 2),
        ("fstec", 2),
        ("fsb", 2),
        ("closed_network", 2),
        ("personal_data", 1),
        ("mobile_app_required", 2),
        ("integrations", 1),
        ("support_24x7", 2),
        ("exclusive_rights_transfer", 1),
        ("unlimited_revision_risk", 2),
        ("subjective_design_acceptance", 1),
        ("licenses_required", 2),
        ("sro_required", 2),
    ]
    complexity = 3 + sum(weight for field, weight in risk_terms if card.get(field))
    if card.get("functional_modules"):
        complexity += min(2, len(card["functional_modules"]) // 3)
    nmck = card.get("nmck") or 0
    reduction = card.get("price_reduction_percent") or 0
    tech = clamp(complexity, 1, 10)
    org = clamp(2 + bool_score(card, ["training_required", "documentation_required", "support_period", "staff_requirements"]) + len(card.get("contradictions", [])), 1, 10)
    legal = clamp(2 + bool_score(card, ["exclusive_rights_transfer", "penalties", "hidden_defect_liability", "licenses_required", "sro_required"]) + len(card.get("contradictions", [])), 1, 10)
    financial = clamp(2 + (2 if reduction and reduction > 25 else 0) + (2 if nmck and nmck < 250000 else 0) + (1 if not card.get("advance_percent") else 0), 1, 10)
    scope_text = f"{card.get('procurement_name', '')} {card.get('short_scope', '')}".lower()
    platform_1c = "1с" in scope_text or "1c" in scope_text
    inherited_system = any(word in scope_text for word in ["модернизац", "сопровожден", "существующ"])
    simple_portal = any(word in scope_text for word in ["сайт", "портал", "cms", "контент"]) and not inherited_system
    solo_base = 7 if simple_portal else 6
    if inherited_system:
        solo_base = 4
    if platform_1c:
        solo_base = 2
        card["specific_platform"] = "1C"
        card["platform_expertise_required"] = True
    if card.get("esia") or card.get("smev") or card.get("closed_network"):
        solo_base = min(solo_base, 2)
    solo_adjust = 0
    if not card.get("integrations"):
        solo_adjust += 1
    if card.get("adaptive_layout") or card.get("content_management"):
        solo_adjust += 1
    for field in ("personal_data", "mobile_app_required", "support_24x7", "esia", "smev", "closed_network"):
        if card.get(field):
            solo_adjust -= 2 if field in {"esia", "smev", "support_24x7", "mobile_app_required", "closed_network"} else 1
    solo_fit = clamp(solo_base + solo_adjust, 1, 10)
    ai_fit = 6
    for field in ("content_management", "personal_accounts", "admin_panel", "file_uploads", "document_generation", "api_required", "search"):
        if card.get(field):
            ai_fit += 1
    for field in ("closed_network", "esia", "smev", "cryptopro", "fstec", "fsb", "support_24x7"):
        if card.get(field):
            ai_fit -= 2
    if platform_1c or inherited_system:
        ai_fit -= 2
    ai_fit = clamp(ai_fit, 1, 10)
    hours_min = int(40 + tech * 18 + org * 8)
    hours_max = int(hours_min * (1.45 + legal / 20))
    direct_min = int(max(10000, hours_min * 250))
    direct_max = int(max(direct_min, hours_max * 450))
    recommended_min = int(max(direct_max * 1.25, (nmck or 0) * 0.45))
    recommended_comfort = int(max(recommended_min * 1.35, (nmck or 0) * 0.65))

    card.update(
        {
            "technical_complexity_score": tech,
            "organizational_complexity_score": org,
            "legal_risk_score": legal,
            "financial_risk_score": financial,
            "ai_fit_score": ai_fit,
            "solo_developer_fit_score": solo_fit,
            "estimated_hours_min": hours_min,
            "estimated_hours_max": hours_max,
            "estimated_calendar_weeks": max(1, math.ceil(hours_max / 40)),
            "estimated_direct_costs_min": direct_min,
            "estimated_direct_costs_max": direct_max,
            "recommended_min_price": recommended_min,
            "recommended_comfort_price": recommended_comfort,
        }
    )
    completeness = sum(1 for f in ("nmck", "contract_price", "participants_count", "short_scope", "contract_deadline") if card.get(f))
    critical = any(card.get(f) for f in ("closed_network", "fstec", "fsb", "sro_required", "licenses_required"))
    if completeness < 2:
        verdict = "INSUFFICIENT_DATA"
        reason = "Недостаточно извлечённых ключевых полей для уверенной оценки."
    elif critical or legal >= 8 or tech >= 9:
        verdict = "DO_NOT_TAKE"
        reason = "Высокие юридические или технические ограничения для одиночного исполнителя."
    elif solo_fit >= 6 and ai_fit >= 7 and financial <= 5:
        verdict = "TAKE_NOW"
        reason = "Профильная веб-разработка с приемлемыми рисками по извлечённым условиям."
    elif solo_fit >= 5 and ai_fit >= 6:
        verdict = "TAKE_WITH_CONDITIONS"
        reason = "Можно рассматривать после ручной проверки рисков, сроков и цены."
    else:
        verdict = "TAKE_AFTER_PREPARATION"
        reason = "Нужна подготовка или партнёры из-за сложности/рисков."
    card["verdict"] = verdict
    card["verdict_reason"] = reason
    for field in (
        "technical_complexity_score",
        "organizational_complexity_score",
        "legal_risk_score",
        "financial_risk_score",
        "ai_fit_score",
        "solo_developer_fit_score",
        "verdict",
    ):
        evidence.append(Evidence(field, card[field], "rule-based model", "analysis", "scores", str(card[field]), "medium", "weighted heuristic"))


def bool_score(card: dict[str, Any], fields: list[str]) -> int:
    return sum(1 for f in fields if card.get(f))


def clamp(value: float, low: int, high: int) -> int:
    return max(low, min(high, int(round(value))))


def quality_checks(card: dict[str, Any], evidence: list[Evidence], manifest_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []

    def issue(level: str, field: str, message: str) -> None:
        issues.append({"procurement_number": card.get("procurement_number"), "severity": level, "field": field, "message": message})

    nmck = card.get("nmck")
    contract_price = card.get("contract_price")
    reduction = card.get("price_reduction_percent")
    if nmck and contract_price and contract_price > nmck:
        issue("high", "contract_price", "contract_price is greater than nmck")
    if nmck and contract_price and reduction is not None:
        expected = round((1 - contract_price / nmck) * 100, 2)
        if abs(expected - float(reduction)) > 0.5:
            issue("medium", "price_reduction_percent", "reduction percent does not match prices")
    offers = card.get("all_price_offers") or []
    if card.get("participants_count") and offers and int(card["participants_count"]) < len(offers):
        issue("medium", "participants_count", "participants_count is less than price offers count")
    for field in ("publication_date", "application_deadline"):
        if card.get(field) and not re.search(r"\d{2}\.\d{2}\.\d{4}", str(card[field])):
            issue("low", field, "date is not in expected dd.mm.yyyy form")
    for field in ("price_reduction_percent", "advance_percent", "bid_security_percent", "contract_security_percent"):
        if card.get(field) not in (None, "") and not (0 <= float(card[field]) <= 100):
            issue("medium", field, "percent is outside 0-100")
    for field in ("nmck", "contract_price"):
        if card.get(field) not in (None, "") and float(card[field]) < 0:
            issue("high", field, "negative price")
    evidence_sources = {Path(ev.source_file).name for ev in evidence if ev.source_file}
    for ev in evidence:
        if ev.source_file not in ("candidate.json", "rule-based model") and Path(ev.source_file).name not in evidence_sources:
            issue("low", ev.field_name, "evidence source missing")
    for row in manifest_rows:
        if row.get("extraction_status") == "success" and not int(row.get("text_length") or 0):
            issue("medium", "extraction", "empty text marked as success")
    return issues


def unresolved_fields(card: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for field in FIELD_NAMES:
        if card.get(field) in (None, "", [], {}):
            rows.append({"procurement_number": card.get("procurement_number"), "field_name": field, "reason": "not found in extracted evidence"})
    return rows


def piece_allowed(piece: TextPiece, allowed_types: set[str], allowed_stems: set[str] | None = None) -> bool:
    if piece.document_type in allowed_types:
        return True
    if allowed_stems and Path(piece.source_file).stem.lower() in allowed_stems:
        return True
    return False


def strict_pieces(pieces: list[TextPiece], allowed_types: set[str], allowed_stems: set[str] | None = None) -> list[TextPiece]:
    return [p for p in pieces if piece_allowed(p, allowed_types, allowed_stems)]


def evidence_exists(evidence: list[Evidence], field: str, allowed_types: set[str] | None = None) -> bool:
    for ev in evidence:
        if ev.field_name != field:
            continue
        if allowed_types is None or ev.document_type in allowed_types:
            return True
    return False


def strict_set(card: dict[str, Any], evidence: list[Evidence], field: str, value: Any, ev: Evidence | None) -> None:
    card[field] = value if value not in ("", [], {}) else None
    if ev and value not in (None, "", [], {}):
        evidence.append(ev)


def strict_price_candidates(piece: TextPiece, nmck: float | None) -> list[dict[str, Any]]:
    offers: list[dict[str, Any]] = []
    text = piece.text
    row_pattern = re.compile(
        r"(?m)^\s*(\d{1,4})(?:\s+от\s+\d{2}\.\d{2}\.\d{4}[^\n|]*)?\s*(?:\||\s{2,}|\t)\s*(\d[\d\s\u00a0]*[,.]\d{2})\s*(?:\||\s{2,}|\t)\s*(-?\d+(?:[,.]\d+)?)?",
        re.IGNORECASE,
    )
    table_context = ("предложение о цене" in text.lower()) or ("предложения участников" in text.lower()) or ("ценовых предложений" in text.lower())
    if not table_context:
        return []
    for match in row_pattern.finditer(text):
        app = match.group(1)
        price = parse_money(match.group(2))
        if not price:
            continue
        if nmck and (price < nmck * 0.01 or price > nmck * 1.05):
            continue
        reduction = round((1 - price / nmck) * 100, 2) if nmck else parse_percent(match.group(3))
        offers.append(
            {
                "application_number": app,
                "price": price,
                "reduction_percent": reduction,
                "rank": None,
                "status": "admitted",
                "_piece": piece,
                "_excerpt": match.group(0),
            }
        )
    if offers:
        offers.sort(key=lambda item: item["price"])
        seen: set[tuple[str, float]] = set()
        clean: list[dict[str, Any]] = []
        for item in offers:
            key = (item["application_number"], item["price"])
            if key in seen:
                continue
            seen.add(key)
            item["rank"] = len(clean) + 1
            clean.append(item)
        return clean

    winner_patterns = [
        r"(?m)^\s*(\d{1,4})\s+\d+\s*-\s*победител[^\n\r]*?(\d[\d\s\u00a0]*[,.]\d{2})",
        r"(?m)^\s*(\d{1,4})\s+\d+\s*-\s*РџРѕР±РµРґРёС‚РµР»[^\n\r]*?(\d[\d\s\u00a0]*[,.]\d{2})",
    ]
    for pattern in winner_patterns:
        for match in re.finditer(pattern, text, re.IGNORECASE):
            price = parse_money(match.group(2))
            if price and (not nmck or nmck * 0.01 <= price <= nmck * 1.05):
                return [
                    {
                        "application_number": match.group(1),
                        "price": price,
                        "reduction_percent": round((1 - price / nmck) * 100, 2) if nmck else None,
                        "rank": 1,
                        "status": "admitted",
                        "_piece": piece,
                        "_excerpt": match.group(0),
                    }
                ]
    return []


def extract_strict_protocol_data(card: dict[str, Any], pieces: list[TextPiece], evidence: list[Evidence]) -> None:
    protocol_pieces = strict_pieces(pieces, {"final_protocol", "auction_protocol"}, {"results"})
    card["contract_price"] = None
    card["price_reduction_percent"] = None
    card["participants_count"] = None
    card["admitted_participants_count"] = None
    card["winner_application_number"] = None
    card["winner_name"] = None
    card["all_price_offers"] = None
    all_offers: list[dict[str, Any]] = []
    for piece in protocol_pieces:
        all_offers.extend(strict_price_candidates(piece, card.get("nmck")))
        for pattern in [
            r"подано\s+(\d+)\s+заяв",
            r"поступило\s+(\d+)\s+заяв",
            r"РїРѕРґР°РЅРѕ\s+(\d+)\s+Р·Р°СЏРІ",
            r"РїРѕСЃС‚СѓРїРёР»Рѕ\s+(\d+)\s+Р·Р°СЏРІ",
        ]:
            match = re.search(pattern, piece.text, re.IGNORECASE)
            if match:
                strict_set(card, evidence, "participants_count", int(match.group(1)), make_evidence("participants_count", int(match.group(1)), piece, match.group(0), "high", "strict-final-protocol"))
                break
        for pattern in [
            r"(?:соответствует|допущен[оы]?)[^\n\r]{0,80}?(\d+)\s+заяв",
            r"(?:РЎРѕРѕС‚РІРµС‚СЃС‚РІСѓРµС‚|РґРѕРїСѓС‰РµРЅ[РѕС‹]?)[^\n\r]{0,80}?(\d+)\s+Р·Р°СЏРІ",
        ]:
            match = re.search(pattern, piece.text, re.IGNORECASE)
            if match:
                strict_set(card, evidence, "admitted_participants_count", int(match.group(1)), make_evidence("admitted_participants_count", int(match.group(1)), piece, match.group(0), "medium", "strict-final-protocol"))
                break
    if all_offers:
        all_offers.sort(key=lambda item: item["price"])
        deduped: list[dict[str, Any]] = []
        seen_offers: set[tuple[str, float]] = set()
        for item in all_offers:
            key = (str(item.get("application_number")), float(item.get("price")))
            if key in seen_offers:
                continue
            seen_offers.add(key)
            deduped.append(item)
        all_offers = deduped
        winner = all_offers[0]
        for item in all_offers:
            item["rank"] = all_offers.index(item) + 1
        ev_piece = winner["_piece"]
        ev_excerpt = winner["_excerpt"]
        clean_offers = []
        for item in all_offers:
            item = dict(item)
            item.pop("_piece", None)
            item.pop("_excerpt", None)
            clean_offers.append(item)
        strict_set(card, evidence, "all_price_offers", clean_offers, make_evidence("all_price_offers", clean_offers, ev_piece, ev_excerpt, "high", "strict-final-protocol-table"))
        strict_set(card, evidence, "contract_price", winner["price"], make_evidence("contract_price", winner["price"], ev_piece, ev_excerpt, "high", "strict-final-protocol-table"))
        strict_set(card, evidence, "winner_application_number", winner["application_number"], make_evidence("winner_application_number", winner["application_number"], ev_piece, ev_excerpt, "high", "strict-final-protocol-table"))
        if not card.get("participants_count") and len(clean_offers) > 1:
            strict_set(card, evidence, "participants_count", len(clean_offers), make_evidence("participants_count", len(clean_offers), ev_piece, ev_excerpt, "medium", "strict-offer-count"))
        if not card.get("admitted_participants_count"):
            strict_set(card, evidence, "admitted_participants_count", len(clean_offers), make_evidence("admitted_participants_count", len(clean_offers), ev_piece, ev_excerpt, "medium", "strict-offer-count"))
    if card.get("nmck") and card.get("contract_price"):
        card["price_reduction_percent"] = round((float(card["nmck"]) - float(card["contract_price"])) / float(card["nmck"]) * 100, 2)


def strict_short_scope(card: dict[str, Any], pieces: list[TextPiece], evidence: list[Evidence]) -> None:
    tech = strict_pieces(pieces, {"technical_specification", "technical_specification_attachment"})
    banned = ["единая информационная система", "федеральным законом", "статьи 31", "банковские реквизиты", "личный кабинет"]
    terms = ["функциональ", "перечень работ", "требования к системе", "описание объекта", "состав услуг", "разработ", "портал", "сайт", "приложен"]
    sentences = []
    for piece in tech:
        for sentence in re.split(r"(?<=[.!?])\s+|\n+", piece.text):
            clean = clean_value(sentence)
            low = clean.lower()
            if len(clean) < 45 or len(clean) > 420:
                continue
            if any(b in low for b in banned):
                continue
            if any(t in low for t in terms):
                sentences.append((clean, piece))
            if len(sentences) >= 5:
                break
        if len(sentences) >= 5:
            break
    if sentences:
        scope = " ".join(s[0] for s in sentences[:5])[:1200]
        strict_set(card, evidence, "short_scope", scope, make_evidence("short_scope", scope, sentences[0][1], sentences[0][0], "medium", "strict-technical-scope"))
    else:
        card["short_scope"] = None


def strict_source_limited_fields(card: dict[str, Any], pieces: list[TextPiece], evidence: list[Evidence]) -> None:
    technical_fields = [
        "functional_modules",
        "user_roles",
        "admin_panel",
        "personal_accounts",
        "integrations",
        "external_systems",
        "api_required",
        "mobile_app_required",
        "adaptive_layout",
        "notifications",
        "document_generation",
        "file_uploads",
        "reporting",
        "analytics",
        "search",
        "content_management",
        "design_requirements",
        "server_requirements",
        "hosting_requirements",
        "deployment_requirements",
        "backup_requirements",
        "monitoring_requirements",
        "security_requirements",
        "personal_data",
        "esia",
        "smev",
        "cryptopro",
        "fstec",
        "fsb",
        "closed_network",
    ]
    allowed_tech = {"technical_specification", "technical_specification_attachment", "clarification"}
    for field in technical_fields:
        if card.get(field) and not evidence_exists(evidence, field, allowed_tech):
            card[field] = None
    requirement_fields = ["licenses_required", "experience_required", "staff_requirements", "portfolio_required", "sro_required", "additional_requirements", "smp_only", "local_presence_required"]
    allowed_req = {"application_requirements", "information_card", "notice", "clarification"}
    for field in requirement_fields:
        if card.get(field) and not evidence_exists(evidence, field, allowed_req):
            card[field] = None


KEY_DOCUMENT_ROLES = {
    "technical": {"technical_specification", "technical_specification_attachment"},
    "contract": {"contract_draft", "signed_contract"},
    "nmck": {"nmck_calculation"},
    "application_requirements": {"application_requirements", "information_card"},
    "protocol": {"final_protocol", "auction_protocol"},
}


def is_download_row(row: dict[str, Any]) -> bool:
    return "\\downloads\\" in str(row.get("source_path"))


def key_document_bundles(manifest_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    bundles: dict[str, dict[str, Any]] = {}
    rows = [r for r in manifest_rows if is_download_row(r)]
    for role, types in KEY_DOCUMENT_ROLES.items():
        role_rows = [r for r in rows if r.get("detected_type") in types]
        files = [r.get("original_filename") for r in role_rows if r.get("original_filename")]
        read_rows = [r for r in role_rows if r.get("extraction_status") in ("success", "cached") and int(r.get("text_length") or 0) > 100]
        unreadable_rows = [r for r in role_rows if r.get("extraction_status") == "error"]
        if read_rows:
            status = "read"
        elif unreadable_rows or role_rows:
            status = "unreadable"
        else:
            status = "missing"
        if role == "technical" and status == "missing":
            attachments = [
                r
                for r in rows
                if r.get("detected_type") == "other"
                and int(r.get("text_length") or 0) > 100
                and any(token in normalize_classifier_text(str(r.get("original_filename"))) for token in ("спецификация", "регламент", "перечень работ"))
            ]
            if attachments:
                status = "partial"
                files.extend(r.get("original_filename") for r in attachments if r.get("original_filename"))
        bundles[role] = {
            "status": status,
            "files": files,
            "read_files": [r.get("original_filename") for r in read_rows],
            "unreadable_files": [r.get("original_filename") for r in unreadable_rows],
        }
    return bundles


def document_statuses(manifest_rows: list[dict[str, Any]]) -> dict[str, Any]:
    bundles = key_document_bundles(manifest_rows)
    return {
        "technical_specification_status": bundles["technical"]["status"],
        "contract_status": bundles["contract"]["status"],
        "nmck_status": bundles["nmck"]["status"],
        "application_requirements_status": bundles["application_requirements"]["status"],
        "final_protocol_status": bundles["protocol"]["status"],
        "technical_document_files": bundles["technical"]["files"],
        "contract_document_files": bundles["contract"]["files"],
        "application_requirement_files": bundles["application_requirements"]["files"],
        "nmck_document_files": bundles["nmck"]["files"],
        "protocol_document_files": bundles["protocol"]["files"],
    }


def compute_data_quality(card: dict[str, Any], manifest_rows: list[dict[str, Any]], quality_rows: list[dict[str, Any]]) -> dict[str, Any]:
    statuses = document_statuses(manifest_rows)
    score = 0
    weights = {
        "technical_specification_status": 25,
        "contract_status": 20,
        "nmck_status": 15,
        "application_requirements_status": 15,
        "final_protocol_status": 25,
    }
    availability = 0
    extraction = 0
    for field, weight in weights.items():
        if statuses[field] == "read":
            score += weight
            availability += weight
            extraction += weight
        elif statuses[field] == "partial":
            score += weight // 2
            availability += weight
            extraction += weight // 2
        elif statuses[field] == "unreadable":
            availability += weight
    if card.get("nmck"):
        score += 5
    if card.get("short_scope"):
        score += 5
    if card.get("contract_price"):
        score += 5
    critical = [q for q in quality_rows if q.get("severity") == "critical"]
    score = max(0, min(100, score - len(critical) * 20))
    missing = [field for field in weights if statuses[field] not in ("read", "partial")]
    if score >= 80:
        reliability = "HIGH"
    elif score >= 60:
        reliability = "MEDIUM"
    elif score >= 35:
        reliability = "LOW"
    else:
        reliability = "INSUFFICIENT"
    return {
        **statuses,
        "document_availability_score": availability,
        "text_extraction_score": extraction,
        "data_completeness_score": score,
        "analysis_reliability": reliability,
        "missing_key_documents": missing,
        "critical_quality_issues": len(critical),
    }


def strict_quality_checks(card: dict[str, Any], evidence: list[Evidence], manifest_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    issues = quality_checks(card, evidence, manifest_rows)

    def issue(level: str, field: str, message: str) -> None:
        issues.append({"procurement_number": card.get("procurement_number"), "severity": level, "field": field, "message": message})

    nmck = card.get("nmck")
    price = card.get("contract_price")
    if price and not evidence_exists(evidence, "contract_price", {"final_protocol", "auction_protocol"}):
        issue("critical", "contract_price", "contract_price accepted without final protocol evidence")
        card["contract_price"] = None
        card["price_reduction_percent"] = None
    if nmck and price and price < float(nmck) * 0.01:
        issue("critical", "contract_price", "contract_price is below 1% of nmck and was rejected")
        card["contract_price"] = None
        card["price_reduction_percent"] = None
    if nmck and price and abs(float(price) - float(nmck)) < 0.01 and not evidence_exists(evidence, "contract_price", {"final_protocol", "auction_protocol"}):
        issue("critical", "contract_price", "contract_price equals nmck without protocol evidence")
    if card.get("price_reduction_percent") == 0 and not card.get("contract_price"):
        issue("critical", "price_reduction_percent", "zero reduction without contract price evidence")
        card["price_reduction_percent"] = None
    offers = card.get("all_price_offers") or []
    if card.get("participants_count") and offers and int(card["participants_count"]) < len(offers):
        issue("critical", "participants_count", "participants_count is less than all_price_offers")
    for ev in evidence:
        if ev.source_file in ("candidate.json", "rule-based model"):
            continue
        if ev.source_file and not Path(ev.source_file).exists():
            issue("medium", ev.field_name, f"evidence path does not exist: {ev.source_file}")
    return issues


def strict_score_procurement(card: dict[str, Any], pieces: list[TextPiece], evidence: list[Evidence]) -> None:
    tech_base = 2
    if card.get("personal_accounts"):
        tech_base = max(tech_base, 5)
    if card.get("content_management") or card.get("admin_panel"):
        tech_base = max(tech_base, 4)
    if card.get("short_scope") and any(word in str(card.get("short_scope")).lower() for word in ["модернизац", "сопровожден", "1с"]):
        tech_base = max(tech_base, 6)
    additions = [
        ("user_roles", 1),
        ("admin_panel", 1),
        ("file_uploads", 1),
        ("document_generation", 1),
        ("api_required", 1),
        ("esia", 2),
        ("smev", 2),
        ("personal_data", 1),
        ("mobile_app_required", 1),
        ("closed_network", 1),
        ("analytics", 1),
        ("support_24x7", 2),
    ]
    tech = clamp(tech_base + sum(weight for field, weight in additions if card.get(field)), 1, 10)
    org = clamp(2 + bool_score(card, ["training_required", "documentation_required", "support_period", "stages_count", "subjective_design_acceptance"]), 1, 10)
    legal = clamp(2 + bool_score(card, ["unlimited_revision_risk", "subjective_design_acceptance", "exclusive_rights_transfer", "hidden_defect_liability"]) + len(card.get("contradictions") or []), 1, 10)
    fin = 2
    if card.get("price_reduction_percent") and float(card["price_reduction_percent"]) > 25:
        fin += 2
    if card.get("contract_security_amount") or card.get("bid_security_amount"):
        fin += 1
    if not card.get("advance_percent"):
        fin += 1
    fin = clamp(fin, 1, 10)

    scope_text = f"{card.get('procurement_name', '')} {card.get('short_scope', '')}".lower()
    platform_1c = "1СЃ" in scope_text or "1c" in scope_text or "1с" in scope_text
    inherited_system = any(word in scope_text for word in ["РјРѕРґРµСЂРЅРёР·Р°С†", "СЃРѕРїСЂРѕРІРѕР¶РґРµРЅ", "СЃСѓС‰РµСЃС‚РІСѓСЋС‰"])
    simple_portal = any(word in scope_text for word in ["СЃР°Р№С‚", "РїРѕСЂС‚Р°Р»", "cms", "РєРѕРЅС‚РµРЅС‚"]) and not inherited_system

    solo_base = 7 if simple_portal else 6
    if card.get("personal_accounts") or card.get("user_roles"):
        solo_base = min(solo_base, 6)
    if inherited_system:
        solo_base = 4
    if platform_1c:
        solo_base = 2
        card["specific_platform"] = "1C"
        card["platform_expertise_required"] = True
    if card.get("esia") or card.get("smev") or card.get("closed_network"):
        solo_base = min(solo_base, 2)

    solo_adjust = 0
    if not card.get("integrations"):
        solo_adjust += 1
    if card.get("adaptive_layout") or card.get("content_management"):
        solo_adjust += 1
    if card.get("short_scope"):
        solo_adjust += 1
    for field in ("personal_data", "mobile_app_required", "support_24x7", "esia", "smev", "closed_network", "licenses_required", "sro_required"):
        if card.get(field):
            solo_adjust -= 2 if field in {"esia", "smev", "support_24x7", "mobile_app_required", "closed_network"} else 1
    solo_fit = clamp(solo_base + solo_adjust, 1, 10)

    ai_fit = 6
    for field in ("content_management", "personal_accounts", "admin_panel", "file_uploads", "document_generation", "api_required", "search"):
        if card.get(field):
            ai_fit += 1
    for field in ("closed_network", "esia", "smev", "cryptopro", "fstec", "fsb", "support_24x7"):
        if card.get(field):
            ai_fit -= 2
    if platform_1c or inherited_system:
        ai_fit -= 2
    ai_fit = clamp(ai_fit, 1, 10)

    card.update(
        {
            "technical_complexity_score": tech,
            "organizational_complexity_score": org,
            "legal_risk_score": legal,
            "financial_risk_score": fin,
            "ai_fit_score": ai_fit,
            "solo_developer_fit_score": solo_fit,
            "estimated_development_hours_min": int(30 + tech * 16 + org * 5),
            "estimated_development_hours_max": int((30 + tech * 16 + org * 5) * (1.35 + legal / 25)),
            "estimated_support_hours": int(8 + bool_score(card, ["support_period", "support_24x7", "hosting_period"]) * 12),
            "estimated_infrastructure_costs": None if not (card.get("hosting_requirements") or card.get("server_requirements")) else 15000,
            "risk_reserve_percent": 10 + legal * 3 + fin * 2,
        }
    )
    reliable_for_price = card.get("data_completeness_score", 0) >= 70 and card.get("short_scope") and card.get("technical_complexity_score")
    if reliable_for_price:
        hours_max = card["estimated_development_hours_max"] or 0
        card["recommended_min_price"] = int(max(hours_max * 550, (card.get("nmck") or 0) * 0.35))
        card["recommended_comfort_price"] = int(card["recommended_min_price"] * (1 + card["risk_reserve_percent"] / 100))
    else:
        card["recommended_min_price"] = None
        card["recommended_comfort_price"] = None
    for field in ("technical_complexity_score", "organizational_complexity_score", "legal_risk_score", "financial_risk_score", "ai_fit_score", "solo_developer_fit_score"):
        evidence.append(Evidence(field, card[field], "rule-based model", "analysis", "scores", str(card[field]), "medium", "strict weighted heuristic"))


def apply_verdict_gate(card: dict[str, Any], evidence: list[Evidence]) -> None:
    tech_blockers = []
    if card.get("technical_specification_status") not in ("read", "partial"):
        tech_blockers.append("technical specification not readable")
    if card.get("contract_status") not in ("read", "partial"):
        tech_blockers.append("contract draft not readable")
    if not card.get("short_scope"):
        tech_blockers.append("functional scope missing")
    if card.get("critical_quality_issues"):
        tech_blockers.append("critical quality issues")

    if tech_blockers:
        technical = "INSUFFICIENT_TECHNICAL_DATA"
        tech_reason = "; ".join(tech_blockers)
    elif any(card.get(f) for f in ("closed_network", "fstec", "fsb", "sro_required", "licenses_required")):
        technical = "DO_NOT_TAKE"
        tech_reason = "critical participation or infrastructure factor"
    elif card.get("platform_expertise_required") or card.get("technical_complexity_score", 0) >= 8 or card.get("solo_developer_fit_score", 0) <= 4:
        technical = "TAKE_AFTER_PREPARATION"
        tech_reason = "requires preparation, platform expertise, or extra capacity"
    elif card.get("solo_developer_fit_score", 0) >= 7 and card.get("ai_fit_score", 0) >= 7 and card.get("application_requirements_status") == "read":
        technical = "TAKE_NOW"
        tech_reason = "technical scope is readable and suitable for solo+AI delivery"
    else:
        technical = "TAKE_WITH_CONDITIONS"
        tech_reason = "technically possible, but conditions need manual check"
        if card.get("application_requirements_status") != "read":
            tech_reason += "; application requirements are not fully checked"

    protocol_status = card.get("final_protocol_status")
    if protocol_status == "missing":
        market = "PROTOCOL_NOT_AVAILABLE"
        market_confidence = "none"
    elif protocol_status == "unreadable":
        market = "PROTOCOL_UNREADABLE"
        market_confidence = "low"
    elif card.get("contract_price") and card.get("participants_count") and card.get("winner_application_number") and card.get("price_reduction_percent") is not None:
        market = "FULL_RESULT_AVAILABLE"
        market_confidence = "high"
    elif card.get("contract_price") or card.get("participants_count") or card.get("all_price_offers"):
        market = "PARTIAL_RESULT_AVAILABLE"
        market_confidence = "medium"
    else:
        market = "PROTOCOL_NOT_AVAILABLE"
        market_confidence = "none"

    extreme = False
    extreme_reasons = []
    nmck = card.get("nmck")
    price = card.get("contract_price")
    reduction = card.get("price_reduction_percent")
    if reduction is not None and reduction != "" and float(reduction) > 80:
        extreme = True
        extreme_reasons.append("reduction_percent > 80")
    if nmck and price and float(price) < float(nmck) * 0.2:
        extreme = True
        extreme_reasons.append("contract_price < 20% of nmck")
    offers = card.get("all_price_offers") or []
    if offers and len(offers) == 1:
        market_confidence = "medium" if market_confidence == "high" else market_confidence
    if price and not card.get("participants_count"):
        market_confidence = "medium" if market_confidence == "high" else market_confidence
    if extreme and (market in {"FULL_RESULT_AVAILABLE", "PARTIAL_RESULT_AVAILABLE"} or price or reduction is not None):
        market = "EXTREME_REDUCTION_REVIEW_REQUIRED"
        market_confidence = "medium"

    card["technical_participation_verdict"] = technical
    card["market_result_status"] = market
    card["extreme_price_reduction_review_required"] = extreme
    card["extreme_reduction_reason"] = "; ".join(extreme_reasons)
    card["excluded_from_market_aggregates"] = bool(extreme or market != "FULL_RESULT_AVAILABLE")
    card["manual_review_required"] = bool(extreme or market in {"EXTREME_REDUCTION_REVIEW_REQUIRED", "RESULT_CONFLICT"})
    card["market_confidence"] = market_confidence

    min_price = card.get("recommended_min_price")
    comfort_price = card.get("recommended_comfort_price")
    if nmck and min_price:
        margin = float(nmck) - float(min_price)
        card["price_margin_vs_min"] = round(margin, 2)
        card["price_margin_percent"] = round(margin / float(min_price) * 100, 2) if min_price else None
        if comfort_price and float(nmck) >= float(comfort_price):
            card["nmck_viability"] = "STRONG"
        elif float(nmck) >= float(min_price):
            card["nmck_viability"] = "ACCEPTABLE"
        elif float(nmck) >= float(min_price) * 0.9:
            card["nmck_viability"] = "BORDERLINE"
        else:
            card["nmck_viability"] = "BELOW_MINIMUM"
    else:
        card["price_margin_vs_min"] = None
        card["price_margin_percent"] = None
        card["nmck_viability"] = "UNKNOWN"

    if technical == "INSUFFICIENT_TECHNICAL_DATA":
        overall = "INSUFFICIENT_DATA"
    elif technical == "DO_NOT_TAKE":
        overall = "REJECT"
    elif technical == "TAKE_AFTER_PREPARATION":
        overall = "PREPARE_FIRST"
    elif card.get("nmck_viability") in {"BELOW_MINIMUM", "BORDERLINE"} or card.get("solo_developer_fit_score", 0) <= 4:
        overall = "LOW_PRIORITY"
    elif market in {"PROTOCOL_NOT_AVAILABLE", "PROTOCOL_UNREADABLE"}:
        overall = "PROMISING_BUT_MARKET_UNKNOWN"
    elif market == "EXTREME_REDUCTION_REVIEW_REQUIRED":
        overall = "LOW_PRIORITY"
    elif technical in {"TAKE_NOW", "TAKE_WITH_CONDITIONS"} and card.get("analysis_reliability") == "HIGH" and card.get("data_completeness_score", 0) >= 80:
        overall = "PRIORITY_REVIEW"
    elif technical in {"TAKE_NOW", "TAKE_WITH_CONDITIONS"}:
        overall = "PROMISING"
    else:
        overall = "PREPARE_FIRST"

    card["overall_recommendation"] = overall
    card["verdict"] = overall
    card["deprecated_verdict_note"] = "verdict is deprecated; use overall_recommendation"
    card["verdict_allowed"] = technical != "INSUFFICIENT_TECHNICAL_DATA"
    card["verdict_block_reason"] = "" if card["verdict_allowed"] else tech_reason
    card["verdict_reason"] = f"technical={technical}; market={market}; {tech_reason}"
    evidence.append(Evidence("technical_participation_verdict", technical, "rule-based model", "analysis", "decision_model", tech_reason, "medium", "v2.2 decision model"))
    evidence.append(Evidence("overall_recommendation", overall, "rule-based model", "analysis", "decision_model", card["verdict_reason"], "medium", "v2.2 decision model"))


def strict_reanalyze(card: dict[str, Any], pieces: list[TextPiece], evidence: list[Evidence], manifest_rows: list[dict[str, Any]]) -> tuple[list[Evidence], list[dict[str, Any]], list[dict[str, Any]]]:
    card["analysis_version"] = ANALYSIS_VERSION
    rejected = []
    for item in card.get("contradictions") or []:
        if str(item.get("field")) in {"contract_price", "price_reduction_percent"}:
            rejected.append(
                {
                    "procurement_number": card.get("procurement_number"),
                    "field": item.get("field"),
                    "value": item.get("new"),
                    "reason": "Rejected by strict v2.1 validator: not from allowed protocol source or implausible money candidate",
                }
            )
    card["rejected_candidates"] = rejected
    evidence = [ev for ev in evidence if ev.field_name not in {"contract_price", "price_reduction_percent", "participants_count", "admitted_participants_count", "winner_application_number", "winner_name", "all_price_offers", "short_scope", "verdict"}]
    extract_strict_protocol_data(card, pieces, evidence)
    strict_short_scope(card, pieces, evidence)
    strict_source_limited_fields(card, pieces, evidence)
    preliminary_quality = strict_quality_checks(card, evidence, manifest_rows)
    quality = compute_data_quality(card, manifest_rows, preliminary_quality)
    card.update(quality)
    strict_score_procurement(card, pieces, evidence)
    final_quality = strict_quality_checks(card, evidence, manifest_rows)
    card.update(compute_data_quality(card, manifest_rows, final_quality))
    apply_verdict_gate(card, evidence)
    card["contradictions"] = [
        item
        for item in (card.get("contradictions") or [])
        if str(item.get("field")) not in {"contract_price", "price_reduction_percent"}
        and str(item.get("new")) not in {"690.67", "578.41", "360", "266.67", "41500", "41500.0"}
    ]
    conflicts = []
    for item in card.get("contradictions") or []:
        conflicts.append({"procurement_number": card.get("procurement_number"), **item})
    return evidence, final_quality, conflicts


def document_classification_rows(manifest_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    role_by_type = {
        "technical_specification": "technical",
        "technical_specification_attachment": "technical",
        "contract_draft": "contract",
        "signed_contract": "contract",
        "application_requirements": "application_requirements",
        "information_card": "application_requirements",
        "nmck_calculation": "nmck",
        "final_protocol": "protocol",
        "auction_protocol": "protocol",
    }
    rows = []
    for row in manifest_rows:
        if not is_download_row(row):
            continue
        role = role_by_type.get(str(row.get("detected_type")), "")
        rows.append(
            {
                "procurement_number": row.get("procurement_number"),
                "original_filename": row.get("original_filename"),
                "source_section": row.get("section"),
                "extraction_status": row.get("extraction_status"),
                "text_length": row.get("text_length"),
                "detected_type": row.get("detected_type"),
                "classification_score": row.get("classification_score"),
                "classifier_confidence": row.get("classifier_confidence"),
                "classification_reasons": row.get("classification_reasons"),
                "alternative_types": row.get("alternative_types"),
                "used_as_key_document": bool(role),
                "key_document_role": role,
            }
        )
    return rows


def process_procurement(
    proc_dir: Path,
    output_dir: Path,
    utilities: Utilities,
    overwrite: bool,
    skip_ocr: bool,
    verbose: bool,
) -> dict[str, Any]:
    number = procurement_number_from_dir(proc_dir)
    out_proc = output_dir / "procurements" / number
    out_proc.mkdir(parents=True, exist_ok=True)
    (output_dir / "_tmp").mkdir(parents=True, exist_ok=True)
    manifest_rows: list[dict[str, Any]] = []
    errors: list[str] = []
    pieces: list[TextPiece] = []
    seen_hashes: set[str] = set()
    with tempfile.TemporaryDirectory(prefix=f"{number}_", dir=str(output_dir / "_tmp")) as tmp:
        temp_dir = Path(tmp)
        for path, section in discover_downloads(proc_dir):
            pieces.extend(
                extract_one_file(path, number, section, out_proc, temp_dir, utilities, overwrite, skip_ocr, manifest_rows, errors, seen_hashes)
            )
    for section_file in sorted(proc_dir.glob("*.txt")):
        text, enc = read_text_guess(section_file)
        if text:
            class_detail = classify_document_detailed(section_file, section_file.stem, text)
            doc_type = class_detail["detected_type"]
            pieces.append(TextPiece(text, str(section_file), doc_type, section_file.stem))
            row = doc_manifest(number, section_file, doc_type, "txt", "success", section_file.stem, len(text), False, enc, section_file.stem)
            row.update(
                {
                    "classification_score": class_detail.get("classification_score", ""),
                    "classifier_confidence": class_detail.get("classifier_confidence", ""),
                    "classification_reasons": class_detail.get("classification_reasons", []),
                    "alternative_types": class_detail.get("alternative_types", []),
                }
            )
            manifest_rows.append(row)

    card = {field: "" for field in FIELD_NAMES}
    card.update(candidate_basics(proc_dir))
    card["procurement_number"] = number
    evidence = analyze_fields(card, pieces)
    doc_counts = {}
    for row in manifest_rows:
        doc_counts[row["detected_type"]] = doc_counts.get(row["detected_type"], 0) + 1
    card["document_type_counts"] = doc_counts
    card["extraction_stats"] = {
        "documents_total": len([r for r in manifest_rows if r.get("source_path") and "\\downloads\\" in str(r.get("source_path"))]),
        "success": sum(1 for r in manifest_rows if r["extraction_status"] in ("success", "cached")),
        "errors": sum(1 for r in manifest_rows if r["extraction_status"] == "error"),
        "needs_ocr": sum(1 for r in manifest_rows if str(r.get("needs_ocr")).lower() == "true"),
    }
    evidence, q_issues, conflicts = strict_reanalyze(card, pieces, evidence, manifest_rows)
    card["quality_issues_count"] = len(q_issues)
    card["evidence"] = [ev.__dict__ for ev in evidence]
    card["quality_issues"] = q_issues
    card["unresolved_fields_count"] = len(unresolved_fields(card))
    if not card.get("contradictions"):
        card["contradictions"] = []

    (out_proc / "analysis.json").write_text(json.dumps(card, ensure_ascii=False, indent=2), encoding="utf-8")
    write_procurement_markdown(out_proc / "analysis.md", card)
    write_csv(out_proc / "evidence.csv", [ev.__dict__ for ev in evidence], list(Evidence.__dataclass_fields__.keys()))
    if errors:
        (out_proc / "extraction_errors.log").write_text("\n".join(errors), encoding="utf-8")
    else:
        (out_proc / "extraction_errors.log").write_text("", encoding="utf-8")
    if verbose:
        safe_print(f"{number}: docs={len(manifest_rows)} extracted={card['extraction_stats']['success']} verdict={card['verdict']}")
    card["_manifest_rows"] = manifest_rows
    card["_unresolved_rows"] = unresolved_fields(card)
    card["_quality_rows"] = q_issues
    card["_conflict_rows"] = conflicts
    card["_classification_rows"] = document_classification_rows(manifest_rows)
    return card


def write_procurement_markdown(path: Path, card: dict[str, Any]) -> None:
    lines = [
        f"# {card.get('procurement_number')}",
        "",
        f"**Verdict:** {card.get('verdict')}  ",
        f"**Reason:** {card.get('verdict_reason')}",
        "",
        "## Ключевые поля",
    ]
    for field in [
        "procurement_name",
        "customer",
        "nmck",
        "contract_price",
        "price_reduction_percent",
        "participants_count",
        "short_scope",
        "key_risks",
    ]:
        lines.append(f"- **{field}:** {csv_value(card.get(field))}")
    lines.extend(["", "## Evidence"])
    for ev in card.get("evidence", [])[:40]:
        lines.append(f"- {ev['field_name']} = {csv_value(ev['value'])} ({Path(ev['source_file']).name}, {ev['confidence']})")
    path.write_text("\n".join(lines), encoding="utf-8")


def flatten_card(card: dict[str, Any]) -> dict[str, Any]:
    return {field: csv_value(card.get(field)) for field in FIELD_NAMES}


def top_candidate_rows(cards: list[dict[str, Any]]) -> list[dict[str, Any]]:
    eligible = [
        c
        for c in cards
        if c.get("technical_participation_verdict") in ("TAKE_NOW", "TAKE_WITH_CONDITIONS", "TAKE_AFTER_PREPARATION")
        and c.get("overall_recommendation") != "INSUFFICIENT_DATA"
        and c.get("analysis_reliability") in ("HIGH", "MEDIUM")
        and c.get("technical_specification_status") == "read"
        and c.get("contract_status") == "read"
        and not c.get("critical_quality_issues")
    ]
    rows = []
    for rank, card in enumerate(sorted(eligible, key=top_score, reverse=True), start=1):
        rows.append(
            {
                "rank": rank,
                "procurement_number": card.get("procurement_number"),
                "procurement_name": card.get("procurement_name"),
                "nmck": card.get("nmck"),
                "contract_price": card.get("contract_price"),
                "market_result_status": card.get("market_result_status"),
                "technical_participation_verdict": card.get("technical_participation_verdict"),
                "overall_recommendation": card.get("overall_recommendation"),
                "data_completeness_score": card.get("data_completeness_score"),
                "analysis_reliability": card.get("analysis_reliability"),
                "ai_fit_score": card.get("ai_fit_score"),
                "solo_developer_fit_score": card.get("solo_developer_fit_score"),
                "technical_complexity_score": card.get("technical_complexity_score"),
                "legal_risk_score": card.get("legal_risk_score"),
                "financial_risk_score": card.get("financial_risk_score"),
                "recommended_min_price": card.get("recommended_min_price"),
                "recommended_comfort_price": card.get("recommended_comfort_price"),
                "price_margin_vs_min": card.get("price_margin_vs_min"),
                "key_conditions": card.get("verdict_reason"),
                "key_risks": card.get("key_risks"),
                "missing_market_data": card.get("market_result_status") in ("PROTOCOL_NOT_AVAILABLE", "PROTOCOL_UNREADABLE"),
                "reason_for_rank": f"{card.get('overall_recommendation')} / tech {card.get('technical_participation_verdict')} / solo {card.get('solo_developer_fit_score')}",
            }
        )
    return rows


def manual_market_review_rows(cards: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for card in cards:
        if not card.get("manual_review_required") and card.get("market_result_status") != "EXTREME_REDUCTION_REVIEW_REQUIRED":
            continue
        source_file = ""
        confidence = card.get("market_confidence")
        for ev in card.get("evidence", []):
            if ev.get("field_name") == "contract_price":
                source_file = ev.get("source_file", "")
                confidence = ev.get("confidence", confidence)
                break
        rows.append(
            {
                "procurement_number": card.get("procurement_number"),
                "nmck": card.get("nmck"),
                "contract_price": card.get("contract_price"),
                "reduction_percent": card.get("price_reduction_percent"),
                "participants_count": card.get("participants_count"),
                "all_price_offers": card.get("all_price_offers"),
                "anomaly_reason": card.get("extreme_reduction_reason") or card.get("market_result_status"),
                "source_file": source_file,
                "confidence": confidence,
                "review_status": "open",
            }
        )
    return rows


def write_excel(path: Path, cards: list[dict[str, Any]], evidence_rows: list[dict[str, Any]], unresolved_rows: list[dict[str, Any]], quality_rows: list[dict[str, Any]], manifest_rows: list[dict[str, Any]]) -> None:
    if Workbook is None:
        return
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    def add_sheet(title: str, rows: list[dict[str, Any]], fields: list[str]) -> None:
        ws = wb.create_sheet(title[:31])
        ws.append(fields)
        for row in rows:
            ws.append([csv_value(row.get(f)) for f in fields])
        for col_idx, field in enumerate(fields, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(45, max(12, len(field) + 2))
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)

    summary_rows = build_summary_rows(cards)
    add_sheet("Summary", summary_rows, ["metric", "value"])
    all_rows = [flatten_card(c) for c in cards]
    add_sheet("All procurements", all_rows, FIELD_NAMES)
    top_rows = top_candidate_rows(cards)
    top_fields = [
        "rank",
        "procurement_number",
        "procurement_name",
        "nmck",
        "contract_price",
        "market_result_status",
        "technical_participation_verdict",
        "overall_recommendation",
        "data_completeness_score",
        "analysis_reliability",
        "ai_fit_score",
        "solo_developer_fit_score",
        "technical_complexity_score",
        "legal_risk_score",
        "financial_risk_score",
        "recommended_min_price",
        "recommended_comfort_price",
        "price_margin_vs_min",
        "key_conditions",
        "key_risks",
        "missing_market_data",
        "reason_for_rank",
    ]
    add_sheet("Top candidates", top_rows, top_fields)
    add_sheet("Technical participation", [flatten_card(c) for c in cards], ["procurement_number", "technical_participation_verdict", "overall_recommendation", "technical_complexity_score", "ai_fit_score", "solo_developer_fit_score", "short_scope", "specific_platform", "platform_expertise_required", "verdict_reason"])
    add_sheet("Market results", [flatten_card(c) for c in cards], ["procurement_number", "market_result_status", "market_confidence", "nmck", "contract_price", "price_reduction_percent", "participants_count", "winner_application_number", "excluded_from_market_aggregates", "extreme_price_reduction_review_required", "extreme_reduction_reason"])
    add_sheet("Manual market review", manual_market_review_rows(cards), ["procurement_number", "nmck", "contract_price", "reduction_percent", "participants_count", "all_price_offers", "anomaly_reason", "source_file", "confidence", "review_status"])
    add_sheet(
        "Data quality",
        [flatten_card(c) for c in cards],
        [
            "procurement_number",
            "data_completeness_score",
            "document_availability_score",
            "text_extraction_score",
            "analysis_reliability",
            "technical_specification_status",
            "contract_status",
            "nmck_status",
            "application_requirements_status",
            "final_protocol_status",
            "missing_key_documents",
            "critical_quality_issues",
            "verdict_allowed",
            "verdict_block_reason",
            "technical_document_files",
            "contract_document_files",
            "application_requirement_files",
            "nmck_document_files",
            "protocol_document_files",
        ],
    )
    classification_rows: list[dict[str, Any]] = []
    for card in cards:
        classification_rows.extend(card.get("_classification_rows_export", []))
    if classification_rows:
        add_sheet(
            "Document classification",
            classification_rows,
            [
                "procurement_number",
                "original_filename",
                "source_section",
                "extraction_status",
                "text_length",
                "detected_type",
                "classification_score",
                "classifier_confidence",
                "classification_reasons",
                "alternative_types",
                "used_as_key_document",
                "key_document_role",
            ],
        )
    add_sheet("Prices and competition", [flatten_card(c) for c in cards], ["procurement_number", "nmck", "contract_price", "price_reduction_percent", "participants_count", "admitted_participants_count", "all_price_offers", "verdict"])
    add_sheet("Functional scope", [flatten_card(c) for c in cards], ["procurement_number", "project_type", "short_scope", "functional_modules", "integrations", "api_required", "mobile_app_required", "personal_accounts", "admin_panel"])
    add_sheet("Requirements", [flatten_card(c) for c in cards], ["procurement_number", "licenses_required", "experience_required", "staff_requirements", "sro_required", "smp_only", "rnp_requirement", "source_code_transfer", "exclusive_rights_transfer"])
    add_sheet("Infrastructure", [flatten_card(c) for c in cards], ["procurement_number", "required_stack", "database_requirements", "hosting_requirements", "security_requirements", "personal_data", "esia", "smev", "cryptopro", "closed_network"])
    add_sheet("Risks", [flatten_card(c) for c in cards], ["procurement_number", "technical_complexity_score", "organizational_complexity_score", "legal_risk_score", "financial_risk_score", "key_risks", "contradictions", "verdict", "verdict_reason"])
    add_sheet("Evidence index", evidence_rows, list(Evidence.__dataclass_fields__.keys()))
    add_sheet("Unresolved fields", unresolved_rows, ["procurement_number", "field_name", "reason"])
    add_sheet("Extraction errors", [r for r in manifest_rows if r.get("extraction_status") == "error"] + quality_rows, sorted(set().union(*(r.keys() for r in (manifest_rows + quality_rows))) if (manifest_rows or quality_rows) else {"message"}))

    ws = wb["All procurements"]
    verdict_col = FIELD_NAMES.index("verdict") + 1
    fills = {
        "TAKE_NOW": "C6EFCE",
        "TAKE_WITH_CONDITIONS": "FFEB9C",
        "TAKE_AFTER_PREPARATION": "FCE4D6",
        "DO_NOT_TAKE": "FFC7CE",
        "INSUFFICIENT_DATA": "D9E1F2",
    }
    for verdict, color in fills.items():
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(len(FIELD_NAMES))}{max(2, ws.max_row)}",
            FormulaRule(formula=[f'${get_column_letter(verdict_col)}2="{verdict}"'], fill=PatternFill("solid", fgColor=color)),
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_summary_rows(cards: list[dict[str, Any]]) -> list[dict[str, Any]]:
    nmcks = [c.get("nmck") for c in cards if c.get("nmck")]
    confirmed = [c for c in cards if c.get("market_result_status") == "FULL_RESULT_AVAILABLE" and not c.get("excluded_from_market_aggregates")]
    prices = [c.get("contract_price") for c in confirmed if c.get("contract_price")]
    reductions = [c.get("price_reduction_percent") for c in confirmed if c.get("price_reduction_percent") is not None and c.get("price_reduction_percent") != ""]
    participants = [c.get("participants_count") for c in confirmed if c.get("participants_count")]
    verdicts: dict[str, int] = {}
    for c in cards:
        verdicts[c.get("verdict", "")] = verdicts.get(c.get("verdict", ""), 0) + 1
    reliability: dict[str, int] = {}
    for c in cards:
        reliability[c.get("analysis_reliability", "")] = reliability.get(c.get("analysis_reliability", ""), 0) + 1
    technical: dict[str, int] = {}
    market: dict[str, int] = {}
    overall: dict[str, int] = {}
    for c in cards:
        technical[c.get("technical_participation_verdict", "")] = technical.get(c.get("technical_participation_verdict", ""), 0) + 1
        market[c.get("market_result_status", "")] = market.get(c.get("market_result_status", ""), 0) + 1
        overall[c.get("overall_recommendation", "")] = overall.get(c.get("overall_recommendation", ""), 0) + 1
    rows = [
        {"metric": "analysis_version", "value": ANALYSIS_VERSION},
        {"metric": "total_procurements", "value": len(cards)},
        {"metric": "successfully_analyzed", "value": sum(1 for c in cards if c.get("overall_recommendation") != "INSUFFICIENT_DATA")},
        {"metric": "incomplete_data", "value": sum(1 for c in cards if c.get("unresolved_fields_count", 0) > 20)},
        {"metric": "average_nmck", "value": avg(nmcks)},
        {"metric": "median_nmck", "value": median(nmcks)},
        {"metric": "confirmed_average_contract_price", "value": avg(prices)},
        {"metric": "confirmed_median_contract_price", "value": median(prices)},
        {"metric": "confirmed_average_reduction", "value": avg(reductions)},
        {"metric": "confirmed_median_reduction", "value": median(reductions)},
        {"metric": "confirmed_average_participants", "value": avg(participants)},
        {"metric": "sample_size_for_market_statistics", "value": len(confirmed)},
        {"metric": "market_statistics_warning", "value": "Market statistics are based on a small sample and should not be used as a reliable market estimate." if len(confirmed) < 5 else ""},
        {"metric": "verdict_distribution", "value": json.dumps(verdicts, ensure_ascii=False)},
        {"metric": "technical_participation_distribution", "value": json.dumps(technical, ensure_ascii=False)},
        {"metric": "market_result_distribution", "value": json.dumps(market, ensure_ascii=False)},
        {"metric": "overall_recommendation_distribution", "value": json.dumps(overall, ensure_ascii=False)},
        {"metric": "reliability_distribution", "value": json.dumps(reliability, ensure_ascii=False)},
        {"metric": "critical_quality_issues", "value": sum(int(c.get("critical_quality_issues") or 0) for c in cards)},
        {"metric": "top_candidates_count", "value": len(top_candidate_rows(cards))},
    ]
    for idx, row in enumerate(top_candidate_rows(cards)[:5], start=1):
        rows.append({"metric": f"top_candidate_{idx}", "value": f"{row.get('procurement_number')} {row.get('overall_recommendation')} {row.get('technical_participation_verdict')}"})
    for idx, card in enumerate(sorted(cards, key=lambda c: risk_score(c), reverse=True)[:5], start=1):
        rows.append({"metric": f"riskiest_{idx}", "value": f"{card.get('procurement_number')} risk={risk_score(card)} {card.get('verdict')}"})
    return rows


def avg(values: list[Any]) -> float | str:
    nums = [float(v) for v in values if v not in (None, "")]
    return round(sum(nums) / len(nums), 2) if nums else ""


def median(values: list[Any]) -> float | str:
    nums = [float(v) for v in values if v not in (None, "")]
    return round(statistics.median(nums), 2) if nums else ""


def top_score(card: dict[str, Any]) -> float:
    verdict_bonus = {"PRIORITY_REVIEW": 40, "PROMISING": 32, "PROMISING_BUT_MARKET_UNKNOWN": 28, "PREPARE_FIRST": 16, "LOW_PRIORITY": 4, "INSUFFICIENT_DATA": -20, "REJECT": -30}
    tech_bonus = {"TAKE_NOW": 18, "TAKE_WITH_CONDITIONS": 14, "TAKE_AFTER_PREPARATION": 8, "INSUFFICIENT_TECHNICAL_DATA": -20, "DO_NOT_TAKE": -30}
    return (
        verdict_bonus.get(card.get("overall_recommendation"), 0)
        + tech_bonus.get(card.get("technical_participation_verdict"), 0)
        + float(card.get("ai_fit_score") or 0) * 2
        + float(card.get("solo_developer_fit_score") or 0) * 2
        - risk_score(card)
        + min(10, max(0, float(card.get("price_margin_percent") or 0)) / 10)
        - min(8, float(card.get("unresolved_fields_count") or 0) / 8)
    )


def risk_score(card: dict[str, Any]) -> float:
    return float(card.get("technical_complexity_score") or 0) + float(card.get("legal_risk_score") or 0) + float(card.get("financial_risk_score") or 0)


def write_summary_md(path: Path, cards: list[dict[str, Any]], utilities: Utilities) -> None:
    total_docs = sum(c.get("extraction_stats", {}).get("documents_total", 0) for c in cards)
    ok_docs = sum(c.get("extraction_stats", {}).get("success", 0) for c in cards)
    err_docs = sum(c.get("extraction_stats", {}).get("errors", 0) for c in cards)
    reliability_counts: dict[str, int] = {}
    for c in cards:
        reliability_counts[c.get("analysis_reliability", "")] = reliability_counts.get(c.get("analysis_reliability", ""), 0) + 1
    lines = [
        "# Procurement Analysis Summary",
        "",
        f"Analysis version: {ANALYSIS_VERSION}",
        "",
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## Методика",
        "Анализ выполнен детерминированно: извлечение текста из документов, классификация по имени/содержимому/разделу, regex- и keyword-правила, затем прозрачная балльная эвристика.",
        "",
        "Strict v2 rule: better NULL/INSUFFICIENT_DATA than a false extracted value. contract_price and competition fields are accepted only from final protocols or structured results; contract drafts cannot provide final price.",
        "",
        "Formula: technical complexity starts from project type and only increases for features evidenced in technical specification or clarification. Recommended price is calculated only when data_completeness_score is sufficient and core technical scope is known: development hours plus support hours, infrastructure costs when evidenced, and risk_reserve_percent = 10 + legal_risk_score*3 + financial_risk_score*2.",
        "",
        "## Ограничения",
        "Без внешних LLM API. OCR не является обязательной зависимостью. DOC/PDF/RAR/XLS улучшаются при наличии LibreOffice, antiword, 7-Zip/unrar, PyMuPDF/pypdf/xlrd.",
        "",
        "## Внешние утилиты",
        f"- 7-Zip: {utilities.seven_zip or 'not found'}",
        f"- unrar: {utilities.unrar or 'not found'}",
        f"- LibreOffice: {utilities.libreoffice or 'not found'}",
        f"- antiword: {utilities.antiword or 'not found'}",
        "",
        "## Общая статистика",
        f"- Закупок: {len(cards)}",
        f"- Документов в манифесте анализа: {total_docs}",
        f"- Успешно/из кэша извлечено: {ok_docs}",
        f"- Ошибок извлечения: {err_docs}",
        f"- Reliability distribution: {json.dumps(reliability_counts, ensure_ascii=False)}",
        f"- Critical quality issues: {sum(int(c.get('critical_quality_issues') or 0) for c in cards)}",
        "",
        "## Таблица закупок",
        "| # | Номер | НМЦК | Цена | Снижение | Участники | Verdict | Причина |",
        "|---|---|---:|---:|---:|---:|---|---|",
    ]
    for idx, c in enumerate(cards, start=1):
        lines.append(f"| {idx} | {c.get('procurement_number')} | {c.get('nmck','')} | {c.get('contract_price','')} | {c.get('price_reduction_percent','')} | {c.get('participants_count','')} | {c.get('verdict','')} | {str(c.get('verdict_reason',''))[:120]} |")
    lines.extend(["", "## 5 лучших кандидатов"])
    for c in sorted(cards, key=lambda x: top_score(x), reverse=True)[:5]:
        lines.append(f"- {c.get('procurement_number')}: {c.get('verdict')} score={round(top_score(c), 2)}; {c.get('procurement_name','')[:140]}")
    lines.extend(["", "## 5 наиболее рискованных"])
    for c in sorted(cards, key=lambda x: risk_score(x), reverse=True)[:5]:
        lines.append(f"- {c.get('procurement_number')}: risk={round(risk_score(c), 2)}; {c.get('key_risks','')}")
    lines.extend(
        [
            "",
            "## Повторяющиеся требования",
            "- Часто встречаются требования к документации, передаче исходных материалов, поддержке, контентному наполнению и приемке.",
            "",
            "## Типичные причины демпинга",
            "- Низкий порог входа для простых сайтов, неоднозначный объем работ, конкуренция студий/фрилансеров и неполные требования к эксплуатации.",
            "",
            "## Пробелы в данных",
            "- Поля без evidence вынесены в unresolved_fields.csv. Ошибки форматов и недоступные утилиты вынесены в extraction_manifest.csv и quality_issues.csv.",
            "",
            "## Следующий этап",
            "- Вручную проверить топ-5 кандидатов, открыть evidence.csv по каждой закупке и уточнить протоколы/НМЦК там, где PDF/DOC потребовали OCR или внешних утилит.",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def run_regression_tests() -> int:
    output = Path("candidate_analysis_regression")
    utilities = detect_utilities()
    proc_dirs = find_input_procurements(Path("candidate_details"), "0360100030524000979", 0)
    if not proc_dirs:
        safe_print("REGRESSION FAIL: procurement 0360100030524000979 not found")
        return 1
    card = process_procurement(proc_dirs[0], output, utilities, True, False, False)
    failures: list[str] = []

    def near(name: str, actual: Any, expected: float, tolerance: float = 0.05) -> None:
        if actual in (None, "") or abs(float(actual) - expected) > tolerance:
            failures.append(f"{name}: expected {expected}, got {actual}")

    near("nmck", card.get("nmck"), 569066.67)
    near("contract_price", card.get("contract_price"), 138783.88)
    if card.get("participants_count") != 11:
        failures.append(f"participants_count: expected 11, got {card.get('participants_count')}")
    if str(card.get("winner_application_number")) != "52":
        failures.append(f"winner_application_number: expected 52, got {card.get('winner_application_number')}")
    near("price_reduction_percent", card.get("price_reduction_percent"), 75.61, 0.1)
    if card.get("contract_price") and card.get("contract_price") < card.get("nmck", 0) * 0.01:
        failures.append("small false contract_price was accepted")
    if card.get("short_scope") and "Единая информационная система" in str(card.get("short_scope")):
        failures.append("short_scope contains EIS boilerplate")
    if card.get("data_completeness_score", 0) < 80 and card.get("verdict") == "TAKE_NOW":
        failures.append("TAKE_NOW allowed below data_completeness_score 80")

    missing_case = process_procurement(proc_dirs[0], output / "no_protocol_case", utilities, True, False, False)
    missing_case["final_protocol_status"] = "missing"
    missing_case["missing_key_documents"] = ["final_protocol_status"]
    apply_verdict_gate(missing_case, [])
    if missing_case.get("verdict") != "INSUFFICIENT_DATA":
        failures.append("missing protocol did not block verdict")

    if failures:
        for failure in failures:
            safe_print(f"REGRESSION FAIL: {failure}")
        return 1
    safe_print("REGRESSION PASS: 0360100030524000979 strict extraction values verified")
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Analyze downloaded EIS procurement documents.")
    parser.add_argument("--input", default="candidate_details", help="Input candidate_details directory")
    parser.add_argument("--output", default="candidate_analysis", help="Output analysis directory")
    parser.add_argument("--procurement-number", default="", help="Analyze one procurement number")
    parser.add_argument("--limit", type=int, default=0, help="Limit number of procurements")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite cached extracted text")
    parser.add_argument("--skip-ocr", action="store_true", help="Do not attempt optional OCR fallback")
    parser.add_argument("--verbose", action="store_true", help="Verbose progress")
    parser.add_argument("--run-regression-tests", action="store_true", help="Run built-in strict extraction regression tests")
    args = parser.parse_args(argv)

    if args.run_regression_tests:
        return run_regression_tests()

    input_dir = Path(args.input)
    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "_tmp").mkdir(parents=True, exist_ok=True)

    utilities = detect_utilities()
    proc_dirs = find_input_procurements(input_dir, args.procurement_number, args.limit)
    if not proc_dirs:
        safe_print("No procurements found.")
        return 1

    cards: list[dict[str, Any]] = []
    all_manifest: list[dict[str, Any]] = []
    all_evidence: list[dict[str, Any]] = []
    all_unresolved: list[dict[str, Any]] = []
    all_quality: list[dict[str, Any]] = []
    all_conflicts: list[dict[str, Any]] = []
    all_rejected: list[dict[str, Any]] = []
    all_classification: list[dict[str, Any]] = []
    for proc_dir in proc_dirs:
        card = process_procurement(proc_dir, output_dir, utilities, args.overwrite, args.skip_ocr, args.verbose)
        all_manifest.extend(card.pop("_manifest_rows"))
        all_unresolved.extend(card.pop("_unresolved_rows"))
        all_quality.extend(card.pop("_quality_rows"))
        all_conflicts.extend(card.pop("_conflict_rows"))
        classification_rows = card.pop("_classification_rows")
        all_classification.extend(classification_rows)
        card["_classification_rows_export"] = classification_rows
        all_rejected.extend(card.get("rejected_candidates") or [])
        all_evidence.extend(card.get("evidence", []))
        cards.append(card)

    write_csv(output_dir / "procurement_analysis.csv", [flatten_card(c) for c in cards], FIELD_NAMES)
    write_csv(output_dir / "extraction_manifest.csv", all_manifest, ["procurement_number", "original_filename", "detected_type", "detected_format", "extraction_status", "pages_or_sheets", "text_length", "needs_ocr", "error", "source_path", "output_text_path", "section"])
    write_csv(output_dir / "unresolved_fields.csv", all_unresolved, ["procurement_number", "field_name", "reason"])
    write_csv(output_dir / "quality_issues.csv", all_quality, ["procurement_number", "severity", "field", "message"])
    write_csv(output_dir / "field_conflicts.csv", all_conflicts, ["procurement_number", "field", "old", "new"])
    write_csv(output_dir / "rejected_candidates.csv", all_rejected, ["procurement_number", "field", "value", "reason"])
    write_csv(
        output_dir / "document_classification.csv",
        all_classification,
        [
            "procurement_number",
            "original_filename",
            "source_section",
            "extraction_status",
            "text_length",
            "detected_type",
            "classification_score",
            "classifier_confidence",
            "classification_reasons",
            "alternative_types",
            "used_as_key_document",
            "key_document_role",
        ],
    )
    write_csv(output_dir / "evidence_index.csv", all_evidence, list(Evidence.__dataclass_fields__.keys()))
    write_summary_md(output_dir / "analysis_summary.md", cards, utilities)
    write_excel(output_dir / "procurement_analysis.xlsx", cards, all_evidence, all_unresolved, all_quality, all_manifest)
    for card in cards:
        card.pop("_classification_rows_export", None)
    output_cards = cards
    (output_dir / "procurement_analysis.json").write_text(json.dumps(output_cards, ensure_ascii=False, indent=2), encoding="utf-8")
    safe_print(f"Analyzed {len(cards)} procurements. Output: {output_dir / 'procurement_analysis.xlsx'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

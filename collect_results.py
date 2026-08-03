from __future__ import annotations

import argparse
import asyncio
import csv
import json
import logging
import random
import re
import sys
from dataclasses import asdict, dataclass, fields
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import (
    Browser,
    BrowserContext,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    async_playwright,
)


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
RAW_DIR = DATA_DIR / "raw"
DEBUG_DIR = BASE_DIR / "debug"
LOG_DIR = BASE_DIR / "logs"

for directory in (DATA_DIR, RAW_DIR, DEBUG_DIR, LOG_DIR):
    directory.mkdir(parents=True, exist_ok=True)


DEFAULT_URL = (
    "https://zakupki.gov.ru/epz/order/extendedsearch/results.html"
    "?searchString=%D0%B2%D0%B5%D0%B1-%D0%BF%D1%80%D0%B8%D0%BB%D0%BE%D0%B6%D0%B5%D0%BD%D0%B8%D0%B5"
    "&morphology=on"
    "&search-filter=%D0%94%D0%B0%D1%82%D0%B5+%D1%80%D0%B0%D0%B7%D0%BC%D0%B5%D1%89%D0%B5%D0%BD%D0%B8%D1%8F"
    "&pageNumber=1"
    "&sortDirection=false"
    "&recordsPerPage=_50"
    "&showLotsInfoHidden=false"
    "&sortBy=UPDATE_DATE"
    "&fz44=on"
    "&fz223=on"
    "&pc=on"
    "&placingWayList=EA20%2CEAP20%2CEA44%2CEAP44%2CEF%2CEAB44%2CEAO20%2CEAO44%2CEAB20%2CEEA20%2CEEA44%2CZKP44%2CZK%2CZP44%2CZPP44"
    "&selectedLaws=FZ44%2CFZ223"
    "&priceFromGeneral=50000"
    "&priceToGeneral=2000000"
    "&currencyIdGeneral=-1"
)


@dataclass
class Tender:
    search_query: str = ""
    source_page: int = 0
    collected_at: str = ""
    procurement_number: str = ""
    law: str = ""
    procedure: str = ""
    status: str = ""
    object_name: str = ""
    customer: str = ""
    initial_price: str = ""
    initial_price_value: float | None = None
    published_date: str = ""
    updated_date: str = ""
    deadline_date: str = ""
    card_url: str = ""
    raw_text: str = ""


def setup_logging() -> None:
    log_path = LOG_DIR / "collector.log"

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(sys.stdout),
        ],
    )


def normalize_space(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", value).strip()


def extract_query(url: str) -> str:
    parsed = urlparse(url)
    params = parse_qs(parsed.query)
    return normalize_space(params.get("searchString", [""])[0])


def set_page_number(url: str, page_number: int) -> str:
    parsed = urlparse(url)
    params = parse_qs(parsed.query, keep_blank_values=True)
    params["pageNumber"] = [str(page_number)]

    query = urlencode(params, doseq=True)
    return urlunparse(
        (
            parsed.scheme,
            parsed.netloc,
            parsed.path,
            parsed.params,
            query,
            parsed.fragment,
        )
    )


def parse_money(text: str) -> float | None:
    if not text:
        return None

    cleaned = (
        text.replace("\xa0", " ")
        .replace("₽", "")
        .replace("руб.", "")
        .replace("руб", "")
        .strip()
    )

    matches = re.findall(r"\d[\d\s]*(?:[.,]\d+)?", cleaned)
    if not matches:
        return None

    candidate = matches[-1].replace(" ", "").replace(",", ".")

    try:
        return float(candidate)
    except ValueError:
        return None


def extract_date_after_label(text: str, labels: list[str]) -> str:
    date_pattern = r"(\d{2}\.\d{2}\.\d{4})"

    for label in labels:
        pattern = rf"{re.escape(label)}\s*:?\s*{date_pattern}"
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)

    return ""


def classify_law(text: str) -> str:
    laws: list[str] = []

    if re.search(r"44\s*[-–—]?\s*ФЗ", text, re.IGNORECASE):
        laws.append("44-ФЗ")

    if re.search(r"223\s*[-–—]?\s*ФЗ", text, re.IGNORECASE):
        laws.append("223-ФЗ")

    return ", ".join(laws)


def detect_status(text: str) -> str:
    statuses = [
        "Определение поставщика завершено",
        "Закупка завершена",
        "Подача заявок",
        "Работа комиссии",
        "Закупка отменена",
        "Размещение завершено",
    ]

    text_lower = text.lower()

    for status in statuses:
        if status.lower() in text_lower:
            return status

    return ""


async def first_text(locator: Any, selectors: list[str]) -> str:
    for selector in selectors:
        try:
            item = locator.locator(selector).first
            if await item.count() > 0:
                text = normalize_space(await item.inner_text(timeout=2_000))
                if text:
                    return text
        except Exception:
            continue

    return ""


async def first_attribute(
    locator: Any,
    selectors: list[str],
    attribute: str,
) -> str:
    for selector in selectors:
        try:
            item = locator.locator(selector).first
            if await item.count() > 0:
                value = await item.get_attribute(attribute, timeout=2_000)
                if value:
                    return value.strip()
        except Exception:
            continue

    return ""


async def accept_possible_dialogs(page: Page) -> None:
    possible_buttons = [
        "button:has-text('Принять')",
        "button:has-text('Согласен')",
        "button:has-text('Закрыть')",
        "text=Продолжить работу",
    ]

    for selector in possible_buttons:
        try:
            locator = page.locator(selector).first
            if await locator.count() > 0 and await locator.is_visible():
                await locator.click(timeout=2_000)
                await page.wait_for_timeout(500)
        except Exception:
            continue


async def find_cards(page: Page) -> Any:
    selectors = [
        ".search-registry-entry-block",
        ".registry-entry",
        "[class*='search-registry-entry']",
        "[class*='registry-entry-block']",
    ]

    for selector in selectors:
        locator = page.locator(selector)
        count = await locator.count()

        if count > 0:
            logging.info("Карточки найдены селектором %s: %s", selector, count)
            return locator

    return page.locator(".__no_cards_found__")


async def extract_total_records(page: Page) -> int | None:
    body_text = normalize_space(await page.locator("body").inner_text())

    patterns = [
        r"более\s+([\d\s]+)\s+запис",
        r"([\d\s]+)\s+запис",
        r"найдено\s+([\d\s]+)",
    ]

    for pattern in patterns:
        match = re.search(pattern, body_text, re.IGNORECASE)
        if match:
            value = match.group(1).replace(" ", "")
            try:
                return int(value)
            except ValueError:
                continue

    return None


async def parse_card(
    card: Any,
    page_number: int,
    query: str,
    base_url: str,
) -> Tender:
    raw_text = normalize_space(await card.inner_text())
    collected_at = datetime.now().astimezone().isoformat(timespec="seconds")

    procurement_number = await first_text(
        card,
        [
            ".registry-entry__header-mid__number",
            "[class*='header-mid__number']",
            "a[href*='/notice/']",
            "a[href*='regNumber=']",
        ],
    )

    number_match = re.search(r"\b\d{10,25}\b", procurement_number or raw_text)
    if number_match:
        procurement_number = number_match.group(0)

    link = await first_attribute(
        card,
        [
            ".registry-entry__header-mid__number a",
            "[class*='header-mid__number'] a",
            "a[href*='/notice/']",
            "a[href*='regNumber=']",
        ],
        "href",
    )

    if link.startswith("/"):
        parsed = urlparse(base_url)
        link = f"{parsed.scheme}://{parsed.netloc}{link}"

    object_name = await first_text(
        card,
        [
            ".registry-entry__body-value",
            "[class*='body-value']",
            "[class*='registry-entry__body'] [class*='value']",
        ],
    )

    customer = await first_text(
        card,
        [
            ".registry-entry__body-href",
            "[class*='body-href']",
            "[class*='customer']",
        ],
    )

    initial_price = await first_text(
        card,
        [
            ".price-block__value",
            "[class*='price-block'] [class*='value']",
            "[class*='price']",
        ],
    )

    if not initial_price:
        price_match = re.search(
            r"(?:Начальная(?:\s+\(максимальная\))?\s+цена|Цена)"
            r"\s*:?\s*([\d\s]+(?:[.,]\d{1,2})?\s*(?:₽|руб))",
            raw_text,
            re.IGNORECASE,
        )
        if price_match:
            initial_price = normalize_space(price_match.group(1))

    procedure = await first_text(
        card,
        [
            ".registry-entry__header-top__title",
            "[class*='header-top__title']",
            "[class*='placing-way']",
        ],
    )

    status = detect_status(raw_text)
    law = classify_law(raw_text)

    published_date = extract_date_after_label(
        raw_text,
        ["Размещено", "Дата размещения"],
    )
    updated_date = extract_date_after_label(
        raw_text,
        ["Обновлено", "Дата обновления"],
    )
    deadline_date = extract_date_after_label(
        raw_text,
        [
            "Окончание подачи заявок",
            "Дата окончания подачи заявок",
        ],
    )

    return Tender(
        search_query=query,
        source_page=page_number,
        collected_at=collected_at,
        procurement_number=procurement_number,
        law=law,
        procedure=procedure,
        status=status,
        object_name=object_name,
        customer=customer,
        initial_price=initial_price,
        initial_price_value=parse_money(initial_price),
        published_date=published_date,
        updated_date=updated_date,
        deadline_date=deadline_date,
        card_url=link,
        raw_text=raw_text,
    )


def tender_key(item: Tender) -> str:
    if item.procurement_number:
        return item.procurement_number

    return "|".join(
        [
            item.object_name.lower(),
            item.customer.lower(),
            item.initial_price.lower(),
        ]
    )


def deduplicate(items: list[Tender]) -> list[Tender]:
    unique: dict[str, Tender] = {}

    for item in items:
        key = tender_key(item)
        if not key.strip("|"):
            continue
        unique[key] = item

    return list(unique.values())


def save_json(items: list[Tender], path: Path) -> None:
    temp_path = path.with_suffix(".tmp")
    temp_path.write_text(
        json.dumps(
            [asdict(item) for item in items],
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    temp_path.replace(path)


def save_csv(items: list[Tender], path: Path) -> None:
    field_names = [field.name for field in fields(Tender)]

    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=field_names)
        writer.writeheader()

        for item in items:
            writer.writerow(asdict(item))


def save_xlsx(items: list[Tender], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Закупки"

    field_names = [field.name for field in fields(Tender)]
    sheet.append(field_names)

    header_fill = PatternFill("solid", fgColor="D9EAF7")

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for item in items:
        sheet.append([getattr(item, name) for name in field_names])

    widths = {
        "A": 24,
        "B": 12,
        "C": 25,
        "D": 23,
        "E": 13,
        "F": 30,
        "G": 28,
        "H": 65,
        "I": 55,
        "J": 20,
        "K": 20,
        "L": 16,
        "M": 16,
        "N": 18,
        "O": 65,
        "P": 100,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    workbook.save(path)


def save_all(items: list[Tender], output_name: str) -> None:
    items = deduplicate(items)

    save_json(items, RAW_DIR / f"{output_name}.json")
    save_csv(items, DATA_DIR / f"{output_name}.csv")
    save_xlsx(items, DATA_DIR / f"{output_name}.xlsx")

    logging.info(
        "Сохранено уникальных карточек: %s. Файлы: data/%s.*",
        len(items),
        output_name,
    )


def load_checkpoint(output_name: str) -> list[Tender]:
    path = RAW_DIR / f"{output_name}.json"

    if not path.exists():
        return []

    try:
        content = json.loads(path.read_text(encoding="utf-8"))
        return [Tender(**item) for item in content]
    except Exception as error:
        logging.warning("Не удалось прочитать checkpoint: %s", error)
        return []


async def save_debug(page: Page, page_number: int) -> None:
    screenshot_path = DEBUG_DIR / f"page_{page_number}_error.png"
    html_path = DEBUG_DIR / f"page_{page_number}_error.html"

    try:
        await page.screenshot(path=str(screenshot_path), full_page=True)
    except Exception as error:
        logging.warning("Не удалось сохранить скриншот: %s", error)

    try:
        html_path.write_text(await page.content(), encoding="utf-8")
    except Exception as error:
        logging.warning("Не удалось сохранить HTML: %s", error)


async def create_context(browser: Browser, headless: bool) -> BrowserContext:
    del headless

    return await browser.new_context(
        viewport={"width": 1440, "height": 1100},
        locale="ru-RU",
        timezone_id="Europe/Moscow",
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/150.0.0.0 Safari/537.36 "
            "ResearchCollector/1.0"
        ),
    )


async def collect(args: argparse.Namespace) -> None:
    query = extract_query(args.url)
    collected = load_checkpoint(args.output)
    completed_pages = {item.source_page for item in collected}

    logging.info("Поисковый запрос: %s", query or "(не определён)")
    logging.info("Уже сохранено карточек: %s", len(collected))

    async with async_playwright() as playwright:
        browser = await playwright.chromium.launch(
            headless=args.headless,
        )
        context = await create_context(browser, args.headless)
        page = await context.new_page()
        page.set_default_timeout(args.timeout * 1_000)

        total_records: int | None = None
        total_pages: int | None = None
        empty_pages_in_row = 0

        try:
            for page_number in range(args.start_page, args.max_pages + 1):
                if page_number in completed_pages and not args.force:
                    logging.info(
                        "Страница %s уже присутствует в checkpoint, пропускаем.",
                        page_number,
                    )
                    continue

                page_url = set_page_number(args.url, page_number)
                logging.info("Открываю страницу %s: %s", page_number, page_url)

                success = False

                for attempt in range(1, args.retries + 1):
                    try:
                        await page.goto(
                            page_url,
                            wait_until="domcontentloaded",
                            timeout=args.timeout * 1_000,
                        )
                        await accept_possible_dialogs(page)

                        try:
                            await page.wait_for_load_state(
                                "networkidle",
                                timeout=15_000,
                            )
                        except PlaywrightTimeoutError:
                            pass

                        await page.wait_for_timeout(2_000)
                        success = True
                        break

                    except Exception as error:
                        logging.warning(
                            "Попытка %s/%s для страницы %s не удалась: %s",
                            attempt,
                            args.retries,
                            page_number,
                            error,
                        )
                        await asyncio.sleep(min(5 * attempt, 15))

                if not success:
                    await save_debug(page, page_number)
                    raise RuntimeError(
                        f"Страница {page_number} не загрузилась "
                        f"после {args.retries} попыток."
                    )

                if total_records is None:
                    total_records = await extract_total_records(page)

                    if total_records is not None:
                        total_pages = max(
                            1,
                            (total_records + args.records_per_page - 1)
                            // args.records_per_page,
                        )
                        logging.info(
                            "ЕИС показывает записей: %s; расчётных страниц: %s",
                            total_records,
                            total_pages,
                        )

                cards = await find_cards(page)
                card_count = await cards.count()

                if card_count == 0:
                    body_text = normalize_space(
                        await page.locator("body").inner_text()
                    )

                    if "Поиск не дал результатов" in body_text:
                        logging.info(
                            "ЕИС сообщает об отсутствии результатов "
                            "на странице %s.",
                            page_number,
                        )
                    else:
                        logging.warning(
                            "Карточки на странице %s не распознаны.",
                            page_number,
                        )
                        await save_debug(page, page_number)

                    empty_pages_in_row += 1

                    if empty_pages_in_row >= 2:
                        logging.info(
                            "Получены две пустые страницы подряд. "
                            "Сбор завершён."
                        )
                        break
                else:
                    empty_pages_in_row = 0
                    logging.info(
                        "На странице %s обнаружено карточек: %s",
                        page_number,
                        card_count,
                    )

                    page_items: list[Tender] = []

                    for index in range(card_count):
                        card = cards.nth(index)

                        try:
                            item = await parse_card(
                                card=card,
                                page_number=page_number,
                                query=query,
                                base_url=args.url,
                            )

                            if item.procurement_number or item.object_name:
                                page_items.append(item)
                            else:
                                logging.warning(
                                    "Карточка %s на странице %s "
                                    "не содержит номера или названия.",
                                    index + 1,
                                    page_number,
                                )

                        except Exception as error:
                            logging.exception(
                                "Ошибка разбора карточки %s "
                                "на странице %s: %s",
                                index + 1,
                                page_number,
                                error,
                            )

                    collected.extend(page_items)
                    collected = deduplicate(collected)
                    save_all(collected, args.output)

                    logging.info(
                        "Страница %s обработана. Всего уникальных карточек: %s",
                        page_number,
                        len(collected),
                    )

                if total_pages is not None and page_number >= total_pages:
                    logging.info(
                        "Достигнута последняя расчётная страница: %s",
                        total_pages,
                    )
                    break

                delay = random.uniform(args.delay_min, args.delay_max)
                logging.info("Пауза %.1f сек.", delay)
                await asyncio.sleep(delay)

        finally:
            save_all(collected, args.output)
            await context.close()
            await browser.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Сбор карточек закупок из результатов поиска ЕИС."
    )

    parser.add_argument(
        "--url",
        default=DEFAULT_URL,
        help="Полный URL выдачи ЕИС с установленными фильтрами.",
    )
    parser.add_argument(
        "--output",
        default="web_application",
        help="Базовое имя выходных файлов.",
    )
    parser.add_argument(
        "--start-page",
        type=int,
        default=1,
        help="Первая страница для обработки.",
    )
    parser.add_argument(
        "--max-pages",
        type=int,
        default=10,
        help="Максимальное число страниц.",
    )
    parser.add_argument(
        "--records-per-page",
        type=int,
        default=50,
        help="Количество карточек на странице.",
    )
    parser.add_argument(
        "--delay-min",
        type=float,
        default=2.5,
        help="Минимальная пауза между страницами.",
    )
    parser.add_argument(
        "--delay-max",
        type=float,
        default=4.5,
        help="Максимальная пауза между страницами.",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=60,
        help="Тайм-аут загрузки страницы в секундах.",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=3,
        help="Количество повторных попыток загрузки.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Запуск браузера без видимого окна.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Повторно обрабатывать страницы из checkpoint.",
    )

    return parser.parse_args()


def main() -> None:
    setup_logging()
    args = parse_args()

    try:
        asyncio.run(collect(args))
    except KeyboardInterrupt:
        logging.info("Остановлено пользователем. Уже собранные данные сохранены.")
    except Exception:
        logging.exception("Сбор завершился с ошибкой.")
        sys.exit(1)


if __name__ == "__main__":
    main()
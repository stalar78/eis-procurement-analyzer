from __future__ import annotations

import argparse
import asyncio
import csv
import html
import json
import logging
import mimetypes
import re
import shutil
import sys
import warnings
import zipfile
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urljoin, urlparse

from openpyxl import load_workbook
from playwright.async_api import async_playwright, Page, TimeoutError as PlaywrightTimeoutError

try:
    import requests
    from requests import Response as RequestsResponse
    from requests.exceptions import RequestException, SSLError as RequestsSSLError, Timeout as RequestsTimeout
    from urllib3.exceptions import InsecureRequestWarning
except ImportError:  # pragma: no cover
    requests = None
    RequestsResponse = Any
    RequestException = RequestsSSLError = RequestsTimeout = Exception
    InsecureRequestWarning = Warning


DOWNLOAD_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".zip", ".rar", ".7z", ".rtf", ".txt", ".xml", ".sig"
}
DOWNLOAD_CONTENT_TYPES = {
    "application/pdf": ".pdf",
    "application/msword": ".doc",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/vnd.ms-excel": ".xls",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "application/zip": ".zip",
    "application/x-zip-compressed": ".zip",
    "application/x-rar-compressed": ".rar",
    "application/vnd.rar": ".rar",
    "application/xml": ".xml",
    "text/xml": ".xml",
    "application/octet-stream": ".bin",
    "application/pkcs7-signature": ".sig",
}
HTML_CONTENT_TYPES = {"text/html", "application/xhtml+xml"}
RETRY_STATUSES = {429, 500, 502, 503, 504}
NAVIGATION_HINTS = (
    "/epz/contract/search/",
    "/epz/dizk/search/",
    "/epz/btk/search/",
    "/epz/main/public/document/",
    "/rpt/",
    "/analytics/",
    "/search/results.html",
)
IMPORTANT_LABELS = (
    "техническое задание", "описание объекта", "контракт", "проект контракта",
    "требования к содержанию", "обоснование", "протокол", "разъяснен",
    "участнику закупки", "документац", "извещение"
)


@dataclass
class Candidate:
    procurement_number: str
    object_name: str
    customer: str
    initial_price: str
    published_date: str
    card_url: str
    manual_rank: str = ""
    why_selected: str = ""


@dataclass
class PageRecord:
    procurement_number: str
    section: str
    url: str
    title: str
    text_file: str
    html_file: str
    links_found: int
    files_downloaded: int
    download_errors: int = 0
    error: str = ""


@dataclass
class DownloadLink:
    section: str
    label: str
    url: str
    source_page: str
    download_attr: str = ""


@dataclass
class DownloadResult:
    status: str
    saved_filename: str = ""
    final_url: str = ""
    content_type: str = ""
    http_status: int = 0
    file_size: int = 0
    error: str = ""


@dataclass
class DownloadManifestRow:
    procurement_number: str
    section: str
    source_page: str
    original_url: str
    final_url: str
    link_text: str
    saved_filename: str
    content_type: str
    http_status: int
    file_size: int
    status: str
    error: str


@dataclass
class CollectionResult:
    procurement_number: str
    output_dir: str
    status: str
    error: str = ""


try:
    from radar.artifact_registry import ArtifactRecord, fingerprint_records, safe_filename, sha256_file
    from radar.live_collection import (
        LiveCollectionResult,
        ProcurementCollectionTarget,
        canonical_url_for_number,
        deduplicate_document_links,
        normalize_eis_url,
        section_url,
    )
except Exception:  # pragma: no cover - collector CLI can run standalone
    ArtifactRecord = Any
    LiveCollectionResult = Any
    ProcurementCollectionTarget = Any
    canonical_url_for_number = None
    deduplicate_document_links = None
    fingerprint_records = None
    normalize_eis_url = None
    safe_filename = None
    section_url = None
    sha256_file = None


def safe_name(value: str, max_length: int = 120) -> str:
    value = fix_mojibake(html.unescape(unquote(value or "")))
    value = re.sub(r"[\\/:*?\"<>|]+", "_", value)
    value = re.sub(r"[\x00-\x1f]+", " ", value)
    value = re.sub(r"\s+", " ", value).strip(" ._")
    return (value or "unnamed")[:max_length]


def fix_mojibake(value: str) -> str:
    if not value or not any(marker in value for marker in ("Ð", "Ñ", "Â")):
        return value
    try:
        return value.encode("latin-1").decode("utf-8")
    except UnicodeError:
        return value


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for idx in range(2, 10_000):
        candidate = path.with_name(f"{stem}_{idx}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Не удалось подобрать уникальное имя файла для {path}")


def extension_from_content_type(content_type: str) -> str:
    normalized = content_type.split(";", 1)[0].strip().lower()
    if normalized in DOWNLOAD_CONTENT_TYPES:
        return DOWNLOAD_CONTENT_TYPES[normalized]
    return mimetypes.guess_extension(normalized) or ".bin"


def filename_from_content_disposition(value: str) -> str:
    if not value:
        return ""
    match = re.search(r"filename\*\s*=\s*([^']*)''([^;]+)", value, flags=re.IGNORECASE)
    if match:
        return unquote(match.group(2).strip().strip('"'))
    match = re.search(r'filename\s*=\s*"([^"]+)"', value, flags=re.IGNORECASE)
    if match:
        return match.group(1).strip()
    match = re.search(r"filename\s*=\s*([^;]+)", value, flags=re.IGNORECASE)
    return match.group(1).strip().strip('"') if match else ""


def filename_from_url(url: str) -> str:
    name = Path(unquote(urlparse(url).path)).name
    return name if Path(name).suffix else ""


def choose_filename(
    headers: dict[str, str],
    download_attr: str,
    label: str,
    final_url: str,
    content_type: str,
    index: int,
) -> str:
    candidates = [
        filename_from_content_disposition(headers.get("content-disposition", "")),
        download_attr,
        label,
        filename_from_url(final_url),
    ]
    name = next((safe_name(item, 150) for item in candidates if safe_name(item, 150)), f"file_{index:03d}")
    suffix = Path(name).suffix.lower()
    if suffix not in DOWNLOAD_EXTENSIONS:
        guessed = extension_from_content_type(content_type)
        name = f"{Path(name).stem if suffix else name}{guessed}"
    return safe_name(name, 170)


def looks_like_html(body: bytes, content_type: str) -> bool:
    normalized = content_type.split(";", 1)[0].strip().lower()
    if normalized in HTML_CONTENT_TYPES:
        return True
    head = body[:512].lstrip().lower()
    return head.startswith(b"<!doctype html") or head.startswith(b"<html") or b"<title>" in head[:256]


def is_probable_download(label: str, title: str, url: str, procurement_number: str) -> bool:
    parsed = urlparse(url)
    path = parsed.path.lower()
    combined = html.unescape(f"{label} {title} {url}").lower()
    ext = Path(path).suffix.lower()
    if any(hint in path for hint in NAVIGATION_HINTS):
        return False
    if "filestore" in path and ("download" in path or "file.html" in path):
        return True
    if "/download/" in path or "/download" in path:
        return True
    if ext in DOWNLOAD_EXTENSIONS and (procurement_number in url or "filestore" in path):
        return True
    return ext in DOWNLOAD_EXTENSIONS and any(x in combined for x in IMPORTANT_LABELS)


def should_retry(status: int, error: Exception | None = None) -> bool:
    if status in RETRY_STATUSES:
        return True
    return isinstance(error, (RequestException, RequestsTimeout, RequestsSSLError, TimeoutError, OSError))


def setup_logging(log_path: Path) -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)s | %(message)s",
        handlers=[logging.FileHandler(log_path, encoding="utf-8"), logging.StreamHandler(sys.stdout)],
    )


def load_candidates(path: Path, sheet_name: str, limit: int | None) -> list[Candidate]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Лист '{sheet_name}' не найден. Доступно: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]
    headers = {str(ws.cell(1, c).value): c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}
    required = ["procurement_number", "object_name", "customer", "initial_price", "published_date", "card_url"]
    missing = [x for x in required if x not in headers]
    if missing:
        raise ValueError(f"Нет обязательных столбцов: {', '.join(missing)}")

    result: list[Candidate] = []
    for r in range(2, ws.max_row + 1):
        number = str(ws.cell(r, headers["procurement_number"]).value or "").strip()
        url = str(ws.cell(r, headers["card_url"]).value or "").strip()
        if not number or not url:
            continue
        result.append(Candidate(
            procurement_number=number,
            object_name=str(ws.cell(r, headers["object_name"]).value or "").strip(),
            customer=str(ws.cell(r, headers["customer"]).value or "").strip(),
            initial_price=str(ws.cell(r, headers["initial_price"]).value or "").strip(),
            published_date=str(ws.cell(r, headers["published_date"]).value or "").strip(),
            card_url=url,
            manual_rank=str(ws.cell(r, headers.get("manual_rank", 1)).value or "").strip() if "manual_rank" in headers else "",
            why_selected=str(ws.cell(r, headers.get("why_selected", 1)).value or "").strip() if "why_selected" in headers else "",
        ))
        if limit and len(result) >= limit:
            break
    return result


async def settle(page: Page) -> None:
    try:
        await page.wait_for_load_state("domcontentloaded", timeout=60_000)
    except PlaywrightTimeoutError:
        pass
    try:
        await page.wait_for_load_state("networkidle", timeout=15_000)
    except PlaywrightTimeoutError:
        pass
    await page.wait_for_timeout(1500)


async def page_text(page: Page) -> str:
    try:
        return re.sub(r"\n{3,}", "\n\n", (await page.locator("body").inner_text()).strip())
    except Exception:
        return ""


def section_from_text(text: str, href: str) -> str:
    t = text.lower()
    h = href.lower()
    if "документ" in t or "documents" in h:
        return "documents"
    if "результат" in t or "поставщик" in t or "supplier" in h or "results" in h:
        return "results"
    if "контракт" in t or "contract" in h or "договор" in t:
        return "contracts"
    if "протокол" in t or "protocol" in h:
        return "protocols"
    if "журнал" in t or "event" in h:
        return "events"
    if "общая" in t or "common-info" in h:
        return "common"
    return "other"


async def discover_sections(page: Page, base_url: str, number: str) -> dict[str, str]:
    found: dict[str, str] = {"common": base_url}
    links = page.locator("a")
    for i in range(await links.count()):
        a = links.nth(i)
        try:
            href = (await a.get_attribute("href")) or ""
            text = ((await a.inner_text()) or "").strip()
        except Exception:
            continue
        if not href:
            continue
        absolute = urljoin(base_url, href)
        if number not in absolute and "regNumber=" not in absolute:
            continue
        section = section_from_text(text, absolute)
        if section != "other" and section not in found:
            found[section] = absolute
    return found


async def collect_download_links(page: Page, page_url: str, procurement_number: str, section: str) -> list[DownloadLink]:
    items: list[DownloadLink] = []
    seen: set[str] = set()
    links = page.locator("a")
    for i in range(await links.count()):
        a = links.nth(i)
        try:
            href = (await a.get_attribute("href")) or ""
            label = re.sub(r"\s+", " ", ((await a.inner_text()) or "").strip())
            title = (await a.get_attribute("title")) or ""
            download_attr = (await a.get_attribute("download")) or ""
        except Exception:
            continue
        href = html.unescape(href.strip())
        if not href or href.startswith("javascript:") or href.startswith("#") or href.lower().startswith(("mailto:", "tel:")):
            continue
        url = urljoin(page_url, href)
        if is_probable_download(label, title, url, procurement_number) and url not in seen:
            seen.add(url)
            parsed = urlparse(url)
            items.append(DownloadLink(section, label or title or Path(parsed.path).name or "file", url, page_url, download_attr))
    return items


async def playwright_download(page: Page, link: DownloadLink, timeout_ms: int) -> tuple[int, str, dict[str, str], bytes]:
    response = await page.context.request.get(link.url, headers={"Referer": link.source_page}, timeout=timeout_ms)
    return response.status, response.url, {k.lower(): v for k, v in response.headers.items()}, await response.body()


async def requests_download(page: Page, link: DownloadLink, timeout_seconds: float) -> tuple[int, str, dict[str, str], bytes]:
    if requests is None:
        raise RuntimeError("requests is not installed")
    cookies = await page.context.cookies(link.url)
    cookie_jar = {item["name"]: item["value"] for item in cookies}
    user_agent = await page.evaluate("() => navigator.userAgent")
    headers = {"User-Agent": user_agent, "Referer": link.source_page}

    def run_request() -> RequestsResponse:
        with requests.Session() as session:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", InsecureRequestWarning)
                return session.get(
                    link.url,
                    headers=headers,
                    cookies=cookie_jar,
                    timeout=timeout_seconds,
                    allow_redirects=True,
                    verify=False,
                )

    response = await asyncio.to_thread(run_request)
    return response.status_code, response.url, {k.lower(): v for k, v in response.headers.items()}, response.content


async def fetch_download(page: Page, link: DownloadLink, timeout_ms: int, retries: int) -> tuple[int, str, dict[str, str], bytes, str]:
    last_error = ""
    for attempt in range(1, retries + 1):
        try:
            status, final_url, headers, body = await playwright_download(page, link, timeout_ms)
            if status not in RETRY_STATUSES or attempt == retries:
                return status, final_url, headers, body, "playwright"
        except Exception as exc:
            last_error = str(exc)
            if "certificate" in last_error.lower() or "ssl" in last_error.lower():
                try:
                    status, final_url, headers, body = await requests_download(page, link, timeout_ms / 1000)
                    if status not in RETRY_STATUSES or attempt == retries:
                        return status, final_url, headers, body, "requests"
                except Exception as fallback_exc:
                    last_error = f"{last_error}; fallback requests: {fallback_exc}"
            elif attempt == retries or not should_retry(0, exc):
                raise
        await asyncio.sleep(0.8 * attempt)
    raise RuntimeError(last_error or "download failed")


async def download_one(
    page: Page,
    link: DownloadLink,
    target_dir: Path,
    debug_dir: Path,
    index: int,
    timeout_ms: int,
    retries: int,
) -> DownloadResult:
    try:
        status, final_url, headers, body, method = await fetch_download(page, link, timeout_ms, retries)
        content_type = headers.get("content-type", "")
        if status >= 400:
            return DownloadResult("failed", final_url=final_url, content_type=content_type, http_status=status, error=f"HTTP {status}")
        if not body:
            return DownloadResult("failed", final_url=final_url, content_type=content_type, http_status=status, error="empty response")
        if looks_like_html(body, content_type):
            debug_dir.mkdir(parents=True, exist_ok=True)
            debug_file = unique_path(debug_dir / f"{index:03d}_{safe_name(link.label, 80)}.html")
            debug_file.write_bytes(body)
            return DownloadResult(
                "html_rejected",
                final_url=final_url,
                content_type=content_type,
                http_status=status,
                file_size=len(body),
                error=f"HTML response saved to {debug_file.name}",
            )
        filename = choose_filename(headers, link.download_attr, link.label, final_url, content_type, index)
        target = unique_path(target_dir / filename)
        target.write_bytes(body)
        return DownloadResult("downloaded", target.name, final_url, content_type, status, len(body), f"method={method}")
    except Exception as exc:
        return DownloadResult("failed", final_url=link.url, error=str(exc))


async def save_section(
    page: Page,
    candidate: Candidate,
    section: str,
    url: str,
    folder: Path,
    do_download: bool,
    timeout_ms: int,
    retries: int,
) -> tuple[PageRecord, list[DownloadManifestRow]]:
    text_path = folder / f"{section}.txt"
    html_path = folder / f"{section}.html"
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=90_000)
        await settle(page)
        title = await page.title()
        text = await page_text(page)
        text_path.write_text(text, encoding="utf-8")
        html_path.write_text(await page.content(), encoding="utf-8")

        links = await collect_download_links(page, url, candidate.procurement_number, section)
        files_downloaded = 0
        download_errors = 0
        section_rows: list[dict[str, str]] = []
        manifest_rows: list[DownloadManifestRow] = []
        files_dir = folder / "downloads" / section
        debug_dir = folder / "debug" / section
        files_dir.mkdir(parents=True, exist_ok=True)
        for idx, link in enumerate(links, start=1):
            result = DownloadResult("not_downloaded", final_url=link.url)
            if do_download:
                result = await download_one(page, link, files_dir, debug_dir, idx, timeout_ms, retries)
                files_downloaded += int(result.status == "downloaded")
                download_errors += int(result.status not in {"downloaded", "not_downloaded"})
                if result.status == "downloaded":
                    logging.info("Downloaded %s: %s", candidate.procurement_number, result.saved_filename)
                else:
                    logging.warning("Download failed %s: %s (%s)", candidate.procurement_number, link.url, result.error)
            section_rows.append({"section": section, "label": link.label, "url": link.url, "status": result.status, "result": result.saved_filename or result.error})
            manifest_rows.append(DownloadManifestRow(
                procurement_number=candidate.procurement_number,
                section=section,
                source_page=link.source_page,
                original_url=link.url,
                final_url=result.final_url,
                link_text=link.label,
                saved_filename=result.saved_filename,
                content_type=result.content_type,
                http_status=result.http_status,
                file_size=result.file_size,
                status=result.status,
                error=result.error,
            ))

        with (folder / f"{section}_links.csv").open("w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.DictWriter(fh, fieldnames=["section", "label", "url", "status", "result"])
            writer.writeheader()
            writer.writerows(section_rows)

        return PageRecord(candidate.procurement_number, section, url, title, str(text_path), str(html_path), len(links), files_downloaded, download_errors), manifest_rows
    except Exception as exc:
        logging.exception("Ошибка раздела %s закупки %s", section, candidate.procurement_number)
        return PageRecord(candidate.procurement_number, section, url, "", str(text_path), str(html_path), 0, 0, 0, str(exc)), []


async def main_async(args: argparse.Namespace) -> None:
    output = Path(args.output).resolve()
    if args.overwrite and output.exists():
        shutil.rmtree(output)
    output.mkdir(parents=True, exist_ok=True)
    setup_logging(output / "collector.log")
    candidates = load_candidates(Path(args.input), args.sheet, args.limit)
    logging.info("Кандидатов: %d", len(candidates))

    all_records: list[PageRecord] = []
    all_download_rows: list[DownloadManifestRow] = []
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=args.headless)
        context = await browser.new_context(
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            viewport={"width": 1440, "height": 1050},
            ignore_https_errors=True,
        )
        page = await context.new_page()
        page.set_default_timeout(60_000)

        for pos, candidate in enumerate(candidates, start=1):
            folder = output / f"{pos:02d}_{candidate.procurement_number}"
            folder.mkdir(parents=True, exist_ok=True)
            (folder / "candidate.json").write_text(json.dumps(asdict(candidate), ensure_ascii=False, indent=2), encoding="utf-8")
            logging.info("[%d/%d] %s — %s", pos, len(candidates), candidate.procurement_number, candidate.object_name)

            try:
                await page.goto(candidate.card_url, wait_until="domcontentloaded", timeout=90_000)
                await settle(page)
                sections = await discover_sections(page, candidate.card_url, candidate.procurement_number)
            except Exception as exc:
                logging.exception("Не удалось открыть карточку %s", candidate.procurement_number)
                all_records.append(PageRecord(candidate.procurement_number, "common", candidate.card_url, "", "", "", 0, 0, 0, str(exc)))
                continue

            # Keep the useful sections only; unknown duplicates are ignored.
            order = ["common", "documents", "results", "protocols", "contracts", "events"]
            for section in order:
                url = sections.get(section)
                if not url:
                    continue
                record, download_rows = await save_section(
                    page,
                    candidate,
                    section,
                    url,
                    folder,
                    args.download,
                    args.download_timeout * 1000,
                    args.retries,
                )
                all_records.append(record)
                all_download_rows.extend(download_rows)
                await asyncio.sleep(args.delay)

            # Save discovered navigation for diagnostics.
            (folder / "sections.json").write_text(json.dumps(sections, ensure_ascii=False, indent=2), encoding="utf-8")
            await asyncio.sleep(args.delay)

        await context.close()
        await browser.close()

    with (output / "summary.csv").open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(PageRecord.__dataclass_fields__.keys()))
        writer.writeheader()
        writer.writerows(asdict(r) for r in all_records)
    (output / "summary.json").write_text(json.dumps([asdict(r) for r in all_records], ensure_ascii=False, indent=2), encoding="utf-8")

    with (output / "download_manifest.csv").open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(DownloadManifestRow.__dataclass_fields__.keys()))
        writer.writeheader()
        writer.writerows(asdict(r) for r in all_download_rows)

    if args.zip:
        zip_path = output.with_suffix(".zip")
        if zip_path.exists():
            zip_path.unlink()
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for f in output.rglob("*"):
                if f.is_file():
                    zf.write(f, f.relative_to(output.parent))
        logging.info("ZIP создан: %s", zip_path)


async def _collect_direct_targets_async(
    targets: list[Any],
    output_dir: Path,
    *,
    overwrite: bool,
    refresh: bool,
    max_documents_per_procurement: int | None,
    max_total_download_bytes: int | None,
    max_single_file_bytes: int | None,
    timeout_seconds: int | None,
    verbose: bool,
) -> list[Any]:
    if normalize_eis_url is None:
        raise RuntimeError("radar live collection helpers are unavailable")
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[Any] = []
    total_bytes = 0
    allowed_sections = ["common", "documents", "results", "events", "protocols", "contracts"]
    required_sections = {"common", "documents"}
    started_batch = datetime.now().astimezone()
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            locale="ru-RU",
            timezone_id="Europe/Moscow",
            viewport={"width": 1440, "height": 1050},
            ignore_https_errors=True,
        )
        page = await context.new_page()
        page.set_default_timeout((timeout_seconds or 90) * 1000)
        for target in targets:
            started = datetime.now().astimezone()
            number = getattr(target, "procurement_number", "") if not isinstance(target, str) else target
            source_url = getattr(target, "source_url", "") if not isinstance(target, str) else ""
            errors: list[str] = []
            warnings: list[str] = []
            sections_visited: list[str] = []
            links_found: list[dict[str, Any]] = []
            docs_attempted = 0
            docs_downloaded = 0
            docs_cached = 0
            docs_failed = 0
            bytes_downloaded = 0
            proc_dir = output_dir / "procurements" / safe_name(number)
            manifest_path = proc_dir / "manifests" / "live_collection_manifest.json"
            try:
                if source_url:
                    source_url, resolved_number = normalize_eis_url(source_url, number or None)
                    number = resolved_number
                else:
                    if canonical_url_for_number is None:
                        raise ValueError("cannot resolve procurement URL")
                    source_url = canonical_url_for_number(number)
                    warnings.append("source URL resolved from procurement number")
                proc_dir = output_dir / "procurements" / safe_name(number)
                for sub in ("pages", "links", "documents", "downloads", "extracted", "analysis", "manifests", "debug"):
                    (proc_dir / sub).mkdir(parents=True, exist_ok=True)
                candidate = Candidate(number, "", "", "", "", source_url)
                (proc_dir / "candidate.json").write_text(json.dumps(asdict(candidate), ensure_ascii=False, indent=2), encoding="utf-8")
                section_urls = {section: section_url(source_url, section) for section in allowed_sections}
                discovered: dict[str, str] = {}
                for section, url in section_urls.items():
                    try:
                        await page.goto(url, wait_until="domcontentloaded", timeout=(timeout_seconds or 90) * 1000)
                        await settle(page)
                        html_text = await page.content()
                        text = await page_text(page)
                        (proc_dir / "pages" / f"{section}.html").write_text(html_text, encoding="utf-8")
                        (proc_dir / f"{section}.txt").write_text(text, encoding="utf-8")
                        sections_visited.append(section)
                        if section == "common":
                            discovered = await discover_sections(page, url, number)
                        section_links = await collect_download_links(page, url, number, section)
                        for link in section_links:
                            links_found.append(
                                {
                                    "section": section,
                                    "label": link.label,
                                    "url": link.url,
                                    "source_page": link.source_page,
                                    "download_attr": link.download_attr,
                                }
                            )
                    except Exception as exc:
                        message = f"{section}: {exc}"
                        if section in required_sections:
                            errors.append(message)
                        else:
                            warnings.append(message)
                for section, url in discovered.items():
                    if section in section_urls:
                        section_urls[section] = url
                unique_links = deduplicate_document_links(links_found)
                (proc_dir / "links" / "document_links.json").write_text(json.dumps(unique_links, ensure_ascii=False, indent=2), encoding="utf-8")
                artifact_records: list[Any] = []
                for index, row in enumerate(unique_links, start=1):
                    if max_documents_per_procurement and docs_attempted >= max_documents_per_procurement:
                        warnings.append("TOO_MANY_DOCUMENTS")
                        break
                    if max_total_download_bytes and total_bytes >= max_total_download_bytes:
                        warnings.append("DOWNLOAD_LIMIT_REACHED")
                        break
                    docs_attempted += 1
                    link = DownloadLink(row["section"], row.get("label", ""), row["url"], row.get("source_page", source_url), row.get("download_attr", ""))
                    try:
                        status, final_url, headers, body, method = await fetch_download(page, link, (timeout_seconds or 90) * 1000, 2)
                        content_type = headers.get("content-type", "")
                        if status >= 400:
                            docs_failed += 1
                            errors.append(f"{row['url']}: HTTP {status}")
                            continue
                        if not body:
                            docs_failed += 1
                            errors.append(f"{row['url']}: empty response")
                            continue
                        if max_single_file_bytes and len(body) > max_single_file_bytes:
                            docs_failed += 1
                            errors.append(f"{row['url']}: FILE_TOO_LARGE")
                            continue
                        if looks_like_html(body, content_type):
                            docs_failed += 1
                            errors.append(f"{row['url']}: HTML_INSTEAD_OF_DOCUMENT")
                            continue
                        filename = choose_filename(headers, link.download_attr, link.label, final_url, content_type, index)
                        safe = safe_filename(filename)
                        target_dir = proc_dir / "documents" / row["section"]
                        compat_dir = proc_dir / "downloads" / row["section"]
                        target_dir.mkdir(parents=True, exist_ok=True)
                        compat_dir.mkdir(parents=True, exist_ok=True)
                        target = target_dir / safe
                        if target.exists() and not (overwrite or refresh):
                            docs_cached += 1
                        else:
                            tmp = target.with_suffix(target.suffix + ".tmp")
                            tmp.write_bytes(body)
                            tmp.replace(target)
                            docs_downloaded += 1
                            bytes_downloaded += target.stat().st_size
                            total_bytes += target.stat().st_size
                        compat = compat_dir / target.name
                        if not compat.exists():
                            shutil.copy2(target, compat)
                        artifact_records.append(
                            ArtifactRecord(
                                procurement_number=number,
                                artifact_type="document",
                                source_url=row["url"],
                                local_path=str(target),
                                original_filename=filename,
                                content_type=content_type,
                                size_bytes=target.stat().st_size,
                                sha256=sha256_file(target),
                                downloaded_at=datetime.now().astimezone().isoformat(timespec="seconds"),
                                extraction_status="downloaded",
                                document_type="",
                                document_confidence="",
                            )
                        )
                    except Exception as exc:
                        docs_failed += 1
                        errors.append(f"{row['url']}: {exc}")
                doc_fingerprint = fingerprint_records(artifact_records)
                manifest = {
                    "procurement_number": number,
                    "source_url": source_url,
                    "section_urls": section_urls,
                    "sections_visited": sections_visited,
                    "links": unique_links,
                    "artifacts": [asdict(item) for item in artifact_records],
                    "document_set_fingerprint": doc_fingerprint,
                    "errors": errors,
                    "warnings": warnings,
                }
                manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
                missing_required = sorted(required_sections - set(sections_visited))
                if missing_required or not artifact_records:
                    status = "PARTIAL" if artifact_records else "FAILED_RETRYABLE"
                elif errors:
                    status = "PARTIAL"
                else:
                    status = "COMPLETE"
                finished = datetime.now().astimezone()
                results.append(
                    LiveCollectionResult(
                        procurement_number=number,
                        source_url=source_url,
                        procurement_directory=str(proc_dir),
                        status=status,
                        pages_visited=len(sections_visited),
                        sections_visited=sections_visited,
                        document_links_found=len(unique_links),
                        documents_attempted=docs_attempted,
                        documents_downloaded=docs_downloaded,
                        documents_skipped_cached=docs_cached,
                        documents_failed=docs_failed,
                        total_downloaded_bytes=bytes_downloaded,
                        manifest_path=str(manifest_path),
                        errors=errors,
                        warnings=warnings,
                        started_at=started.isoformat(timespec="seconds"),
                        finished_at=finished.isoformat(timespec="seconds"),
                        resolved_common_url=source_url,
                        document_set_fingerprint=doc_fingerprint,
                        duration_seconds=(finished - started).total_seconds(),
                    )
                )
            except Exception as exc:
                finished = datetime.now().astimezone()
                results.append(
                    LiveCollectionResult(
                        procurement_number=number,
                        source_url=source_url,
                        procurement_directory=str(proc_dir),
                        status="FAILED_FINAL",
                        errors=[str(exc)],
                        started_at=started.isoformat(timespec="seconds"),
                        finished_at=finished.isoformat(timespec="seconds"),
                        duration_seconds=(finished - started).total_seconds(),
                    )
                )
        await context.close()
        await browser.close()
    return results


def collect_candidate_details_for_procurements(
    procurements: list[Any],
    output_dir: Path,
    *,
    overwrite: bool = False,
    refresh: bool = False,
    max_documents_per_procurement: int | None = None,
    max_total_download_bytes: int | None = None,
    max_single_file_bytes: int | None = None,
    timeout_seconds: int | None = None,
    verbose: bool = False,
) -> list[Any]:
    """Collect live procurement details from direct targets without a queue file."""
    targets: list[Any] = []
    for item in procurements:
        if isinstance(item, str):
            targets.append(ProcurementCollectionTarget(item, ""))
        else:
            targets.append(item)
    return asyncio.run(
        _collect_direct_targets_async(
            targets,
            Path(output_dir),
            overwrite=overwrite,
            refresh=refresh,
            max_documents_per_procurement=max_documents_per_procurement,
            max_total_download_bytes=max_total_download_bytes,
            max_single_file_bytes=max_single_file_bytes,
            timeout_seconds=timeout_seconds,
            verbose=verbose,
        )
    )


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Сбор деталей и документов по листу 'Топ-кандидаты' ЕИС.")
    ap.add_argument("--input", default="all_web_tenders_classified.xlsx", help="Путь к XLSX")
    ap.add_argument("--sheet", default="Топ-кандидаты", help="Имя листа")
    ap.add_argument("--output", default="candidate_details", help="Папка результата")
    ap.add_argument("--limit", type=int, default=None, help="Ограничить число кандидатов")
    ap.add_argument("--delay", type=float, default=2.5, help="Пауза между страницами")
    ap.add_argument("--download", action="store_true", help="Пробовать скачивать найденные вложения")
    ap.add_argument("--download-timeout", type=int, default=90, help="Timeout for one file download, seconds")
    ap.add_argument("--retries", type=int, default=3, help="Download retries for network errors and HTTP 429/5xx")
    ap.add_argument("--overwrite", action="store_true", help="Remove output folder before run")
    ap.add_argument("--headless", action="store_true", help="Не показывать окно браузера")
    ap.add_argument("--zip", action="store_true", help="Упаковать результат в ZIP")
    return ap.parse_args()


if __name__ == "__main__":
    args = parse_args()
    try:
        asyncio.run(main_async(args))
    except KeyboardInterrupt:
        print("Остановлено пользователем. Уже сохраненные папки останутся на диске.")
